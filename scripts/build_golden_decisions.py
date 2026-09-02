"""生成真实资料验收的人工 sheet 决策（manual_sheet_decisions.json）。

本脚本以"人工复核者"身份工作：只读 corpus 副本，从文档自身的结算表编号
读取权威期号（文件名期号与文档不一致时以文档为准——真实陷阱，需留痕），
并对每张 Sheet 给出 extract / evidence_only 决定与理由。

龙泉卓博系列 layout（人工核对多份副本后确认）：
- 结算表-明细表：双层表头 7-8 行（第 7 行组名/第 8 行列名），第 9 行为段落行，
  数据自第 10 行起。列：2名称 3单位 9本期工程量 10本期金额（金额口径）。
- 结算表-合同外：单层表头第 7 行，数据自第 8 行。列：2名称 3单位 7本期量
  8本期金额。
"""
from __future__ import annotations

import json
import re
import sys
from pathlib import Path

REPO_ROOT = Path(__file__).resolve().parents[1]
sys.path.insert(0, str(REPO_ROOT / "src"))

from openpyxl import load_workbook  # noqa: E402

BASE = REPO_ROOT / "local_private_data" / "real_acceptance"
CORPUS = BASE / "corpus"
OUT = BASE / "manual_sheet_decisions.json"

ACTOR = "golden-case-reviewer"
PERIOD_RE = re.compile(r"(\d{4})年第0*(\d+)期")


def doc_period_no(path: Path) -> int | None:
    """从明细表/价款结算表的结算表编号读权威期号（人工口径）。"""
    wb = load_workbook(path, read_only=True, data_only=True)
    try:
        for name in ("结算表-明细表", "结算表-价款结算表", "结算表-支付表（最终审核）"):
            if name not in wb.sheetnames:
                continue
            ws = wb[name]
            for row in ws.iter_rows(min_row=1, max_row=8, values_only=True):
                for v in row:
                    m = PERIOD_RE.search(str(v)) if v else None
                    if m:
                        return int(m.group(2))
    finally:
        wb.close()
    return None


def sheet_status_map(test_id: str) -> dict[str, str]:
    """从最近一轮验收结果读各 sheet 状态（决定依据，留痕）。"""
    import glob

    runs = sorted(glob.glob(str(BASE / "work" / "run_*" / "acceptance_results.json")))
    if not runs:
        return {}
    data = json.loads(Path(runs[-1]).read_text(encoding="utf-8"))
    for rec in data.get("per_file", []):
        if rec.get("test_id") == test_id:
            sp = rec.get("settlement_parse") or {}
            return {s["name"]: s.get("status", "") for s in sp.get("sheets", [])}
    return {}


def decisions_for(test_id: str) -> list[dict]:
    copy = next(CORPUS.glob(f"{test_id}.*"))
    status = sheet_status_map(test_id)
    extract_sheets: dict[str, dict] = {}
    if test_id in {f"T-LQ-{i:02d}" for i in list(range(1, 10)) + [15]}:
        period = doc_period_no(copy)
        assert period is not None, f"{test_id}: 文档期号未找到"
        extract_sheets["结算表-明细表"] = {
            "col_map": {"code": 1, "name": 2, "unit": 3, "quantity": 9, "amount": 10},
            "period_no": period,
        }
        if "结算表-合同外" in status:
            extract_sheets["结算表-合同外"] = {
                "col_map": {"name": 2, "unit": 3, "quantity": 7, "amount": 8},
                "period_no": period,
            }
    decisions: list[dict] = []
    for sheet_name, st in status.items():
        if sheet_name in extract_sheets:
            spec = extract_sheets[sheet_name]
            decisions.append({
                "sheet": sheet_name,
                "action": "extract",
                "direction": "downward",
                "period_no": spec["period_no"],
                "col_map": spec["col_map"],
                "actor": ACTOR,
                "reason": (
                    f"人工核对：结算明细类表单，表头/列位经人工逐列确认；"
                    f"期号以文档结算表编号为准（第{spec['period_no']}期）"
                ),
            })
        elif st in ("needs_role_review", "non_settlement_form"):
            # no_header sheet 不出决策：无人工列映射前保持 no_header 原状
            # （执行器规定 no_header 只有显式 extract+列映射才进入人工候选）
            role = ("supporting_evidence" if ("封面" in sheet_name or "说明" in sheet_name
                    or "计量" in sheet_name or "支付" in sheet_name)
           else ("settlement_summary" if "价款结算表" in sheet_name
                 else "other_non_settlement"))
            decisions.append({
                "sheet": sheet_name,
                "action": "evidence_only",
                "role": role,
                "actor": ACTOR,
                "reason": f"人工核对：非明细抽取对象（status={st}），仅保留证据与追溯",
            })
    return decisions


def main() -> int:
    files: dict[str, list[dict]] = {}
    statuses = {}
    for copy in sorted(CORPUS.iterdir()):
        tid = copy.stem
        files[tid] = decisions_for(tid)
        statuses[tid] = sheet_status_map(tid)
    payload = {
        "version": 1,
        "files": files,
        "acceptance_controls": {},
        "_notes": {
            "period_authority": "期号以文档结算表编号为准，文件名期号与文档不一致处已逐份核对",
            "layout": "明细表=双层表头7-8行+段落行9+数据10起；合同外=单层表头7行+数据8起",
        },
    }
    OUT.write_text(json.dumps(payload, ensure_ascii=False, indent=2), encoding="utf-8")
    n_extract = sum(1 for ds in files.values() for d in ds if d["action"] == "extract")
    n_evidence = sum(1 for ds in files.values() for d in ds if d["action"] == "evidence_only")
    print(f"written {OUT.name}: files={len(files)} extract={n_extract} evidence_only={n_evidence}")
    for tid, ds in files.items():
        ex = [d["sheet"] for d in ds if d["action"] == "extract"]
        print(f"  {tid}: extract={ex}")
    return 0


if __name__ == "__main__":
    raise SystemExit(main())

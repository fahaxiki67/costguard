"""Windows Excel/WPS 真实 COM 实测（P2 验收工具，只读原始资料纪律下运行）。

子命令：
  com_modes      四模式 COM 验证：强制 Excel / 强制 WPS(Ket) / 回退链 / 无 COM 降级；
                 同时验证：附着不误杀用户已开实例、工作簿只读关闭、无进程残留。
  roundtrip      导出→COM 打开→重算→保存→复开→与程序结果逐格核对；
                 含审核底稿 G==H（公式复核=程序值）与差异列自洽检查。

用法：
  uv run python scripts/com_roundtrip_check.py com_modes [--workdir DIR]
  uv run python scripts/com_roundtrip_check.py roundtrip [--workdir DIR]

只使用 examples/demo 合成数据；真实资料路径仅作为跳转目标（只读打开）。
结果打印 JSON 到 stdout，供发布清单归档。
"""
from __future__ import annotations

import argparse
import json
import shutil
import subprocess
import sys
import time
from pathlib import Path

REPO_ROOT = Path(__file__).resolve().parents[1]
sys.path.insert(0, str(REPO_ROOT / "src"))

_SPREADSHEET_PROCS = ("EXCEL.EXE", "et.exe", "wps.exe", "ket.exe", "et.exe.old")


def list_spreadsheet_processes() -> dict[str, list[int]]:
    out_map: dict[str, list[int]] = {}
    for name in sorted({p.upper() for p in _SPREADSHEET_PROCS}):
        out = subprocess.run(
            ["tasklist", "/FI", f"IMAGENAME eq {name}", "/FO", "CSV", "/NH"],
            capture_output=True, text=True, check=False, encoding="utf-8",
            errors="replace")
        pids: list[int] = []
        for line in out.stdout.splitlines():
            parts = [p.strip('"') for p in line.split('","')]
            if len(parts) >= 2 and parts[0].lower() == name.lower():
                try:
                    pids.append(int(parts[1]))
                except ValueError:
                    continue
        out_map[name] = pids
    return out_map


def wait_processes_gone(before_pids: dict[str, list[int]], timeout: float = 10.0) -> dict[str, list[int]]:
    """等待 before 里不存在的“新增进程”全部退出；返回超时后仍存活的进程。

    WPS(et.exe/wps.exe) COM Server 完全退出有数十秒延迟；调用方对
    WPS 场景应给足 timeout（脚本统一用 45s），否则会把正常退出延迟
    误判成进程泄漏。
    """
    deadline = time.time() + timeout
    while True:
        current = list_spreadsheet_processes()
        leftover = {
            name: [p for p in pids if p not in set(before_pids.get(name, []))]
            for name, pids in current.items() if pids
        }
        if not any(leftover.values()) or time.time() >= deadline:
            return leftover
        time.sleep(1.0)


def _com_client():
    import pythoncom
    import win32com.client
    return pythoncom, win32com.client


def open_isolated_app(progid: str, *, visible: bool = False):
    """启动独立表格程序实例（DispatchEx），调用方负责 Quit。"""
    pythoncom, win32com = _com_client()
    pythoncom.CoInitialize()
    app = win32com.DispatchEx(progid)
    try:
        app.Visible = visible
        app.DisplayAlerts = False
    except Exception:  # noqa: BLE001  WPS 个别属性缺失可容忍
        pass
    return app, pythoncom


def quit_app(app, pythoncom) -> None:
    try:
        app.Quit()
    except Exception:  # noqa: BLE001
        pass
    app = None
    try:
        pythoncom.CoUninitialize()
    except Exception:  # noqa: BLE001
        pass


# ---------------------------------------------------------------------------
# com_modes：四模式 COM 验证
# ---------------------------------------------------------------------------

def _make_probe_workbook(workdir: Path) -> tuple[Path, Path]:
    from openpyxl import Workbook

    path = workdir / "com_probe_工程量台账.xlsx"
    wb = Workbook()
    ws = wb.active
    ws.title = "第一期结算"
    ws.append(["清单编码", "名称", "数量", "单价", "合价"])
    ws.append(["0101", "挖土方", 100, 25.5, "=C2*D2"])
    ws.append(["0102", "回填土", 200, 12.3, "=C3*D3"])
    wb.save(path)
    # 守护工作簿：与跳转目标不同文件，避免跳转关闭自己打开的工作簿时
    # 连带清空守护会话（同文件会被识别为同一工作簿）。
    guard_path = workdir / "com_guard_守护台账.xlsx"
    wb2 = Workbook()
    ws2 = wb2.active
    ws2.title = "守护Sheet"
    ws2.append(["A", "B"])
    ws2.append(["守护", 1])
    wb2.save(guard_path)
    return path, guard_path


def run_jump_case(mode: str, progids_json: str, workdir_json: str) -> dict:
    """单案例执行体（由父进程以子进程方式调用，避免 COM 卡死拖垮整体）。"""
    from jiadun.platform import spreadsheet_jump as sj

    progids = tuple(progids_json) if progids_json else None
    workdir = Path(workdir_json)
    workdir.mkdir(parents=True, exist_ok=True)
    probe, guard_file = _make_probe_workbook(workdir)
    before = list_spreadsheet_processes()
    # 预先开一个用户实例并持有不同工作簿，验证跳转不得误杀
    guard_app, guard_pyt, guard_progid = None, None, None
    for pid_ in ("Excel.Application", "Ket.Application", "ET.Application"):
        try:
            guard_app, guard_pyt = open_isolated_app(pid_, visible=False)
            guard_progid = pid_
            break
        except Exception:  # noqa: BLE001
            continue
    guard_alive_before = None
    guard_book = None
    if guard_app is not None:
        try:
            guard_book = guard_app.Workbooks.Open(str(guard_file), 0, True)
            guard_alive_before = guard_book.FullName
        except Exception:  # noqa: BLE001
            guard_book = None
    try:
        target = sj.JumpTarget(
            file_path=probe, sheet_name="第一期结算", row=2, col=5)
        outcome = sj.open_in_spreadsheet(target, progids=progids)
    finally:
        guard_alive_after = None
        if guard_app is not None:
            try:
                guard_alive_after = guard_app.Workbooks.Count >= 1
                if guard_book is not None:
                    guard_book.Close(False)
            except Exception:  # noqa: BLE001
                guard_alive_after = None
            quit_app(guard_app, guard_pyt)
    leftover = wait_processes_gone(before, timeout=45.0)
    late_cleared = False
    if any(leftover.values()):
        # WPS 空闲会话回收可能超过 45s：150s 内清掉视为通过（如实记录）
        leftover = wait_processes_gone(
            {**before, **{k: set() for k in leftover}}, timeout=105.0)
        late_cleared = not any(leftover.values())
    return {
        "mode": mode,
        "progids": list(progids) if progids is not None else "default",
        "outcome": outcome,
        "guard_app": guard_progid,
        "guard_workbook_survived": (
            bool(guard_alive_before) and bool(guard_alive_after)
            if guard_app is not None else None),
        "leftover_processes": {k: v for k, v in leftover.items() if v},
        "late_cleared": late_cleared,
    }


def com_modes(workdir: Path) -> dict:
    workdir.mkdir(parents=True, exist_ok=True)
    results: dict = {"workdir": str(workdir), "modes": []}
    cases = [
        ("excel_only", ("Excel.Application",)),
        ("wps_only", ("Ket.Application",)),
        ("fallback_chain", ("No.Such.Application", "Excel.Application")),
    ]
    for mode, progids in cases:
        # 每个案例独立子进程 + 超时：外部表格程序卡死时能杀掉并如实记录
        proc = subprocess.Popen(
            [sys.executable, str(Path(__file__).resolve()), "_case",
             "--mode", mode, "--progids", json.dumps(progids),
             "--workdir", str(workdir / mode)],
            stdout=subprocess.PIPE, stderr=subprocess.PIPE, text=True,
            encoding="utf-8", errors="replace")
        try:
            out, err = proc.communicate(timeout=210)
            result = json.loads(out.strip().splitlines()[-1])
        except subprocess.TimeoutExpired:
            proc.kill()
            proc.wait(timeout=15)
            result = {"mode": mode, "outcome": "case_timeout",
                      "note": "COM 操作超过 210s 被强制终止（外部程序无响应）"}
        except Exception as exc:  # noqa: BLE001
            result = {"mode": mode, "outcome": "case_error",
                      "note": f"{type(exc).__name__}: {exc} {err[-200:] if err else ''}"}
        results["modes"].append(result)

    # 模式4：无任何 COM（对应“两者都没有”）→ 降级 opener；用哑 opener 避免真开
    workdir4 = workdir / "no_com"
    workdir4.mkdir(parents=True, exist_ok=True)
    probe, _ = _make_probe_workbook(workdir4)
    before = list_spreadsheet_processes()
    opened: list[Path] = []
    target = sj_target(probe)
    outcome = sj_open(target, opener=opened.append)
    leftover = wait_processes_gone(before, timeout=45.0)
    results["modes"].append({
        "mode": "no_com_fallback",
        "progids": [],
        "outcome": outcome,
        "opened": [str(opened[0])] if opened else [],
        "leftover_processes": {k: v for k, v in leftover.items() if v},
    })
    ok = all(
        (m["outcome"] == "located" and m["guard_workbook_survived"] is not False
         and not m.get("leftover_processes"))
        for m in results["modes"][:3])
    ok = ok and results["modes"][3]["outcome"] == "opened_only"
    results["verdict"] = "PASS" if ok else "FAIL"
    return results


def sj_target(probe: Path):
    from jiadun.platform import spreadsheet_jump as sj

    return sj.JumpTarget(file_path=probe, sheet_name="第一期结算", row=2, col=5)


def sj_open(target, *, opener):
    from jiadun.platform import spreadsheet_jump as sj

    return sj.open_in_spreadsheet(target, progids=(), opener=opener, platform="nt")


# ---------------------------------------------------------------------------
# roundtrip：导出→COM 打开→重算→保存→复开→核对
# ---------------------------------------------------------------------------

def _build_project(workdir: Path):
    from jiadun.core import demo as demo_core
    from jiadun.core.engine import aggregate as agg_mod
    from jiadun.core.models import project as pm

    ws = workdir / "JiadunProjects"
    info = demo_core.provision_demo_project(ws)
    info, conn = pm.open_project(Path(info.workspace_path))
    aggs = agg_mod.aggregate_project(conn, info.project_id, include_all_directions=True)
    agg_mod.persist_period_totals(conn, info.project_id, aggs)
    return info, conn


def _sheet_values(path: Path, data_only: bool) -> dict[str, dict[tuple[int, int], object]]:
    from openpyxl import load_workbook

    wb = load_workbook(path, data_only=data_only, read_only=True)
    out: dict[str, dict[tuple[int, int], object]] = {}
    try:
        for ws in wb.worksheets:
            grid: dict[tuple[int, int], object] = {}
            for row in ws.iter_rows():
                for cell in row:
                    if cell.value is not None:
                        grid[(cell.row, cell.column)] = cell.value
            out[ws.title] = grid
    finally:
        wb.close()
    return out


def _values_equal(a, b) -> bool:
    if a is None or b is None:
        return a is None and b is None
    if isinstance(a, (int, float)) and isinstance(b, (int, float)):
        return abs(float(a) - float(b)) < 1e-6
    return str(a) == str(b)


def roundtrip(workdir: Path, progid: str = "Excel.Application") -> dict:
    from jiadun.core.export import excel_export

    workdir.mkdir(parents=True, exist_ok=True)
    info, conn = _build_project(workdir)
    try:
        exports = Path(info.workspace_path) / "exports"
        xlsx = excel_export.export_workbook(conn, info.project_id, exports)
    finally:
        conn.close()
    original = _sheet_values(xlsx, data_only=False)

    copy_path = workdir / "roundtrip_recalc.xlsx"
    shutil.copy2(xlsx, copy_path)
    before = list_spreadsheet_processes()
    app, pythoncom = open_isolated_app(progid, visible=False)
    recalc_error: str | None = None
    try:
        book = app.Workbooks.Open(str(copy_path))
        try:
            if hasattr(app, "CalculateFullRebuild"):
                app.CalculateFullRebuild()
            elif hasattr(app, "CalculateFull"):
                app.CalculateFull()
            else:
                app.Calculate()
        except Exception as exc:  # noqa: BLE001
            recalc_error = f"calculate: {exc}"
        book.Save()
        book.Close(False)
        book = None
    except Exception as exc:  # noqa: BLE001
        recalc_error = f"open/save: {exc}"
    finally:
        quit_app(app, pythoncom)
    leftover = wait_processes_gone(before, timeout=45.0)

    report: dict = {
        "progid": progid,
        "export_file": str(xlsx),
        "recalc_copy": str(copy_path),
        "recalc_error": recalc_error,
        "leftover_processes": {k: v for k, v in leftover.items() if v},
    }
    if recalc_error:
        report["verdict"] = "FAIL"
        return report

    recalced = _sheet_values(copy_path, data_only=True)
    drift: list[str] = []
    for sheet, grid in original.items():
        rgrid = recalced.get(sheet, {})
        for coord, val in grid.items():
            rval = rgrid.get(coord)
            if isinstance(val, str) and val.startswith("="):
                # 公式：重算后必须有缓存值且不是错误值
                if rval is None or (isinstance(rval, str) and rval.startswith("#")):
                    drift.append(f"{sheet}!{coord} 公式无有效计算值: {rval!r}")
            elif not _values_equal(val, rval):
                drift.append(f"{sheet}!{coord} {val!r} -> {rval!r}")
    report["cell_drift_count"] = len(drift)
    report["cell_drift_samples"] = drift[:20]

    # 审核底稿语义核对：G(公式合价) == H(程序值)；J(差异) == ROUND(G-I,2)
    audit = recalced.get("审核底稿", {})
    bad_g, bad_j, n_rows = [], [], 0
    for (r, c), val in audit.items():
        if c == 8 and isinstance(val, (int, float)):  # H 列程序计算合价
            n_rows += 1
            g = audit.get((r, 7))
            i = audit.get((r, 9))
            j = audit.get((r, 10))
            if not isinstance(g, (int, float)) or abs(float(g) - float(val)) > 0.005:
                bad_g.append(f"行{r}: G={g!r} != H={val!r}")
            if isinstance(i, (int, float)) and not (
                    isinstance(j, (int, float))
                    and abs(float(j) - round(float(g) - float(i), 2)) <= 0.011):
                bad_j.append(f"行{r}: J={j!r} != ROUND(G-I,2)")
    report["audit_rows_checked"] = n_rows
    report["audit_formula_mismatch"] = bad_g[:10]
    report["audit_diff_mismatch"] = bad_j[:10]
    leak = bool(report["leftover_processes"])
    report["verdict"] = (
        "PASS" if not drift and not bad_g and not bad_j and not leak and n_rows > 0
        else "FAIL")
    return report


def main() -> int:
    parser = argparse.ArgumentParser(description=__doc__)
    parser.add_argument("command", choices=["com_modes", "roundtrip", "_case"])
    parser.add_argument("--workdir", default=None)
    parser.add_argument("--progid", default="Excel.Application")
    parser.add_argument("--mode", default=None)
    parser.add_argument("--progids", default="[]")
    args = parser.parse_args()
    workdir = Path(args.workdir) if args.workdir else (
        Path(__file__).resolve().parent.parent / "cg_runs" / "com_check")
    if args.command == "_case":
        result = run_jump_case(args.mode, json.loads(args.progids), str(workdir))
        print(json.dumps(result, ensure_ascii=False))
        return 0
    if args.command == "com_modes":
        report = com_modes(workdir / "modes")
    else:
        report = roundtrip(workdir / "roundtrip", progid=args.progid)
    print(json.dumps(report, ensure_ascii=False, indent=2))
    return 0 if report.get("verdict") == "PASS" else 1


if __name__ == "__main__":
    raise SystemExit(main())

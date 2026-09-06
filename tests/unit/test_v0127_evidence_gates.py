"""v0.1.27 基线核查修复的门控回归测试（fix/v0127-evidence-gates）。

覆盖四类问题，全部先红后绿：
- 聚合键含归一单位/特征：同名变体合并、混合单位绝不跨单位求和；
- fuzzy_name 合并单位复检：名称相似 ≥97 但单位不同 → 不可比禁合并（C3）；
  85–97 相似组给出"疑似相关"提示（C3b）；
- 镜像比较跨单位阻断：两侧单位不同 → 数量/单价/金额差值一律 None（C2）；
- 隐藏/深度隐藏 Sheet 导入门控：不自动确认、不写期次/明细（C4）。
"""

from decimal import Decimal
from pathlib import Path

import pytest
from openpyxl import Workbook

from jiadun.core.db import migrations
from jiadun.core.engine import settlement_io
from jiadun.core.engine.aggregate import aggregate_project
from jiadun.core.matching import matching
from jiadun.core.matching.mirror import build_mirror_comparison


@pytest.fixture()
def project_db(tmp_path):
    db_path = tmp_path / "project.db"
    migrations.migrate(db_path, tmp_path / "backups")
    conn = migrations.connect(db_path)
    with conn:
        project_id = conn.execute(
            "INSERT INTO projects(name, schema_version, workspace_path, created_at)"
            " VALUES (?, ?, ?, ?)",
            ("门控回归", migrations.LATEST_SCHEMA_VERSION, str(tmp_path), "2026"),
        ).lastrowid
        period_ids = []
        for period_no in (1, 2):
            period_ids.append(
                conn.execute(
                    "INSERT INTO settlement_periods(project_id, period_no, title, direction)"
                    " VALUES (?, ?, ?, 'upward')",
                    (project_id, period_no, f"第{period_no}期"),
                ).lastrowid
            )
    yield conn, project_id, period_ids
    conn.close()


def _line(conn, period_id, *, code="", name="", unit="", quantity="1", amount="10", feature=""):
    with conn:
        conn.execute(
            "INSERT INTO line_items(period_id, code, name, feature, unit, quantity, amount)"
            " VALUES (?, ?, ?, ?, ?, ?, ?)",
            (period_id, code, name, feature, unit, quantity, amount),
        )


# ---------------------------------------------------------------- C3：模糊合并单位复检
def test_fuzzy_merge_with_mixed_units_becomes_incomparable(project_db):
    """名称相似 ≥97 且单位不同：合并组必须判不可比并保留双单位，不得 probable。"""
    conn, project_id, (period_1, _period_2) = project_db
    base = "现浇C25商品混凝土垫层泵送商品砼运至现场含浇捣养护及运输费"
    with conn:
        id1 = conn.execute(
            "INSERT INTO line_items(period_id, name, unit, quantity) VALUES (?,?,?,?)",
            (period_1, base, "m³", "100"),
        ).lastrowid
        id2 = conn.execute(
            "INSERT INTO line_items(period_id, name, unit, quantity) VALUES (?,?,?,?)",
            (period_1, base + "率", "m²", "3000"),
        ).lastrowid

    groups = matching.match_items(conn, project_id)

    merged = [g for g in groups if id1 in g.item_ids and id2 in g.item_ids]
    assert len(merged) == 1, "两行应进入同一组供人工拆分"
    group = merged[0]
    assert group.level == matching.INCOMPARABLE, f"实际级别 {group.level}"
    assert group.units == {"m2", "m3"}
    assert any("单位不一致" in note for note in group.notes)


def test_similar_name_probable_groups_get_review_hint(project_db):
    """相似度 85–97 的两个高概率组不合并，但必须留下'疑似相关'提示（C3b）。"""
    conn, project_id, (period_1, _period_2) = project_db
    with conn:
        conn.execute(
            "INSERT INTO line_items(period_id, name, unit) VALUES (?,?,?)",
            (period_1, "现浇商品混凝土垫层", "m³"),
        )
        conn.execute(
            "INSERT INTO line_items(period_id, name, unit) VALUES (?,?,?)",
            (period_1, "现浇商品混凝土垫层 泵送", "m³"),
        )

    groups = matching.match_items(conn, project_id)

    notes = [note for g in groups for note in g.notes]
    assert any("疑似相关" in note for note in notes), f"notes={notes}"


# ---------------------------------------------------------------- C2：镜像跨单位阻断
def test_mirror_cross_unit_numeric_fields_incomparable(project_db):
    """两侧单位不同：数量/单价/金额差值必须为 None 且标注不可比（C2）。"""
    conn, project_id, (period_1, period_2) = project_db
    with conn:
        conn.execute("UPDATE settlement_periods SET direction='downward' WHERE id=?", (period_2,))
    _line(conn, period_1, code="S2", name="商砼", unit="m²", quantity="10", amount="100")
    _line(conn, period_2, code="S2", name="商砼", unit="m³", quantity="2", amount="200")
    contract = matching.run_contract.ensure_run_contract(conn, project_id)
    with conn:
        match_id = conn.execute(
            "INSERT INTO matches(project_id, group_key, item_ids_json, level, method, score,"
            " status, run_signature, run_id) VALUES (?, ?, ?, ?, ?, ?, 'pending', ?, ?)",
            (project_id, "downward:code:S2", "[]", matching.PROBABLE, "code_exact", 0.8,
             contract.signature, contract.run_id),
        ).lastrowid

    result = build_mirror_comparison(conn, project_id, int(match_id))

    by_field = {f.field: f for f in result.fields}
    assert by_field["unit"].status == "单位存在差异"
    for field_name in ("quantity", "unit_price", "amount"):
        assert by_field[field_name].difference is None, field_name
        assert by_field[field_name].difference_rate is None, field_name
        assert by_field[field_name].status == "不可比（两侧单位不同）", field_name
    assert result.quantity_difference is None
    assert result.amount_difference is None
    assert result.amount_difference_rate is None


# ---------------------------------------------------------------- C5：聚合键含归一单位/特征
def test_aggregate_splits_same_name_different_units(project_db):
    """同码同名不同单位：聚合必须分行且标记不可比，绝不出现跨单位合计。"""
    conn, project_id, (period_1, period_2) = project_db
    _line(conn, period_1, code="S1", name="商砼", unit="m²", quantity="10", amount="100")
    _line(conn, period_2, code="S1", name="商砼", unit="m³", quantity="2", amount="200")

    aggs = aggregate_project(conn, project_id, direction="upward")

    assert len(aggs) == 2
    units = {agg.item_key.rsplit("|u:", 1)[1] for agg in aggs}
    assert units == {"m2", "m3"}
    assert not any(agg.cum_qty == Decimal("12") for agg in aggs)
    assert all(
        any("不可比" in warning for warning in agg.warnings) for agg in aggs
    )


def test_aggregate_merges_normalized_name_variants_same_unit(project_db):
    """同名变体（空格差异）同单位：归一后应合并为同一累计。"""
    conn, project_id, (period_1, period_2) = project_db
    _line(conn, period_1, name="水泥砂浆 楼地面", unit="m²", amount="10")
    _line(conn, period_2, name="水泥砂浆楼地面", unit="m²", amount="20")

    aggs = aggregate_project(conn, project_id, direction="upward")

    assert len(aggs) == 1
    assert aggs[0].cum_amount == Decimal("30")


# ---------------------------------------------------------------- C4：隐藏 Sheet 门控
def _write_workbook(path: Path) -> None:
    wb = Workbook()
    ws = wb.active
    ws.title = "第三期对上"
    ws.append(["编码", "名称", "项目特征", "单位", "数量", "单价"])
    ws.append(["010401001", "垫层", "C25商品混凝土", "m³", 100, 350])
    hidden = wb.create_sheet("备注底稿")
    hidden.sheet_state = "hidden"
    hidden.append(["编码", "名称", "项目特征", "单位", "数量", "单价"])
    hidden.append(["010401001", "垫层", "C25商品混凝土(底稿口径)", "m²", 3000, 11.67])
    very_hidden = wb.create_sheet("废弃草稿")
    very_hidden.sheet_state = "veryHidden"
    very_hidden.append(["编码", "名称", "项目特征", "单位", "数量", "单价"])
    very_hidden.append(["010401001", "垫层", "C25商品混凝土(草稿口径)", "m²", 9999, 1])
    wb.save(path)


@pytest.fixture()
def hidden_sheet_project(tmp_path):
    src = tmp_path / "结算_含隐藏表.xlsx"
    _write_workbook(src)
    db_path = tmp_path / "project.db"
    migrations.migrate(db_path, tmp_path / "backups")
    conn = migrations.connect(db_path)
    with conn:
        project_id = conn.execute(
            "INSERT INTO projects(name, schema_version, workspace_path, created_at)"
            " VALUES ('隐藏门控', 1, ?, datetime('now'))",
            (str(tmp_path),),
        ).lastrowid
    conn.commit()
    yield conn, project_id, tmp_path, src
    conn.close()


def test_hidden_sheets_are_never_auto_confirmed(hidden_sheet_project):
    """隐藏/veryHidden Sheet 不得自动确认，也不得写入期次与明细（C4）。"""
    conn, project_id, project_dir, src = hidden_sheet_project

    report = settlement_io.import_settlement_file(
        conn, project_id, project_dir, src, period_no=3, direction="upward"
    )

    sheet_states = {
        sr.sheet_name: (sr.status, getattr(sr, "state_code", None)) for sr in report.sheets
    }
    assert sheet_states["第三期对上"][0] == "parsed"
    assert sheet_states["备注底稿"][0] != "parsed"
    assert sheet_states["废弃草稿"][0] != "parsed"

    statuses = {
        row["sheet_name"]: row["sheet_status"]
        for row in conn.execute("SELECT sheet_name, sheet_status FROM raw_sheets")
    }
    assert statuses["第三期对上"] == "confirmed"
    assert statuses["备注底稿"] == "pending"
    assert statuses["废弃草稿"] == "pending"

    rows = conn.execute(
        """SELECT rs.sheet_name, li.id FROM line_items li
           LEFT JOIN raw_sheets rs ON rs.id = li.sheet_id"""
    ).fetchall()
    assert all(row["sheet_name"] == "第三期对上" for row in rows), "隐藏表行不得写入明细"
    assert len(rows) == 1

    periods = conn.execute("SELECT title FROM settlement_periods").fetchall()
    assert len(periods) == 1, "隐藏表不得创建结算期次"
    assert "备注底稿" not in periods[0]["title"]

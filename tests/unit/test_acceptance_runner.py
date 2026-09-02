"""验收执行器非破坏性 + 表单路由（监督第五/六/七轮，先红后绿）。

1) runner 必须使用时间戳 run 目录（微秒唯一，同秒碰撞安全）：重复运行不得
   覆盖/删除既有 run 结果（基线保留）；支持可恢复续跑且不破坏未完成现场；
2) 键值对表单（支付审批单类）路由为 non_settlement_form 待人工状态：
   绝不写入 settlement/contract 模型；候选仅存通用 evidence 表；
   纯表单导入归类为 partial/needs_manual_review，不进入结算计算与导出；
3) steps.compute 必须反映真实计算（有 groups/checks/复算结果），不得用
   解析成功代替。
"""
import hashlib
import json
import subprocess
from pathlib import Path

import pytest


def _make_form_workbook(path: Path) -> None:
    """脱敏复现 R08 结构：键值对支付审批单（无列式表头、大量合并）。"""
    import openpyxl

    wb = openpyxl.Workbook()
    wb.remove(wb.active)
    ws = wb.create_sheet("VYSMSZ")
    ws.cell(row=1, column=1, value=" ")  # 近空 sheet
    ws2 = wb.create_sheet("支付审批单")
    rows = [
        ("××公司（脱敏示例）",),
        ("资金支付审批单",),
        ("申请单位名称：", None, "甲单位（脱敏）"),  # 键带冒号 + 值在右侧第2列（间隔布局）
        ("合同名称", "某工程（脱敏）"),
        ("合同编号", "HZ-001"),
        ("收款方信息", "户名", "乙单位（脱敏）"),
        ("", "开户行", "某银行"),
        ("本次申请支付金额", None, "1,000,000.00"),  # 金额在 C 列（col3）
    ]
    for r, row in enumerate(rows, start=1):
        for c, v in enumerate(row, start=1):
            if v is not None:
                ws2.cell(row=r, column=c, value=v)
    ws2.merge_cells("A1:J1")
    ws2.merge_cells("A2:J2")
    ws2.merge_cells("A6:A7")
    wb.save(path)


@pytest.fixture()
def runner_env(tmp_path, monkeypatch):
    """给 runner 注入临时 base 目录与 13 行假 corpus（脱敏）。"""
    import scripts.real_acceptance_run as runner

    base = tmp_path / "real_acceptance"
    corpus = base / "corpus"
    corpus.mkdir(parents=True)
    rows = []
    for i in range(1, 14):
        src = corpus / f"T{i:02d}_sample.xlsx"
        _make_form_workbook(src) if i == 1 else _make_simple(src)
        rows.append({
            "test_id": f"T{i:02d}",
            "source_path": f"orig/T{i:02d}",
            "copy_path": f"corpus/T{i:02d}_sample.xlsx",
            "sha256": runner.sha256_of(src),
            "purpose": "脱敏回归",
        })
    manifest = base / "manifest.csv"
    import csv

    with open(manifest, "w", encoding="utf-8", newline="") as f:
        w = csv.DictWriter(f, fieldnames=["test_id", "source_path", "copy_path", "sha256", "purpose"])
        w.writeheader()
        w.writerows(rows)
    monkeypatch.setattr(runner, "BASE", base)
    monkeypatch.setattr(runner, "WORK", base / "work")
    monkeypatch.setattr(runner, "MANIFEST", manifest)
    monkeypatch.setattr(runner, "DECISIONS", base / "manual_sheet_decisions.json")
    return runner, base


def _make_simple(src: Path) -> None:
    import openpyxl

    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "第1期"
    for c, v in enumerate(["清单编码", "清单名称", "单位", "工程量", "综合单价", "合价"], start=1):
        ws.cell(row=1, column=c, value=v)
    ws.cell(row=2, column=1, value="K1")
    ws.cell(row=2, column=2, value="某清单")
    ws.cell(row=2, column=3, value="m3")
    ws.cell(row=2, column=4, value=10)
    ws.cell(row=2, column=5, value=10)
    ws.cell(row=2, column=6, value=100)
    wb.save(src)


def _git_test(root: Path, *args: str) -> None:
    """执行临时 Git 仓库命令，不读取或修改当前工作仓库。"""
    subprocess.run(["git", *args], cwd=root, check=True, capture_output=True)


def _expected_head_diff_hash(root: Path) -> str:
    raw = subprocess.run(
        ["git", "diff", "HEAD", "--binary", "--", ".", ":!local_private_data"],
        cwd=root,
        check=True,
        capture_output=True,
    ).stdout
    return hashlib.sha256(raw).hexdigest()


def test_acceptance_bundle_hash_covers_staged_unstaged_and_mixed_changes(tmp_path):
    """运行包的 tracked diff 必须覆盖三种工作树状态且排除私密目录。"""
    from jiadun.core.acceptance import bundle

    repo = tmp_path / "repo"
    repo.mkdir()
    (repo / "tracked.txt").write_text("base\n", encoding="utf-8")
    (repo / "other.txt").write_text("base\n", encoding="utf-8")
    private = repo / "local_private_data"
    private.mkdir()
    (private / "secret.txt").write_text("secret-base\n", encoding="utf-8")
    _git_test(repo, "init", "-q")
    _git_test(repo, "add", ".")
    subprocess.run(
        [
            "git",
            "-c",
            "user.name=Jiadun Test",
            "-c",
            "user.email=jiadun-test@example.invalid",
            "commit",
            "-qm",
            "initial",
        ],
        cwd=repo,
        check=True,
        capture_output=True,
    )

    tracked = repo / "tracked.txt"
    other = repo / "other.txt"
    secret = private / "secret.txt"

    # 仅暂存：旧实现 git diff（不带 HEAD）会错误返回空 patch。
    tracked.write_text("staged\n", encoding="utf-8")
    secret.write_text("secret-staged\n", encoding="utf-8")
    _git_test(repo, "add", "tracked.txt", "local_private_data/secret.txt")
    assert bundle._tracked_diff_hash(repo) == _expected_head_diff_hash(repo)
    assert "local_private_data" not in subprocess.run(
        ["git", "diff", "HEAD", "--binary", "--", ".", ":!local_private_data"],
        cwd=repo,
        check=True,
        capture_output=True,
        text=True,
    ).stdout

    # 仅未暂存：混合状态前先把索引恢复到 HEAD，保留工作区内容。
    _git_test(repo, "reset", "-q", "HEAD")
    tracked.write_text("unstaged\n", encoding="utf-8")
    secret.write_text("secret-unstaged\n", encoding="utf-8")
    assert bundle._tracked_diff_hash(repo) == _expected_head_diff_hash(repo)

    # 混合状态：一个文件已暂存，另一个仅在工作区修改，同时仍不能泄露私密目录。
    tracked.write_text("mixed-staged\n", encoding="utf-8")
    other.write_text("mixed-unstaged\n", encoding="utf-8")
    secret.write_text("secret-mixed\n", encoding="utf-8")
    _git_test(repo, "add", "tracked.txt", "local_private_data/secret.txt")
    assert bundle._tracked_diff_hash(repo) == _expected_head_diff_hash(repo)


def test_decimal_warning_forces_with_findings() -> None:
    """Decimal 复算警告必须阻止无条件 passed，即使其他门槛全绿。"""
    import scripts.real_acceptance_run as runner

    status = runner.classify_technical_validation(
        technical_execution_complete=True,
        ab_check_status="ab_passed",
        evidence_status="available",
        high_findings=0,
        control_status="not_available",
        anomaly_total=0,
        decimal_warning_groups=1,
    )

    assert status == "with_findings"


def test_corpus_preflight_classifies_missing_and_hash_mismatch(runner_env):
    """真实资料副本缺失/哈希不符必须可分类为 pending，而不是裸断言。"""
    runner, base = runner_env
    missing = base / "corpus" / "T12_sample.xlsx"
    missing.unlink()
    altered = base / "corpus" / "T13_sample.xlsx"
    altered.write_bytes(altered.read_bytes() + b"altered")

    with open(base / "manifest.csv", encoding="utf-8") as f:
        import csv

        records = list(csv.DictReader(f))
    pre = runner.verify_corpus(records)

    summary = runner.summarize_corpus_preflight(pre)

    assert summary == {
        "status": "pending",
        "record_count": 13,
        "ready_count": 11,
        "pending_count": 2,
        "missing_test_ids": ["T12"],
        "hash_mismatch_test_ids": ["T13"],
    }


def test_corpus_preflight_records_hash_read_error_as_pending(runner_env, monkeypatch):
    """文件存在但无法读取哈希时，也必须进入完整性待复核而非中断整批。"""
    runner, base = runner_env

    def unreadable(_path):
        raise PermissionError("simulated file lock")

    monkeypatch.setattr(runner, "sha256_of", unreadable)
    with open(base / "manifest.csv", encoding="utf-8") as f:
        import csv

        records = list(csv.DictReader(f))
    pre = runner.verify_corpus(records)

    assert pre[0]["exists"] is True
    assert pre[0]["hash_match"] is False
    assert pre[0]["hash_error"].startswith("PermissionError:")
    summary = runner.summarize_corpus_preflight(pre)
    assert summary["status"] == "pending"
    assert summary["pending_count"] == 13
    pending = runner._pending_source_result(pre[0])
    assert pending["preflight"]["reason"] == "hash_unreadable"
    assert "无法读取 SHA-256" in pending["preflight"]["message"]


def test_runner_records_partial_corpus_as_pending_without_processing_bad_files(
    runner_env, monkeypatch
):
    """预检不完整时仍生成可恢复 run；坏副本不得进入解析/计算。"""
    runner, base = runner_env
    missing = base / "corpus" / "T12_sample.xlsx"
    missing.unlink()
    altered = base / "corpus" / "T13_sample.xlsx"
    altered.write_bytes(altered.read_bytes() + b"altered")

    seen: list[str] = []
    original_inspect = runner.inspect_file

    def track_inspect(test_id, *args, **kwargs):
        seen.append(test_id)
        return original_inspect(test_id, *args, **kwargs)

    monkeypatch.setattr(runner, "inspect_file", track_inspect)
    runner.main()

    run = sorted((base / "work").glob("run_*"))[-1]
    report = json.loads((run / "acceptance_results.json").read_text(encoding="utf-8"))
    assert report["preflight"] == {
        "status": "pending",
        "record_count": 13,
        "ready_count": 11,
        "pending_count": 2,
        "missing_test_ids": ["T12"],
        "hash_mismatch_test_ids": ["T13"],
    }
    assert "T12" not in seen and "T13" not in seen
    assert report["hash_check"]["before_all_match"] is False
    assert report["hash_check"]["after_all_match"] is False
    human_report = (run / "LOCAL_ACCEPTANCE_REPORT.md").read_text(encoding="utf-8")
    assert "## 副本预检" in human_report
    assert "待补资料：T12" in human_report
    assert "哈希待复核：T13" in human_report
    assert "完整性预检未全部通过" in human_report

    for test_id, reason in (("T12", "missing"), ("T13", "hash_mismatch")):
        result = json.loads((run / "done" / f"{test_id}.json").read_text(encoding="utf-8"))
        assert result["preflight"]["status"] == "pending"
        assert result["preflight"]["reason"] == reason
        assert result["steps"]["technical_execution_complete"] is False
        assert result["steps"]["technical_validation_status"] == "not_run_or_incomplete"
        assert result["steps"]["overall_acceptance_status"] == "pending_source_data"
        assert result["steps"]["verification_level"] == "insufficient"
        assert result.get("decimal_recompute") in (None, {})
        assert result.get("dual_path_check") in (None, {}, [])


def test_pending_source_can_resume_in_same_run_after_copy_is_restored(runner_env):
    """补齐并校正副本后，同一 run 的 pending marker 应重新进入流程。"""
    runner, base = runner_env
    source = base / "corpus" / "T13_sample.xlsx"
    original = source.read_bytes()
    source.unlink()

    runner.main()
    run = sorted((base / "work").glob("run_*"))[-1]
    pending = json.loads((run / "done" / "T13.json").read_text(encoding="utf-8"))
    assert pending["preflight"]["reason"] == "missing"

    source.write_bytes(original)
    runner.main(run_dir=run)
    resumed = json.loads((run / "done" / "T13.json").read_text(encoding="utf-8"))
    assert resumed.get("preflight") in (None, {})
    assert resumed["steps"]["import"] is True
    assert resumed["steps"]["overall_acceptance_status"] != "pending_source_data"


def test_resume_downgrades_completed_marker_when_copy_changes(runner_env):
    """已完成 marker 不能掩盖续跑时发现的副本篡改。"""
    runner, base = runner_env
    runner.main()
    run = sorted((base / "work").glob("run_*"))[-1]
    source = base / "corpus" / "T13_sample.xlsx"
    source.write_bytes(source.read_bytes() + b"changed-after-run")

    runner.main(run_dir=run)
    result = json.loads((run / "done" / "T13.json").read_text(encoding="utf-8"))

    assert result["preflight"]["reason"] == "hash_mismatch"
    assert result["steps"]["overall_acceptance_status"] == "pending_source_data"
    assert result.get("decimal_recompute") in (None, {})
    assert result.get("dual_path_check") in (None, {}, [])


def test_post_run_copy_mutation_invalidates_processed_result(runner_env, monkeypatch):
    """处理期间源副本被修改时，已生成结果必须整体降级为待补资料。"""
    runner, base = runner_env
    original_inspect = runner.inspect_file

    def inspect_then_mutate(test_id, purpose, copy, *args, **kwargs):
        result = original_inspect(test_id, purpose, copy, *args, **kwargs)
        if test_id == "T01":
            copy.write_bytes(copy.read_bytes() + b"changed-during-run")
        return result

    monkeypatch.setattr(runner, "inspect_file", inspect_then_mutate)
    runner.main()

    run = sorted((base / "work").glob("run_*"))[-1]
    report = json.loads((run / "acceptance_results.json").read_text(encoding="utf-8"))
    result = next(item for item in report["per_file"] if item["test_id"] == "T01")

    assert report["hash_check"]["after_all_match"] is False
    assert report["hash_check"]["modified_copies"] == ["T01"]
    assert report["hash_check"]["invalidated_results"] == ["T01"]
    assert result["preflight"]["reason"] == "modified_after_processing"
    assert result["steps"]["overall_acceptance_status"] == "pending_source_data"
    assert result["steps"]["technical_execution_complete"] is False
    assert result.get("decimal_recompute") in (None, {})
    assert result.get("dual_path_check") in (None, {}, [])


def test_pipeline_stage_failure_is_recorded_and_run_remains_recoverable(
    runner_env, monkeypatch
):
    """异常检测失败不能中断整批，必须留下 done marker 和汇总报告。"""
    runner, base = runner_env
    from jiadun.core.anomalies import engine as anomaly_engine

    def fail_anomalies(*_args, **_kwargs):
        raise RuntimeError("anomaly boom")

    monkeypatch.setattr(anomaly_engine, "run_anomalies", fail_anomalies)
    runner.main()

    run = sorted((base / "work").glob("run_*"))[-1]
    assert (run / "acceptance_results.json").is_file()
    assert (run / "LOCAL_ACCEPTANCE_REPORT.md").is_file()
    result = json.loads((run / "done" / "T02.json").read_text(encoding="utf-8"))
    assert result["pipeline_error"]["stage"] == "anomalies"
    assert "anomaly boom" in result["pipeline_error"]["error"]
    assert result["steps"]["anomalies"] is False
    assert result["steps"]["matches"] is False
    assert result["steps"]["technical_execution_complete"] is False
    assert result["steps"]["overall_acceptance_status"] == "not_passed"


def test_failed_marker_can_be_retried_after_input_or_stage_is_fixed(runner_env, monkeypatch):
    """失败 marker 不得被续跑逻辑永久当成已完成结果。"""
    runner, base = runner_env
    from jiadun.core.anomalies import engine as anomaly_engine

    original = anomaly_engine.run_anomalies
    calls = {"count": 0}

    def fail_once(*args, **kwargs):
        calls["count"] += 1
        if calls["count"] == 1:
            raise RuntimeError("retryable anomaly failure")
        return original(*args, **kwargs)

    monkeypatch.setattr(anomaly_engine, "run_anomalies", fail_once)
    runner.main()
    run = sorted((base / "work").glob("run_*"))[-1]
    failed = json.loads((run / "done" / "T02.json").read_text(encoding="utf-8"))
    assert failed["pipeline_error"]["stage"] == "anomalies"

    runner.main(run_dir=run)
    retried = json.loads((run / "done" / "T02.json").read_text(encoding="utf-8"))
    assert "pipeline_error" not in retried
    assert retried["import"]["ok"] is True
    assert (run / "验收-T02_r2").exists()


def test_unexpected_file_exception_is_recorded_without_aborting_batch(runner_env, monkeypatch):
    """文件级未预期异常也要转为 marker，后续 test_id 仍可继续。"""
    runner, base = runner_env
    original_inspect = runner.inspect_file

    def fail_one(test_id, *args, **kwargs):
        if test_id == "T02":
            raise OSError("synthetic inspect failure")
        return original_inspect(test_id, *args, **kwargs)

    monkeypatch.setattr(runner, "inspect_file", fail_one)
    runner.main()

    run = sorted((base / "work").glob("run_*"))[-1]
    failed = json.loads((run / "done" / "T02.json").read_text(encoding="utf-8"))
    completed = json.loads((run / "done" / "T03.json").read_text(encoding="utf-8"))
    assert failed["pipeline_error"]["stage"] == "inspect_file"
    assert "synthetic inspect failure" in failed["pipeline_error"]["error"]
    assert failed["steps"]["technical_execution_complete"] is False
    assert completed["test_id"] == "T03"
    assert (run / "acceptance_results.json").is_file()


def test_word_export_failure_does_not_leave_partial_excel_result(runner_env, monkeypatch):
    """成对导出任一格式失败时，本次 Excel 不得作为看似完整成果残留。"""
    runner, base = runner_env
    from jiadun.core.export import excel_export

    original_docx = excel_export.export_management_summary_docx
    failed_once = {"value": False}

    def fail_first_docx(*args, **kwargs):
        if not failed_once["value"]:
            failed_once["value"] = True
            raise RuntimeError("synthetic Word export failure")
        return original_docx(*args, **kwargs)

    monkeypatch.setattr(excel_export, "export_management_summary_docx", fail_first_docx)
    runner.main()

    run = sorted((base / "work").glob("run_*"))[-1]
    first = json.loads((run / "done" / "T02.json").read_text(encoding="utf-8"))
    exports = list((run / "验收-T02" / "exports").glob("*"))
    assert first["export"]["error"]
    assert first["steps"]["technical_execution_complete"] is False
    assert exports == []


def test_partial_export_files_are_rolled_back_when_export_raises(runner_env, monkeypatch):
    """导出函数已写出半文件后抛异常时，验收包装层仍必须清理现场。"""
    runner, base = runner_env
    from jiadun.core.export import excel_export

    original_xlsx = excel_export.export_workbook
    original_docx = excel_export.export_management_summary_docx
    raised = {"xlsx": False, "docx": False}

    def partial_xlsx(conn, project_id, export_dir):
        if not raised["xlsx"]:
            raised["xlsx"] = True
            export_dir.mkdir(parents=True, exist_ok=True)
            (export_dir / "partial.xlsx").write_bytes(b"partial")
            raise RuntimeError("synthetic partial Excel failure")
        return original_xlsx(conn, project_id, export_dir)

    def partial_docx(conn, project_id, export_dir):
        if not raised["docx"]:
            raised["docx"] = True
            export_dir.mkdir(parents=True, exist_ok=True)
            (export_dir / "partial.docx").write_bytes(b"partial")
            raise RuntimeError("synthetic partial Word failure")
        return original_docx(conn, project_id, export_dir)

    monkeypatch.setattr(excel_export, "export_workbook", partial_xlsx)
    monkeypatch.setattr(excel_export, "export_management_summary_docx", partial_docx)
    runner.main()

    run = sorted((base / "work").glob("run_*"))[-1]
    failed_xlsx = json.loads((run / "done" / "T02.json").read_text(encoding="utf-8"))
    failed_docx = json.loads((run / "done" / "T03.json").read_text(encoding="utf-8"))
    assert failed_xlsx["pipeline_error"]["stage"] == "export_excel"
    assert failed_docx["pipeline_error"]["stage"] == "export_word"
    assert list((run / "验收-T02" / "exports").glob("*")) == []
    assert list((run / "验收-T03" / "exports").glob("*")) == []


def test_nested_same_basename_partial_export_is_rolled_back(tmp_path):
    """嵌套目录中已有同名文件时，新半文件仍必须被识别并清理。"""
    import scripts.real_acceptance_run as runner

    export_dir = tmp_path / "exports"
    old = export_dir / "old" / "partial.xlsx"
    old.parent.mkdir(parents=True)
    old.write_bytes(b"old")
    existing_paths = {old.relative_to(export_dir).as_posix()}
    new = export_dir / "new" / "partial.xlsx"
    new.parent.mkdir(parents=True)
    new.write_bytes(b"new")

    assert runner._new_export_files(export_dir, existing_paths) == [new]


def test_corrupt_done_marker_is_recorded_and_does_not_abort_batch(runner_env):
    """损坏的续跑 marker 应转为结构化失败，后续文件仍可继续处理。"""
    runner, base = runner_env
    runner.main()
    run = sorted((base / "work").glob("run_*"))[-1]
    marker = run / "done" / "T02.json"
    marker.write_text("{broken", encoding="utf-8")

    runner.main(run_dir=run)

    failed = json.loads(marker.read_text(encoding="utf-8"))
    continued = json.loads((run / "done" / "T03.json").read_text(encoding="utf-8"))
    assert failed["pipeline_error"]["stage"] == "done_marker"
    assert "JSONDecodeError" in failed["pipeline_error"]["error"]
    assert failed["steps"]["technical_execution_complete"] is False
    assert continued["test_id"] == "T03"


def test_semantically_incomplete_done_marker_is_not_reused(runner_env):
    """合法 JSON 但缺关键步骤/身份的 marker 必须重新处理。"""
    runner, base = runner_env
    runner.main()
    run = sorted((base / "work").glob("run_*"))[-1]
    marker = run / "done" / "T02.json"
    marker.write_text(
        json.dumps({"test_id": "T02", "steps": {}, "import": {"ok": False}}),
        encoding="utf-8",
    )

    runner.main(run_dir=run)

    repaired = json.loads(marker.read_text(encoding="utf-8"))
    assert repaired["marker"]["test_id"] == "T02"
    assert repaired["steps"]["import"] is True
    assert "pipeline_error" not in repaired


def test_done_marker_with_wrong_test_id_is_not_reused(runner_env):
    """不同 test_id 的合法 marker 不能因文件名相同而错归属。"""
    runner, base = runner_env
    runner.main()
    run = sorted((base / "work").glob("run_*"))[-1]
    marker = run / "done" / "T02.json"
    previous = json.loads(marker.read_text(encoding="utf-8"))
    previous["marker"]["test_id"] = "T03"
    previous["test_id"] = "T03"
    marker.write_text(json.dumps(previous), encoding="utf-8")

    runner.main(run_dir=run)

    repaired = json.loads(marker.read_text(encoding="utf-8"))
    assert repaired["marker"]["test_id"] == "T02"
    assert repaired["test_id"] == "T02"


def test_corrupt_manual_decisions_are_structured_and_recoverable(runner_env):
    """损坏人工决定文件不能在批次入口裸抛异常或锁死界面。"""
    runner, base = runner_env
    (base / "manual_sheet_decisions.json").write_text("{broken", encoding="utf-8")

    runner.main()

    run = sorted((base / "work").glob("run_*"))[-1]
    failed = json.loads((run / "done" / "T01.json").read_text(encoding="utf-8"))
    continued = json.loads((run / "done" / "T02.json").read_text(encoding="utf-8"))
    assert failed["pipeline_error"]["stage"] == "manual_decisions"
    assert "JSONDecodeError" in failed["pipeline_error"]["error"]
    assert failed["steps"]["technical_execution_complete"] is False
    assert continued["test_id"] == "T02"


@pytest.mark.parametrize(
    ("verification_level", "range_unproven_sheets"),
    [("insufficient", 0), ("findings", 0), ("sufficient", 1)],
)
def test_acceptance_never_marks_unproven_crosscheck_as_passed(
    verification_level, range_unproven_sheets
) -> None:
    import scripts.real_acceptance_run as runner

    status = runner.classify_technical_validation(
        technical_execution_complete=True,
        ab_check_status="ab_passed",
        evidence_status="available",
        high_findings=0,
        control_status="passed",
        anomaly_total=0,
        decimal_warning_groups=0,
        verification_level=verification_level,
        range_unproven_sheets=range_unproven_sheets,
    )
    assert status == "with_findings"


class TestRunnerNonDestructive:
    def test_timestamped_runs_preserve_previous(self, runner_env):
        """两次运行产生两个 run 目录：第一次结果必须原样保留（基线不覆盖）。"""
        runner, base = runner_env
        runner.main()
        runs = sorted((base / "work").glob("run_*"))
        assert len(runs) == 1, f"首次运行应产生 1 个 run 目录: {runs}"
        first_result = runs[0] / "acceptance_results.json"
        first_bytes = first_result.read_bytes()

        runner.main()  # 第二次运行
        runs = sorted((base / "work").glob("run_*"))
        assert len(runs) == 2, "第二次运行必须新建 run 目录"
        # 第一次 run 的结果逐字节保留
        assert runs[0].joinpath("acceptance_results.json").read_bytes() == first_bytes
        # 每轮都包含全部 13 个 test_id 项目
        projects = [d.name for d in runs[1].iterdir() if d.is_dir() and d.name.startswith("验收-")]
        assert len(projects) == 13

    def test_timestamped_runs_preserve_human_reports(self, runner_env):
        """每轮 Markdown 报告随 run 保存，不能覆盖上一轮验收记录。"""
        runner, base = runner_env
        runner.main()
        first = sorted((base / "work").glob("run_*"))[0]
        first_report = first / "LOCAL_ACCEPTANCE_REPORT.md"
        assert first_report.is_file()
        first_bytes = first_report.read_bytes()

        runner.main()
        runs = sorted((base / "work").glob("run_*"))
        assert len(runs) == 2
        assert runs[0].joinpath("LOCAL_ACCEPTANCE_REPORT.md").read_bytes() == first_bytes
        assert runs[1].joinpath("LOCAL_ACCEPTANCE_REPORT.md").is_file()

    def test_resume_skips_completed(self, runner_env):
        """可恢复重跑：同 run 目录续跑时跳过已完成 test_id。"""
        runner, base = runner_env
        runner.main()
        runs = sorted((base / "work").glob("run_*"))
        first = runs[-1]
        done = {d.name for d in first.iterdir() if d.is_dir() and d.name.startswith("验收-")}
        # 续跑同一 run 目录：全部已完成 → 不重复建项目
        runner.main(run_dir=first)
        after = {d.name for d in first.iterdir() if d.is_dir() and d.name.startswith("验收-")}
        assert after == done, "续跑不得重建已完成项目"


class TestFormRouting:
    def test_form_sheet_routed_out_of_settlement_model(self, tmp_path):
        """键值对支付表单必须路由为 non_settlement_form：

        - 不生成 settlement_period / line_items / period_totals（不得污染结算模型）；
        - 保留原 Sheet 与单元格（保真层不变）；
        - 事实候选进既有 contract_facts（含原文、单元格位置、证据ID、待人工状态）；
        - 审计留痕。
        """
        from jiadun.core.engine import settlement_io
        from jiadun.core.models import project as pm

        src = tmp_path / "form.xlsx"
        _make_form_workbook(src)
        info = pm.create_project("表单路由", tmp_path / "ws")
        info, conn = pm.open_project(Path(info.workspace_path))
        try:
            report = settlement_io.import_settlement_file(
                conn, info.project_id, Path(info.workspace_path), src)
            form_sheet = next(s for s in report.sheets if s.sheet_name == "支付审批单")
            assert form_sheet.status == "non_settlement_form", \
                f"表单 sheet 应路由为 non_settlement_form，得到 {form_sheet.status}"
            assert any("人工" in n or "表单" in n for n in form_sheet.notes), \
                f"必须给出可恢复的人工复核提示: {form_sheet.notes}"
            joined = "".join(form_sheet.notes)
            assert "通用 evidence 人工复核入口" in joined, \
                f"notes 必须指向通用 evidence 人工复核入口: {joined}"
            assert "合同" not in joined and "contract" not in joined.lower(), \
                f"notes 不得暗示写入合同模型: {joined}"

            # 结算与合同模型全部零污染（contract_docs/facts 是合同模块专用，
            # 表单路由不得写入，否则 contract_risks 会产生虚假合同风险）
            for table in ("settlement_periods", "line_items", "period_totals",
                          "contract_docs", "contract_facts"):
                c = conn.execute(f"SELECT COUNT(*) c FROM {table}").fetchone()["c"]
                assert c == 0, f"{table} 被表单路由污染（{c} 行）"

            # 保真层保留原 Sheet
            assert conn.execute(
                "SELECT COUNT(*) c FROM raw_sheets WHERE sheet_name='支付审批单'").fetchone()["c"] == 1
            assert conn.execute(
                "SELECT COUNT(*) c FROM raw_cells WHERE row=8 AND col=3").fetchone()["c"] == 1

            # 候选仅存通用证据表：带行列、原文、证据ID、待人工确认
            import json as _json

            evs = conn.execute(
                "SELECT id, kind, summary, sources_json FROM evidence WHERE kind='form_field_candidate'"
            ).fetchall()
            assert evs, "表单键值对应作为待人工确认的证据候选"
            # 金额候选必须用真实值列：金额在 C 列 → location 精确为 行8列3
            amt = next(e for e in evs if "支付金额" in e["summary"])
            assert "1,000,000.00" in amt["summary"], amt["summary"]
            amt_src = _json.loads(amt["sources_json"])[0]
            assert amt_src["location"] == "行8列3", \
                f"金额位置必须为真实值列 行8列3，得到 {amt_src['location']}"
            # 带冒号键（申请单位名称：）+ 右值间隔布局 → 候选不得遗漏，位置为真实值列 行3列3
            unit = next(e for e in evs if "申请单位名称" in e["summary"])
            assert "甲单位（脱敏）" in unit["summary"]
            assert _json.loads(unit["sources_json"])[0]["location"] == "行3列3"
            # 可反向定位：evidence 的行列必须对应 raw_cells 中的非空原格
            sheet_id = conn.execute(
                "SELECT id FROM raw_sheets WHERE sheet_name='支付审批单'").fetchone()["id"]
            for e in evs:
                src = _json.loads(e["sources_json"])[0]
                assert src["location"].startswith("行") and src["quote"], \
                    "每条候选必须带行列位置与原文"
                assert "待人工确认" in e["summary"]
                row_no = int(src["location"][1:src["location"].index("列")])
                col_no = int(src["location"][src["location"].index("列") + 1:])
                cell = conn.execute(
                    "SELECT raw_value FROM raw_cells WHERE sheet_id=? AND row=? AND col=?",
                    (sheet_id, row_no, col_no)).fetchone()
                assert cell and (cell["raw_value"] or "").strip(), \
                    f"证据位置不可反向定位: {src['location']}"

            # 回归：表单导入不得产生虚假合同风险
            from jiadun.core.contracts import extract as contract_extract

            assert contract_extract.contract_risks(conn, info.project_id) == []

            # 审计留痕
            from jiadun.core.evidence import audit as audit_log

            entries = audit_log.history_for(conn, info.project_id)
            assert any("form" in e.action or "表单" in e.reason for e in entries)
        finally:
            conn.close()

    def test_form_like_detection_hint(self, tmp_path):
        """键值对表单结构必须给出可恢复诊断提示（而非裸 no_header）。"""
        from jiadun.core.parsing.excel_parser import parse_file
        from jiadun.core.parsing.header_detect import detect_form_like

        src = tmp_path / "form.xlsx"
        _make_form_workbook(src)
        result = parse_file(src, "xlsx")
        ws2 = next(s for s in result.sheets if s.sheet_name == "支付审批单")
        cells = {(c.row, c.col): (c.raw_value or "") for c in ws2.cells}
        assert detect_form_like(cells, ws2.merged_ranges) in ("strong", "weak")
        plain = next(s for s in result.sheets if s.sheet_name == "VYSMSZ")
        plain_cells = {(c.row, c.col): (c.raw_value or "") for c in plain.cells}
        assert detect_form_like(plain_cells, plain.merged_ranges) is None

class TestPureFormProjectHandling:
    def test_pure_form_project_partial_and_skips_settlement_pipeline(self, runner_env, monkeypatch):
        """纯表单项目：status=partial（非 ok/failed），不进计算/异常/匹配/导出，
        steps 单列 non_settlement_form_needs_manual_review，技术执行状态为 False。"""
        runner, base = runner_env
        # T01 换成纯表单，其余 simple 清单
        manifest_t01 = base / "corpus" / "T01_sample.xlsx"
        _make_form_workbook(manifest_t01)
        import csv

        rows = []
        for i in range(1, 14):
            src = base / "corpus" / f"T{i:02d}_sample.xlsx"
            rows.append({
                "test_id": f"T{i:02d}", "source_path": f"orig/T{i:02d}",
                "copy_path": f"corpus/T{i:02d}_sample.xlsx",
                "sha256": runner.sha256_of(src), "purpose": "脱敏回归",
            })
        with open(base / "manifest.csv", "w", encoding="utf-8", newline="") as f:
            w = csv.DictWriter(f, fieldnames=["test_id", "source_path", "copy_path", "sha256", "purpose"])
            w.writeheader()
            w.writerows(rows)

        runner.main()
        runs = sorted((base / "work").glob("run_*"))
        t01 = json.loads((runs[-1] / "done" / "T01.json").read_text(encoding="utf-8"))
        steps = t01.get("steps", {})
        assert steps.get("non_settlement_form_needs_manual_review") is True, \
            f"纯表单必须单列 needs_manual_review: {steps}"
        assert steps.get("technical_execution_complete") is False
        assert steps.get("technical_validation_status") == "not_run_or_incomplete"
        assert steps.get("overall_acceptance_status") == "needs_manual_review"
        assert "full_pipeline" not in steps
        for k in ("compute", "anomalies", "matches", "excel", "word"):
            assert steps.get(k) is False, f"纯表单 steps.{k} 必须为 False: {steps}"
        # 口径：settlement parse 仅在 report.status=='ok' 时算成功；
        # 纯表单 partial 不得计入 parse
        assert steps.get("parse") is False, f"纯表单 partial 不得计入 parse: {steps}"
        assert (t01.get("settlement_parse") or {}).get("ok") is False, \
            "settlement_parse.ok 必须仅在 status=='ok' 时为 True"

        # ImportReport 状态：纯表单 = partial 且 needs_manual_review（非 ok 非普通 failed）
        rec = t01
        sp = rec.get("settlement_parse") or {}
        assert sp.get("status") == "partial", f"纯表单导入状态应为 partial: {sp}"
        assert sp.get("needs_manual_review") is True


class TestInterruptSafety:
    def test_interrupted_site_preserved_on_rerun(self, runner_env):
        """中断续跑不得删除未完成现场：旧项目目录保留，新目录带序号生成。"""
        runner, base = runner_env
        runner.main()
        runs = sorted((base / "work").glob("run_*"))
        first = runs[-1]
        # 模拟中断：T13 无 done marker，且项目目录留有现场标记
        (first / "done" / "T13.json").unlink(missing_ok=True)
        site = first / "验收-T13"
        site.mkdir(exist_ok=True)
        (site / "interrupted.marker").write_text("现场")
        # 删除另一个 test 的 marker 模拟两个未完成
        (first / "done" / "T12.json").unlink(missing_ok=True)

        runner.main(run_dir=first)  # 续跑
        # T13 旧现场必须保留（不得 rmtree）
        assert (site / "interrupted.marker").exists(), "中断现场被删除"
        # 重跑生成新目录（带序号），不覆盖旧目录
        new_dirs = sorted(first.glob("验收-T13*"))
        assert len(new_dirs) >= 2 and new_dirs[-1].name != "验收-T13"


class TestRunDirUniqueness:
    def test_same_second_runs_both_kept(self, runner_env):
        """同一秒连续两次运行：两个 run 目录都必须保留（微秒/唯一后缀）。"""
        runner, base = runner_env
        runner.main()
        runner.main()
        runs = sorted((base / "work").glob("run_*"))
        assert len(runs) == 2, f"同秒两次运行应产生 2 个 run 目录: {runs}"


class TestComputeTruthfulness:
    def test_compute_false_when_all_computation_fails(self, runner_env, monkeypatch):
        """steps.compute 必须反映真实计算：groups 与 checks 全部失败时 compute=False
        （不得用解析成功代替）。"""
        runner, base = runner_env
        from jiadun.core.engine import aggregate as agg_mod
        from jiadun.core.engine import crosscheck as cc_mod

        def boom(*a, **k):
            raise RuntimeError("computation broken")
        monkeypatch.setattr(agg_mod, "aggregate_project", boom)
        monkeypatch.setattr(cc_mod, "run_crosscheck", boom)
        runner.main()
        runs = sorted((base / "work").glob("run_*"))
        t02 = json.loads((runs[-1] / "done" / "T02.json").read_text(encoding="utf-8"))
        steps = t02.get("steps", {})
        assert steps.get("compute") is False, \
            f"复算失败时 compute 必须为 False: {steps}（不得用解析成功代替）"


class TestStepsRoleFormSeparation:
    def test_role_review_not_mislabeled_as_form(self, runner_env, monkeypatch):
        """角色审阅文件不得被误标为表单路由；steps 两列互斥分列。"""
        runner, base = runner_env
        # T01 换成弱表头台账（→ needs_role_review），其余 simple 清单
        import csv

        from test_acceptance_runner import _make_form_workbook  # noqa: F401

        # 弱台账样例
        src = base / "corpus" / "T01_sample.xlsx"
        import openpyxl

        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = "台账明细"
        for c, v in enumerate(["合同编号", "合同名称", "承包人", "合同金额"], start=1):
            ws.cell(row=1, column=c, value=v)
        for i in range(2, 6):
            ws.cell(row=i, column=1, value=f"HT-{i}")
            ws.cell(row=i, column=2, value=f"某合同{i}")
            ws.cell(row=i, column=3, value=f"对方{i}")
            ws.cell(row=i, column=4, value=1000 + i)
        wb.save(src)
        rows = []
        for i in range(1, 14):
            p = base / "corpus" / f"T{i:02d}_sample.xlsx"
            rows.append({"test_id": f"T{i:02d}", "source_path": f"orig/T{i:02d}",
                         "copy_path": f"corpus/T{i:02d}_sample.xlsx",
                         "sha256": runner.sha256_of(p), "purpose": "脱敏回归"})
        with open(base / "manifest.csv", "w", encoding="utf-8", newline="") as f:
            w = csv.DictWriter(f, fieldnames=["test_id", "source_path", "copy_path", "sha256", "purpose"])
            w.writeheader()
            w.writerows(rows)

        runner.main()
        runs = sorted((base / "work").glob("run_*"))
        t01 = json.loads((runs[-1] / "done" / "T01.json").read_text(encoding="utf-8"))
        steps = t01.get("steps", {})
        assert steps.get("non_settlement_spreadsheet_needs_role_review") is True, \
            f"角色审阅必须单列: {steps}"
        assert steps.get("non_settlement_form_needs_manual_review") is False, \
            f"角色审阅不得误标为表单路由: {steps}"
        rec = t01
        assert (rec.get("role_review") or {}).get("needs_manual_review") is True
        assert rec.get("form_route") in (None, {})


class TestManualSheetDecisions:
    def test_private_decisions_resolve_form_and_amount_only_sheet(self, runner_env):
        """私有人工决定应可安全完成：表单仅作证据，金额型明细按明确行列抽取。"""
        runner, base = runner_env
        import csv

        import openpyxl

        src = base / "corpus" / "T01_sample.xlsx"
        _make_form_workbook(src)
        wb = openpyxl.load_workbook(src)
        ws = wb.create_sheet("计算明细表")
        ws.cell(4, 1, "序号")
        ws.cell(4, 2, "项目名称")
        ws.cell(4, 3, "分包商申报")
        ws.cell(4, 4, "项目部审核")
        ws.cell(5, 3, "金额(元)")
        ws.cell(5, 4, "金额(元)")
        ws.cell(6, 1, "1")
        ws.cell(6, 2, "暂估价设备本期结算")
        ws.cell(6, 3, "3370591.52")
        ws.cell(6, 4, "3370591.52")
        ws.cell(10, 2, "金额合计")
        ws.cell(10, 4, "3370591.52")
        wb.save(src)

        rows = []
        for i in range(1, 14):
            p = base / "corpus" / f"T{i:02d}_sample.xlsx"
            rows.append({
                "test_id": f"T{i:02d}", "source_path": f"orig/T{i:02d}",
                "copy_path": f"corpus/T{i:02d}_sample.xlsx",
                "sha256": runner.sha256_of(p), "purpose": "脱敏回归",
            })
        with open(base / "manifest.csv", "w", encoding="utf-8", newline="") as f:
            w = csv.DictWriter(
                f, fieldnames=["test_id", "source_path", "copy_path", "sha256", "purpose"])
            w.writeheader()
            w.writerows(rows)

        (base / "manual_sheet_decisions.json").write_text(json.dumps({
            "version": 1,
            "files": {
                "T01": [
                    {"sheet": "支付审批单", "action": "evidence_only",
                     "role": "non_settlement_form", "actor": "验收复核人",
                     "reason": "人工确认：支付审批表仅作证据"},
                    {"sheet": "计算明细表", "action": "extract",
                     "actor": "验收复核人", "reason": "人工确认项目部审核金额列与唯一明细行",
                     "direction": "downward", "period_no": 2,
                     "col_map": {"code": 1, "name": 2, "amount": 4},
                     "header_range": [4, 5], "data_range": [6, 6]},
                ]
            },
        }, ensure_ascii=False), encoding="utf-8")

        runner.main()
        run = sorted((base / "work").glob("run_*"))[-1]
        rec = json.loads((run / "done" / "T01.json").read_text(encoding="utf-8"))
        assert rec["steps"]["technical_execution_complete"] is True
        assert rec["steps"]["technical_validation_status"] == "with_findings"
        assert rec["steps"]["overall_acceptance_status"] == "pending_wps_with_findings"
        assert rec["steps"]["ab_check_status"] == "ab_passed"
        assert rec["steps"]["control_status"] == "not_available"
        assert rec["steps"]["wps"] == "pending_manual"
        assert "full_pipeline" not in rec["steps"]
        assert rec["settlement_parse"]["status"] == "ok_after_manual_confirmation"
        assert len(rec["manual_sheet_decisions"]) == 2
        assert rec.get("form_route") in (None, {})
        assert rec.get("role_review") in (None, {})

        import sqlite3

        conn = sqlite3.connect(run / "验收-T01" / "project.db")
        conn.row_factory = sqlite3.Row
        row = conn.execute(
            "SELECT li.quantity, li.amount, sp.period_no, sp.direction "
            "FROM line_items li JOIN settlement_periods sp ON sp.id=li.period_id"
        ).fetchone()
        conn.close()
        assert row["quantity"] is None and row["amount"] == "3370591.52"
        assert row["period_no"] == 2 and row["direction"] == "downward"


class TestAcceptanceControls:
    def test_bridge_is_evidence_and_difference_remains_open_anomaly(
        self, runner_env, tmp_path
    ):
        runner, _ = runner_env
        from jiadun.core.models import project as pm

        info = pm.create_project("控制值记录", tmp_path / "workspace")
        info, conn = pm.open_project(Path(info.workspace_path))
        try:
            recorded = runner.record_acceptance_controls(conn, info.project_id, {
                "bridges": [{
                    "summary": "小计加税金桥接",
                    "steps": [{"formula": "100+3", "value": "103"}],
                    "sources": [{"sheet": "计算明细表", "location": "J16:J17"}],
                }],
                "differences": [{
                    "severity": "medium",
                    "summary": "公式值与旧缓存差异0.48元",
                    "sources": [{"sheet": "计算明细表", "location": "G7,K7"}],
                }],
            })
            assert [r["kind"] for r in recorded] == ["bridge", "difference"]
            assert conn.execute(
                "SELECT COUNT(*) c FROM evidence WHERE project_id=?",
                (info.project_id,),
            ).fetchone()["c"] == 2
            anomaly = conn.execute(
                "SELECT rule_id, severity, status FROM anomalies WHERE project_id=?",
                (info.project_id,),
            ).fetchone()
            assert dict(anomaly) == {
                "rule_id": "acceptance_control_difference",
                "severity": "medium",
                "status": "open",
            }
        finally:
            conn.close()

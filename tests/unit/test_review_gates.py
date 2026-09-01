"""独立复核后补充的 P0 闸门回归测试。

这些测试专门覆盖“数值碰巧一致但证据条件不充分”的反例，避免只用全量
pytest 通过来替代业务验收。
"""
from __future__ import annotations

import json
from decimal import Decimal
from pathlib import Path

import openpyxl
import pytest
from openpyxl.utils import get_column_letter


def _make_book(
    path: Path,
    *,
    two_grand_totals: bool = False,
    single_detail: bool = False,
    missing_detail_amount: bool = False,
    hidden_rows: list[int] | None = None,
    hidden_cols: list[int] | None = None,
) -> None:
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "第1期明细"
    for col, value in enumerate(("项目编码", "项目名称", "工程量", "综合单价", "合价"), 1):
        ws.cell(1, col, value)
    detail_rows = [("A1", "项目A", 2, 100, 200)]
    if not single_detail:
        detail_rows.append(("A2", "项目B", 1, 300, 300))
    if missing_detail_amount:
        code, name, quantity, unit_price, _amount = detail_rows[0]
        detail_rows[0] = (code, name, quantity, unit_price, None)
    for row in detail_rows:
        ws.append(row)
    control_total = 200 if single_detail else 500
    ws.append((None, "本页小计", None, None, control_total))
    ws.append((None, "合计", None, None, control_total))
    if two_grand_totals:
        ws.append((None, "总计", None, None, control_total))
    for row_number in hidden_rows or []:
        ws.row_dimensions[int(row_number)].hidden = True
    for column_number in hidden_cols or []:
        ws.column_dimensions[get_column_letter(int(column_number))].hidden = True
    wb.save(path)


def _import_project(
    tmp_path: Path,
    *,
    two_grand_totals: bool = False,
    single_detail: bool = False,
    missing_detail_amount: bool = False,
    hidden_rows: list[int] | None = None,
    hidden_cols: list[int] | None = None,
):
    from jiadun.core.engine import settlement_io
    from jiadun.core.models import project as pm

    info = pm.create_project("复核闸门", tmp_path / "ws")
    info, conn = pm.open_project(Path(info.workspace_path))
    src = tmp_path / "第1期结算.xlsx"
    _make_book(
        src,
        two_grand_totals=two_grand_totals,
        single_detail=single_detail,
        missing_detail_amount=missing_detail_amount,
        hidden_rows=hidden_rows,
        hidden_cols=hidden_cols,
    )
    report = settlement_io.import_settlement_file(
        conn, info.project_id, Path(info.workspace_path), src, direction="upward"
    )
    period_id = conn.execute(
        "SELECT id FROM settlement_periods WHERE project_id=?", (info.project_id,)
    ).fetchone()["id"]
    return info, conn, report, period_id


def test_automatic_import_persists_range_evidence(tmp_path):
    _info, conn, _report, period_id = _import_project(tmp_path)
    try:
        row = conn.execute(
            """SELECT data_row_start, data_row_end, data_range_status,
                      data_range_method, data_range_evidence_json
               FROM table_headers WHERE sheet_id=(
                   SELECT sheet_id FROM line_items WHERE period_id=? LIMIT 1
               )""",
            (period_id,),
        ).fetchone()
        assert row is not None
        assert (row["data_row_start"], row["data_row_end"]) == (2, 5)
        assert row["data_range_status"] == "inferred"
        assert row["data_range_method"] == "last_non_empty_row"
        evidence = json.loads(row["data_range_evidence_json"])
        assert evidence["data_range"] == [2, 5]
        assert evidence["method"] == "last_non_empty_row"
    finally:
        conn.close()


def test_automatic_import_persists_sheet_coverage_proof_and_rows(tmp_path):
    _info, conn, _report, period_id = _import_project(tmp_path)
    try:
        proof = conn.execute(
            """SELECT * FROM sheet_coverage_proofs
               WHERE period_id=? ORDER BY id DESC LIMIT 1""",
            (period_id,),
        ).fetchone()
        assert proof is not None
        assert proof["raw_data_row_count"] == proof["classified_row_count"]
        assert proof["classified_detail_rows"] == 2
        assert proof["business_rows_used"] == 2
        assert proof["ab_row_set_status"] == "same_row_set"
        assert proof["ab_independence_level"] == "shared_extractor"
        assert proof["evidence_id"] is not None
        rows = conn.execute(
            "SELECT COUNT(*) AS c FROM row_classifications WHERE proof_id=?",
            (proof["id"],),
        ).fetchone()
        assert rows["c"] == proof["raw_data_row_count"]
        evidence = conn.execute(
            "SELECT kind, sources_json, steps_json FROM evidence WHERE id=?",
            (proof["evidence_id"],),
        ).fetchone()
        assert evidence["kind"] == "sheet_coverage_proof"
        assert json.loads(evidence["sources_json"])[0]["sheet_id"] == proof["sheet_id"]
        assert json.loads(evidence["steps_json"])[0]["ab_independence_level"] == "shared_extractor"
    finally:
        conn.close()


def test_automatic_import_binds_coverage_proof_to_final_run_contract(tmp_path):
    """整份文件导入完成后，覆盖证明与 Evidence 必须绑定同一 run_id。"""
    from jiadun.core.contracts import run_contract

    info, conn, _report, period_id = _import_project(tmp_path)
    try:
        current = run_contract.get_current_contract(conn, info.project_id)
        assert current is not None
        proof = conn.execute(
            "SELECT run_signature, run_id, evidence_id FROM sheet_coverage_proofs "
            "WHERE period_id=? ORDER BY id DESC LIMIT 1",
            (period_id,),
        ).fetchone()
        assert proof is not None
        assert (proof["run_signature"], proof["run_id"]) == (
            current.signature, current.run_id
        )
        evidence = conn.execute(
            "SELECT run_signature, run_id, scope FROM evidence WHERE id=?",
            (proof["evidence_id"],),
        ).fetchone()
        assert tuple(evidence) == (current.signature, current.run_id, "current")
    finally:
        conn.close()


def test_historical_coverage_proof_is_not_reused_after_contract_change(tmp_path):
    """合同变化后旧证明必须退出当前范围，不能支撑绿色校核。"""
    from jiadun.core.contracts import run_contract
    from jiadun.core.engine import crosscheck
    from jiadun.core.evidence import audit as audit_log

    info, conn, _report, period_id = _import_project(tmp_path)
    try:
        first = run_contract.get_current_contract(conn, info.project_id)
        assert first is not None
        audit_log.record_audit(
            conn, info.project_id, "复核人", "confirm", f"period:{period_id}",
            {"status": "current"}, {"status": "recheck"}, "触发新运行合同的复核确认",
        )
        second = run_contract.ensure_run_contract(conn, info.project_id)
        assert second.run_id != first.run_id
        result = crosscheck.check_period(conn, period_id)
        assert result.coverage_proof_status == "unproven"
        assert result.verification_level == "insufficient"
        assert any("逐行覆盖证明未完整" in note for note in result.notes)
    finally:
        conn.close()


def test_unproven_range_blocks_green_even_when_a_b_c_match(tmp_path):
    from jiadun.core.engine import crosscheck

    _info, conn, _report, period_id = _import_project(tmp_path)
    try:
        header = conn.execute(
            """SELECT sheet_id, header_row_lo, header_row_hi, col_map_json,
                      confidence, needs_review
                 FROM table_headers WHERE sheet_id=(
                     SELECT rs.id FROM raw_sheets rs WHERE rs.period_id=? LIMIT 1
                 ) ORDER BY id DESC LIMIT 1""",
            (period_id,),
        ).fetchone()
        assert header is not None
        conn.execute("DELETE FROM table_headers WHERE sheet_id=?", (header["sheet_id"],))
        conn.execute(
            """INSERT INTO table_headers(
                   sheet_id, header_row_lo, header_row_hi, col_map_json,
                   confidence, needs_review, data_range_status, data_range_method)
               VALUES (?, ?, ?, ?, ?, ?, 'unproven', 'unknown')""",
            (
                header["sheet_id"], header["header_row_lo"], header["header_row_hi"],
                header["col_map_json"], header["confidence"], header["needs_review"],
            ),
        )
        result = crosscheck.check_period(conn, period_id)
        assert result.status == "match"
        assert result.control_status == "match"
        assert result.range_unproven_sheets == 1
        assert result.verification_level == "insufficient"
        assert any("取数范围完整性无法证明" in note for note in result.notes)
    finally:
        conn.close()


def test_partial_sheet_coverage_blocks_green_even_when_a_b_c_match(tmp_path):
    """A/B 同时沿用同一错误行集时，覆盖证明缺口仍必须 fail-closed。"""
    from jiadun.core.engine import crosscheck
    from jiadun.core.parsing import coverage_proof

    _info, conn, _report, period_id = _import_project(tmp_path)
    try:
        original = conn.execute(
            """SELECT * FROM sheet_coverage_proofs
               WHERE period_id=? ORDER BY id DESC LIMIT 1""",
            (period_id,),
        ).fetchone()
        assert original is not None
        partial = coverage_proof.SheetCoverageProof(
            raw_row_start=original["raw_row_start"],
            raw_row_end=original["raw_row_end"],
            raw_col_start=original["raw_col_start"],
            raw_col_end=original["raw_col_end"],
            raw_data_row_count=original["raw_data_row_count"],
            classified_row_count=original["classified_row_count"],
            counts={
                "detail": original["classified_detail_rows"],
                "subtotal": original["excluded_subtotal_rows"],
                "title": original["excluded_title_rows"],
                "note": original["excluded_note_rows"],
                "blank": original["excluded_blank_rows"],
                "tail_note": original["excluded_tail_note_rows"],
                "orphan_numeric": original["excluded_orphan_numeric_rows"],
                "parse_failed": original["excluded_parse_failed_rows"],
                "pending_review": max(0, original["pending_rows"] - original["excluded_tail_note_rows"]),
                "unrecognized": original["unrecognized_rows"],
            },
            raw_amount_total=original["raw_amount_total"],
            detail_amount_total=original["detail_amount_total"],
            business_rows_used=original["business_rows_used"],
            proof_status="partial",
            proof_reason=["synthetic_parser_error"],
            ab_row_set_status=original["ab_row_set_status"],
            ab_row_set_hash=original["ab_row_set_hash"],
            ab_independence_level=original["ab_independence_level"],
        )
        new_proof_id = coverage_proof.persist_sheet_coverage_proof(
            conn,
            project_id=original["project_id"],
            file_id=original["file_id"],
            batch_id=original["batch_id"],
            sheet_id=original["sheet_id"],
            period_id=original["period_id"],
            direction=original["direction"],
            proof=partial,
            rows=[],
            run_signature=original["run_signature"],
            run_id=original["run_id"],
            evidence_id=original["evidence_id"],
        )
        assert new_proof_id > original["id"]
        result = crosscheck.check_period(conn, period_id)
        assert result.status == "match"
        assert result.control_status == "match"
        assert result.path_a_total == result.path_b_total == Decimal("500")
        assert result.coverage_proof_status == "partial"
        assert result.ab_row_set_status == "same_row_set"
        assert result.ab_independence_level == "shared_extractor"
        assert result.verification_level == "insufficient"
        assert any("逐行覆盖证明未完整" in note for note in result.notes)
        assert any("非独立" in note for note in result.notes)
    finally:
        conn.close()


def test_c_control_source_retains_original_cell_and_evidence(tmp_path):
    from jiadun.core.engine import crosscheck

    _info, conn, _report, period_id = _import_project(tmp_path)
    try:
        result = crosscheck.check_period(conn, period_id)
        assert result.c_control_source is not None
        assert result.c_control_source["sheet_id"] is not None
        assert result.c_control_source["row"] == 5
        assert result.c_control_source["raw_value"] == "500"
        assert result.c_control_source["selection_rule"]
        assert result.c_control_evidence_id is not None
        evidence = json.loads(crosscheck.make_evidence(1, result))
        c_step = next(step for step in evidence["steps"] if step["step"] == "C 控制值来源")
        assert c_step["result"]["source_evidence_id"] == result.c_control_evidence_id
    finally:
        conn.close()


def test_missing_raw_amount_cannot_be_sufficient_when_control_matches(tmp_path):
    from jiadun.core.engine import crosscheck

    _info, conn, _report, period_id = _import_project(
        tmp_path, single_detail=True, missing_detail_amount=True
    )
    try:
        conn.execute(
            "UPDATE line_items SET amount=NULL, quantity='2', unit_price='100' "
            "WHERE period_id=? AND json_extract(flags_json, '$.subtotal') IS NULL",
            (period_id,),
        )
        result = crosscheck.check_period(conn, period_id)
        assert result.path_a_total == Decimal("200")
        assert result.path_b_total == Decimal("200")
        assert result.raw_subtotal == Decimal("200")
        assert result.diff_ab == Decimal("0")
        assert result.control_status == "match"
        assert result.verification_level == "insufficient"
        assert result.amount_source == "calculated"
        assert any("原始金额缺失" in note for note in result.notes)
    finally:
        conn.close()


@pytest.mark.parametrize(("column", "value"), [
    ("hidden_rows_json", "[3]"),
    ("hidden_cols_json", "[5]"),
])
def test_hidden_visibility_blocks_inferred_range(tmp_path, column, value):
    """自动推断不能把隐藏行/列当作已证明完整；必须人工确认范围。"""
    from jiadun.core.engine import crosscheck

    import_kwargs = {"hidden_rows": [3]} if column == "hidden_rows_json" else {"hidden_cols": [5]}
    _info, conn, _report, period_id = _import_project(tmp_path, **import_kwargs)
    try:
        result = crosscheck.check_period(conn, period_id)
        assert result.range_unproven_sheets == 1
        assert result.verification_level == "insufficient"
        assert any("取数范围完整性无法证明" in note for note in result.notes)
    finally:
        conn.close()


def test_persisted_overall_status_is_not_match_when_range_unproven(tmp_path):
    """旧 period_totals 读取面也不能把证据不足写成整体 match。"""
    from jiadun.core.engine import aggregate, crosscheck

    info, conn, _report, period_id = _import_project(tmp_path, hidden_rows=[3])
    try:
        aggs = aggregate.aggregate_project(conn, info.project_id, direction="upward")
        aggregate.persist_period_totals(conn, info.project_id, aggs)
        result = crosscheck.run_crosscheck(
            conn, info.project_id, [1], direction="upward"
        )[0]
        row = conn.execute(
            """SELECT cross_check_status, verification_level
               FROM period_totals WHERE period_id=? LIMIT 1""",
            (period_id,),
        ).fetchone()
        assert result.verification_level == "insufficient"
        assert row["verification_level"] == "insufficient"
        assert row["cross_check_status"] != "match"
    finally:
        conn.close()


def test_confirmed_non_settlement_sheet_is_not_pending(tmp_path):
    from jiadun.core.engine import crosscheck, settlement_io

    info, conn, _report, period_id = _import_project(tmp_path)
    try:
        batch_id = conn.execute(
            "SELECT id FROM parse_batches WHERE file_id=(SELECT source_file_id FROM settlement_periods WHERE id=?)",
            (period_id,),
        ).fetchone()["id"]
        sheet_id = conn.execute(
            """INSERT INTO raw_sheets(batch_id, sheet_index, sheet_name, n_rows, n_cols)
               VALUES (?,?,?,?,?)""",
            (batch_id, 99, "汇总控制表", 3, 3),
        ).lastrowid
        assert settlement_io.pending_sheet_count(conn, info.project_id) == 1
        before = crosscheck.check_period(conn, period_id)
        assert before.pending_sheets == 1

        settlement_io.confirm_sheet_non_settlement_role(
            conn, info.project_id, int(sheet_id), "复核人", "settlement_summary",
            "人工核对：该表为控制汇总，仅存证，不进入明细结算模型",
        )
        assert settlement_io.pending_sheet_count(conn, info.project_id) == 0
        after = crosscheck.check_period(conn, period_id)
        assert after.pending_sheets == 0
        assert after.verification_level == "sufficient"
        assert conn.execute(
            "SELECT COUNT(*) c FROM evidence WHERE project_id=? AND kind='sheet_role_confirmed'",
            (info.project_id,),
        ).fetchone()["c"] == 1
    finally:
        conn.close()


def test_non_settlement_confirmation_historicalizes_old_crosscheck_evidence(tmp_path):
    """人工门控改变运行范围时，旧校核 Evidence 也必须退出 current。"""
    from jiadun.core.engine import aggregate, crosscheck, settlement_io

    info, conn, _report, period_id = _import_project(tmp_path)
    try:
        aggregate.persist_period_totals(
            conn,
            info.project_id,
            aggregate.aggregate_project(conn, info.project_id, direction="upward"),
        )
        crosscheck.run_crosscheck(
            conn, info.project_id, [1], direction="upward"
        )[0]
        old_evidence_id = int(conn.execute(
            "SELECT evidence_id FROM crosscheck_results WHERE period_id=?",
            (period_id,),
        ).fetchone()["evidence_id"])
        old_proof_evidence_id = int(conn.execute(
            "SELECT evidence_id FROM sheet_coverage_proofs WHERE period_id=? ORDER BY id DESC LIMIT 1",
            (period_id,),
        ).fetchone()["evidence_id"])
        old_contract = conn.execute(
            """SELECT run_id, signature FROM run_contracts
               WHERE project_id=? AND invalidated_at IS NULL""",
            (info.project_id,),
        ).fetchone()

        batch_id = conn.execute(
            """SELECT pb.id FROM parse_batches pb
               JOIN source_files sf ON sf.id=pb.file_id
               WHERE sf.project_id=? ORDER BY pb.id DESC LIMIT 1""",
            (info.project_id,),
        ).fetchone()["id"]
        sheet_id = conn.execute(
            """INSERT INTO raw_sheets(batch_id, sheet_index, sheet_name, n_rows, n_cols)
               VALUES (?,?,?,?,?)""",
            (batch_id, 88, "汇总控制表", 3, 3),
        ).lastrowid

        settlement_io.confirm_sheet_non_settlement_role(
            conn,
            info.project_id,
            int(sheet_id),
            "复核人",
            "settlement_summary",
            "人工核对：该表只作为控制证据，不进入明细结算模型",
        )

        evidence = conn.execute(
            """SELECT scope, historical_reason, run_id, run_signature
               FROM evidence WHERE id=?""",
            (old_evidence_id,),
        ).fetchone()
        assert evidence["scope"] == "historical"
        assert evidence["historical_reason"]
        assert (evidence["run_id"], evidence["run_signature"]) == (
            old_contract["run_id"], old_contract["signature"]
        )
        event = conn.execute(
            """SELECT before_scope, after_scope, event_type
               FROM evidence_events WHERE evidence_id=? ORDER BY id DESC LIMIT 1""",
            (old_evidence_id,),
        ).fetchone()
        assert tuple(event) == ("current", "historical", "invalidated")
        proof_evidence = conn.execute(
            "SELECT scope, historical_reason FROM evidence WHERE id=?",
            (old_proof_evidence_id,),
        ).fetchone()
        assert proof_evidence["scope"] == "historical"
        assert proof_evidence["historical_reason"]
    finally:
        conn.close()


def test_crosscheck_result_survives_reopen(tmp_path):
    from jiadun.core.engine import aggregate, crosscheck
    from jiadun.core.models import project as pm

    info, conn, _report, period_id = _import_project(tmp_path)
    workspace = Path(info.workspace_path)
    try:
        aggregate.persist_period_totals(
            conn,
            info.project_id,
            aggregate.aggregate_project(conn, info.project_id, direction="upward"),
        )
        result = crosscheck.run_crosscheck(
            conn, info.project_id, [1], direction="upward"
        )[0]
        stored = conn.execute(
            """SELECT verification_level, status, detail_rows,
                      excluded_subtotal_rows, pending_sheets, range_unproven_sheets,
                      classified_detail_rows, business_rows_used,
                      coverage_proof_status, ab_row_set_status,
                      ab_independence_level, c_control_evidence_id, run_id, evidence_id
                      FROM crosscheck_results WHERE period_id=?""",
            (period_id,),
        ).fetchone()
        assert stored is not None
        assert stored["verification_level"] == result.verification_level
        assert stored["status"] == "match"
        assert stored["detail_rows"] == result.detail_rows
        assert stored["excluded_subtotal_rows"] == result.excluded_subtotal_rows
        assert stored["pending_sheets"] == result.pending_sheets == 0
        assert stored["range_unproven_sheets"] == result.range_unproven_sheets == 0
        assert stored["classified_detail_rows"] == result.classified_detail_rows == 2
        assert stored["business_rows_used"] == result.business_rows_used == 2
        assert stored["coverage_proof_status"] == "complete"
        assert stored["ab_row_set_status"] == "same_row_set"
        assert stored["ab_independence_level"] == "shared_extractor"
        assert stored["c_control_evidence_id"] == result.c_control_evidence_id
        active = conn.execute(
            "SELECT run_id FROM run_contracts WHERE project_id=? AND invalidated_at IS NULL",
            (info.project_id,),
        ).fetchone()
        assert active is not None
        assert stored["run_id"] == active["run_id"]
        assert stored["evidence_id"] is not None
        evidence_row = conn.execute(
            "SELECT run_id, scope, historical_reason FROM evidence WHERE id=?",
            (stored["evidence_id"],)
        ).fetchone()
        assert evidence_row["run_id"] == active["run_id"]
        assert evidence_row["scope"] == "current"
        assert evidence_row["historical_reason"] is None
    finally:
        conn.close()

    _reopened, reopened = pm.open_project(workspace)
    try:
        row = reopened.execute(
            "SELECT verification_level, evidence_id FROM crosscheck_results WHERE period_id=?",
            (period_id,),
        ).fetchone()
        assert row is not None
        assert row["verification_level"] == result.verification_level
        assert row["evidence_id"] is not None
    finally:
        reopened.close()


def test_period_totals_carries_run_verification_level(tmp_path):
    from jiadun.core.engine import aggregate, crosscheck

    info, conn, _report, period_id = _import_project(tmp_path)
    try:
        aggs = aggregate.aggregate_project(conn, info.project_id, direction="upward")
        aggregate.persist_period_totals(conn, info.project_id, aggs)
        result = crosscheck.run_crosscheck(
            conn, info.project_id, [1], direction="upward"
        )[0]
        row = conn.execute(
            "SELECT cross_check_status, verification_level FROM period_totals WHERE period_id=? LIMIT 1",
            (period_id,),
        ).fetchone()
        assert row is not None
        assert row["verification_level"] == result.verification_level
        if result.verification_level == "sufficient":
            assert row["cross_check_status"] == "match"
        else:
            assert row["cross_check_status"] != "match"
    finally:
        conn.close()


def test_multiple_grand_totals_make_control_unavailable(tmp_path):
    from jiadun.core.engine import crosscheck

    _info, conn, _report, period_id = _import_project(tmp_path, two_grand_totals=True)
    try:
        result = crosscheck.check_period(conn, period_id)
        assert result.control_status == "not_available"
        assert result.verification_level == "insufficient"
        assert any("控制值不唯一" in note for note in result.notes)
    finally:
        conn.close()


def test_reaggregate_invalidates_persisted_crosscheck_result(tmp_path):
    from jiadun.core.engine import aggregate, crosscheck

    info, conn, _report, period_id = _import_project(tmp_path)
    try:
        aggregate.persist_period_totals(
            conn,
            info.project_id,
            aggregate.aggregate_project(conn, info.project_id, direction="upward"),
        )
        crosscheck.run_crosscheck(conn, info.project_id, [1], direction="upward")
        assert conn.execute(
            "SELECT COUNT(*) c FROM crosscheck_results WHERE period_id=?", (period_id,)
        ).fetchone()["c"] == 1
        aggs = aggregate.aggregate_project(conn, info.project_id, direction="upward")
        aggregate.persist_period_totals(conn, info.project_id, aggs)
        assert conn.execute(
            "SELECT COUNT(*) c FROM crosscheck_results WHERE period_id=?", (period_id,)
        ).fetchone()["c"] == 0
    finally:
        conn.close()


def test_empty_match_reason_does_not_partially_write(tmp_path):
    from jiadun.core.evidence.audit import AuditReasonRequiredError
    from jiadun.core.matching import matching

    info, conn, _report, _period_id = _import_project(tmp_path)
    try:
        groups = matching.match_items(conn, info.project_id, direction="upward")
        matching.save_matches(conn, info.project_id, groups)
        match_id = conn.execute(
            "SELECT id FROM matches WHERE project_id=? LIMIT 1", (info.project_id,)
        ).fetchone()["id"]
        with pytest.raises(AuditReasonRequiredError):
            matching.confirm_match(conn, info.project_id, match_id, "user", " ")
        row = conn.execute("SELECT status FROM matches WHERE id=?", (match_id,)).fetchone()
        assert row["status"] == "pending"
        assert conn.execute(
            "SELECT COUNT(*) c FROM evidence WHERE project_id=? AND kind='match_confirmation'",
            (info.project_id,),
        ).fetchone()["c"] == 0
        assert conn.execute(
            "SELECT COUNT(*) c FROM audit_log WHERE project_id=? AND action='confirm_match'",
            (info.project_id,),
        ).fetchone()["c"] == 0
    finally:
        conn.close()


def test_business_fallbacks_hide_unknown_codes():
    from jiadun.core.anomalies.rules import _dir_label
    from jiadun.ui.labels import rule_zh

    assert _dir_label("upward") == "[对上结算] "
    assert _dir_label("downward") == "[对下结算] "
    assert rule_zh("new_rule_v99") == "其他审核问题"

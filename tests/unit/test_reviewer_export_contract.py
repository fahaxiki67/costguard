import pytest
from openpyxl import Workbook

from jiadun.core.db import migrations
from jiadun.core.export import excel_export


@pytest.fixture()
def project_db(tmp_path):
    db_path = tmp_path / "project.db"
    migrations.migrate(db_path, tmp_path / "backups")
    conn = migrations.connect(db_path)
    with conn:
        project_id = conn.execute(
            "INSERT INTO projects(name, schema_version, workspace_path, created_at)"
            " VALUES (?, ?, ?, ?)",
            ("导出维度回归", migrations.LATEST_SCHEMA_VERSION, str(tmp_path), "2026"),
        ).lastrowid
        period_ids = []
        for period_no, direction in ((1, "upward"), (2, "upward"), (3, "downward")):
            period_ids.append(
                conn.execute(
                    "INSERT INTO settlement_periods(project_id, period_no, title, direction)"
                    " VALUES (?, ?, ?, ?)",
                    (project_id, period_no, f"第{period_no}期", direction),
                ).lastrowid
            )
    yield conn, project_id, period_ids
    conn.close()


def _line(conn, period_id, *, code, name, feature, unit, quantity, amount):
    with conn:
        conn.execute(
            "INSERT INTO line_items(period_id, code, name, feature, unit, quantity, amount)"
            " VALUES (?, ?, ?, ?, ?, ?, ?)",
            (period_id, code, name, feature, unit, quantity, amount),
        )


def test_export_keeps_feature_and_unit_visible_when_keys_are_split(project_db):
    conn, project_id, (up_1, up_2, down_1) = project_db
    _line(conn, up_1, code="S1", name="商砼", feature="C25垫层", unit="m²",
          quantity="10", amount="100")
    _line(conn, up_2, code="S1", name="商砼", feature="C25垫层", unit="m³",
          quantity="2", amount="200")
    _line(conn, down_1, code="S2", name="商砼", feature="C30梁", unit="m³",
          quantity="3", amount="300")
    _line(conn, up_1, code="S2", name="商砼", feature="C30梁", unit="m²",
          quantity="10", amount="100")

    wb = Workbook()
    wb.remove(wb.active)
    sheet_name = excel_export.export_settlement_summary(
        conn, project_id, wb, direction="upward"
    )
    summary = wb[sheet_name]
    headers = [cell.value for cell in summary[1]]
    assert "项目特征" in headers
    unit_col = headers.index("单位")
    feature_col = headers.index("项目特征")
    s1_rows = [
        row for row in summary.iter_rows(min_row=2, values_only=True)
        if row[0] == "S1"
    ]
    assert {row[unit_col] for row in s1_rows} == {"m2", "m3"}
    assert {row[feature_col] for row in s1_rows} == {"C25垫层"}

    wb2 = Workbook()
    wb2.remove(wb2.active)
    excel_export.export_updown_comparison(conn, project_id, wb2)
    comparison = wb2["对上对下对比表"]
    comparison_headers = [cell.value for cell in comparison[1]]
    assert "对上单位" in comparison_headers
    assert "对下单位" in comparison_headers
    note_col = comparison_headers.index("口径说明")
    diff_col = comparison_headers.index("金额差异(公式)")
    up_unit_col = comparison_headers.index("对上单位")
    down_unit_col = comparison_headers.index("对下单位")
    s2_rows = [
        row for row in comparison.iter_rows(min_row=2, values_only=True)
        if row[0] == "S2"
    ]
    assert {row[up_unit_col] for row in s2_rows} == {"m2", ""}
    assert {row[down_unit_col] for row in s2_rows} == {"", "m3"}
    assert all(row[diff_col] is None for row in s2_rows)
    assert all("不可比" in str(row[note_col]) for row in s2_rows)

    wb3 = Workbook()
    wb3.remove(wb3.active)
    excel_export.export_diff_sheets(conn, project_id, wb3)
    diff = wb3["金额差异表"]
    diff_headers = [cell.value for cell in diff[1]]
    assert "项目特征" in diff_headers
    assert "单位" in diff_headers
    diff_unit_col = diff_headers.index("单位")
    s1_diff_rows = [
        row for row in diff.iter_rows(min_row=2, values_only=True)
        if row[1] == "S1"
    ]
    assert {row[diff_unit_col] for row in s1_diff_rows} == {"m2", "m3"}

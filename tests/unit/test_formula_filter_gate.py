"""P1 公式缓存、筛选和数据区合并元数据的 fail-closed 回归。

这些用例只使用人工构造的脱敏工作簿，验证解析层保留结构事实，导入层不把
未知可见性/不可信公式当成已确认输入，异常规则能够给出可定位的证据缺口。
"""
from __future__ import annotations

import json
import sys
import zipfile
from pathlib import Path
from types import SimpleNamespace
from xml.etree import ElementTree

import openpyxl
import pytest
from openpyxl.worksheet.filters import AutoFilter
from openpyxl.worksheet.table import Table

from jiadun.core.anomalies import engine, rules
from jiadun.core.db import migrations
from jiadun.core.engine import crosscheck, settlement_io
from jiadun.core.models import project as project_model
from jiadun.core.parsing import excel_parser


def _make_workbook(
    path: Path,
    *,
    amount_formula: str = "=C2*D2",
    control_formula: str | None = None,
    with_filter: bool = False,
    with_table_filter: bool = False,
    merged_data: bool = False,
) -> None:
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "第1期明细"
    for col, value in enumerate(("项目编码", "项目名称", "工程量", "综合单价", "合价"), 1):
        ws.cell(1, col, value)
    ws.append(("A1", "项目A", 2, 100, amount_formula))
    ws.append(("A2", "项目B", 1, 300, 300))
    ws.append((None, "合计", None, None, 500))
    if control_formula is not None:
        ws["E4"] = control_formula
    if merged_data:
        # 第二条明细的名称保持为空。旧实现会自动把 B2 的值复制到 B3，
        # 造成错误的名称/匹配；正确行为应保留空值并进入人工复核。
        ws.merge_cells("B2:B3")
    if with_filter:
        ws.auto_filter.ref = "A1:E4"
        ws.auto_filter.add_filter_column(1, ["A1"])
    if with_table_filter:
        table = Table(displayName="ItemsTable", ref="A1:E4")
        table.autoFilter = AutoFilter(ref="A1:E4")
        table.autoFilter.add_filter_column(1, ["A1"])
        ws.add_table(table)
    wb.save(path)


def _set_formula_cache(path: Path, values: dict[str, str]) -> None:
    """为测试工作簿写入最小 OOXML 缓存值，模拟 Office 已重算结果。"""
    tmp = path.with_suffix(".cached.xlsx")
    namespace = "{http://schemas.openxmlformats.org/spreadsheetml/2006/main}"
    with zipfile.ZipFile(path, "r") as source, zipfile.ZipFile(tmp, "w") as target:
        for item in source.infolist():
            payload = source.read(item.filename)
            if item.filename == "xl/worksheets/sheet1.xml":
                root = ElementTree.fromstring(payload)
                for cell in root.iter(f"{namespace}c"):
                    coordinate = cell.attrib.get("r")
                    if coordinate not in values:
                        continue
                    value = cell.find(f"{namespace}v")
                    if value is None:
                        value = ElementTree.SubElement(cell, f"{namespace}v")
                    value.text = values[coordinate]
                payload = ElementTree.tostring(root, encoding="utf-8", xml_declaration=True)
            target.writestr(item, payload)
    tmp.replace(path)


def _make_duplicate_header_workbook(path: Path) -> None:
    """构造“首段明细 + 跨页重复表头 + 控制行”的脱敏反例。"""
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "第1期明细"
    headers = ("项目编码", "项目名称", "工程量", "综合单价", "合价")
    for col, value in enumerate(headers, 1):
        ws.cell(1, col, value)
    ws.append(("A1", "项目A", 1, 100, 100))
    for col, value in enumerate(headers, 1):
        ws.cell(3, col, value)
    ws.append(("A2", "项目B", 2, 100, 200))
    ws.append((None, "小计", None, None, 300))
    ws.append((None, "合计", None, None, 300))
    wb.save(path)


@pytest.fixture()
def imported_project(tmp_path: Path):
    info = project_model.create_project("公式筛选闸门", tmp_path / "ws")
    info, conn = project_model.open_project(Path(info.workspace_path))
    yield info, conn, Path(info.workspace_path)
    conn.close()


def _import_fixture(info, conn, project_dir: Path, src: Path):
    report = settlement_io.import_settlement_file(
        conn, info.project_id, project_dir, src, direction="upward"
    )
    sheet = conn.execute(
        "SELECT * FROM raw_sheets WHERE batch_id=? ORDER BY id LIMIT 1",
        (report.batch_id,),
    ).fetchone()
    period = conn.execute(
        "SELECT id FROM settlement_periods WHERE project_id=? ORDER BY id LIMIT 1",
        (info.project_id,),
    ).fetchone()
    return report, sheet, (int(period["id"]) if period else None)


def test_parser_records_formula_cache_state_and_filter_table_metadata(tmp_path: Path):
    src = tmp_path / "公式筛选.xlsx"
    _make_workbook(src, with_filter=True, with_table_filter=True)

    result = excel_parser.parse_file(src, "xlsx")

    assert result.status == "ok"
    sheet = result.sheets[0]
    assert sheet.auto_filter_ref
    assert sheet.filter_state == "unknown"
    assert sheet.filter_conditions
    assert any(item["source"] == "worksheet.auto_filter" for item in sheet.filter_conditions)
    assert any(item["source"] == "table.auto_filter" for item in sheet.filter_conditions)
    assert sheet.table_ranges and sheet.table_ranges[0]["name"] == "ItemsTable"
    formula = next(cell for cell in sheet.cells if cell.is_formula)
    assert formula.formula_cache_status == "missing"
    assert sheet.formula_metadata["unverified_count"] >= 1
    assert sheet.formula_metadata["uncached_count"] >= 1


def test_import_persists_filter_and_formula_risk_and_keeps_sheet_pending(imported_project, tmp_path):
    info, conn, project_dir = imported_project
    src = tmp_path / "筛选公式.xlsx"
    _make_workbook(src, with_filter=True)

    report, sheet, _period_id = _import_fixture(info, conn, project_dir, src)

    assert report.status == "partial"
    assert report.needs_manual_review is True
    assert sheet["filter_state"] == "unknown"
    assert json.loads(sheet["filter_conditions_json"])
    assert json.loads(sheet["formula_metadata_json"])["unverified_count"] >= 1
    assert sheet["sheet_status"] == "pending"
    assert "筛选" in sheet["sheet_status_reason"]


def test_formula_without_cache_is_structural_high_risk_with_source_location(imported_project, tmp_path):
    info, conn, project_dir = imported_project
    src = tmp_path / "无缓存公式.xlsx"
    _make_workbook(src)
    _report, _sheet, period_id = _import_fixture(info, conn, project_dir, src)

    findings = rules.rule_formula_issues(conn, info.project_id)
    formula_findings = [finding for finding in findings if finding.rule_id == "formula_no_cache"]
    assert formula_findings
    finding = formula_findings[0]
    assert finding.severity == "high"
    assert finding.details["blocks_verification"] is True
    assert finding.details["row"] == 2
    assert finding.details["col"] == 5
    assert finding.details["cache_status"] == "missing"
    assert period_id is not None


def test_formula_semantics_mismatch_is_verifiable_for_qty_price_and_control(imported_project, tmp_path):
    info, conn, project_dir = imported_project
    src = tmp_path / "公式语义错误.xlsx"
    _make_workbook(src, amount_formula="=C2+D2", control_formula="=C2+D2")
    _set_formula_cache(src, {"E2": "302", "E4": "3"})
    _report, _sheet, _period_id = _import_fixture(info, conn, project_dir, src)

    findings = rules.rule_formula_semantics(conn, info.project_id)
    mismatch = [finding for finding in findings if finding.rule_id == "formula_semantics_mismatch"]
    control = [
        finding for finding in findings
        if finding.rule_id == "formula_control_semantics_mismatch"
    ]
    assert mismatch
    assert mismatch[0].severity == "high"
    assert mismatch[0].details["expected_fields"] == ["quantity", "unit_price"]
    assert mismatch[0].details["blocks_verification"] is True
    assert control
    assert control[0].severity == "high"
    assert control[0].details["expected_field"] == "amount"


def test_data_area_merge_is_not_automatically_anchor_copied(imported_project, tmp_path):
    info, conn, project_dir = imported_project
    src = tmp_path / "明细合并.xlsx"
    _make_workbook(src, merged_data=True)
    _report, sheet, _period_id = _import_fixture(info, conn, project_dir, src)

    items = conn.execute(
        "SELECT name, flags_json FROM line_items ORDER BY id"
    ).fetchall()
    assert len(items) >= 2
    assert items[1]["name"] != "项目A"
    flags = json.loads(items[1]["flags_json"] or "{}")
    assert flags.get("merge_anchor_copy") is not True
    assert sheet["sheet_status"] == "pending"
    assert "合并" in sheet["sheet_status_reason"]


def test_hidden_rows_or_columns_make_import_pending(imported_project, tmp_path):
    """隐藏行/列不能在导入层显示为已确认，必须留下可见的范围复核入口。"""
    info, conn, project_dir = imported_project
    src = tmp_path / "隐藏行列.xlsx"
    _make_workbook(src)
    wb = openpyxl.load_workbook(src)
    ws = wb[wb.sheetnames[0]]
    ws.row_dimensions[3].hidden = True
    ws.column_dimensions["E"].hidden = True
    wb.save(src)

    report, sheet, _period_id = _import_fixture(info, conn, project_dir, src)

    assert report.status == "partial"
    assert report.needs_manual_review is True
    assert sheet["sheet_status"] == "pending"
    assert "隐藏" in sheet["sheet_status_reason"]


def test_xls_parser_exposes_capability_limitations_as_partial(monkeypatch, tmp_path):
    """xlrd 无法保留公式/合并拓扑时，解析结果必须显式降级而非伪装完整。"""

    class FakeCell:
        def __init__(self, value):
            self.ctype = 1
            self.value = value

    class FakeSheet:
        name = "第1期明细"
        nrows = 2
        ncols = 1
        rowinfo_map = {}
        colinfo_map = {}

        def cell(self, row, col):
            return FakeCell("x" if row == 0 else "1")

    class FakeBook:
        nsheets = 1

        def sheet_by_index(self, _index):
            return FakeSheet()

        def release_resources(self):
            return None

    fake_xlrd = SimpleNamespace(open_workbook=lambda _path: FakeBook())
    monkeypatch.setitem(sys.modules, "xlrd", fake_xlrd)

    result = excel_parser.parse_xls(tmp_path / "sample.xls")

    assert result.status == "partial"
    assert result.stats["capability_limited"] is True
    assert result.stats["limitations"]
    assert result.sheets[0].formula_metadata["capability_limited"] is True
    assert result.sheets[0].formula_metadata["limitations"]


def test_partial_parser_import_never_becomes_confirmed(monkeypatch, imported_project, tmp_path):
    """解析器部分成功仍可预览，但不能让工作表/期次进入绿色确认状态。"""
    info, conn, project_dir = imported_project
    src = tmp_path / "部分解析.xls"
    src.write_bytes(b"synthetic xls placeholder")
    from jiadun.core.parsing.excel_parser import CellRecord, ParseResult, SheetRecord

    values = [
        (1, 1, "项目编码"), (1, 2, "项目名称"), (1, 3, "工程量"),
        (1, 4, "综合单价"), (1, 5, "合价"),
        (2, 1, "A1"), (2, 2, "项目A"), (2, 3, "1"), (2, 4, "100"), (2, 5, "100"),
        (3, 2, "合计"), (3, 5, "100"),
    ]
    fake_sheet = SheetRecord(
        sheet_index=0,
        sheet_name="第1期明细",
        n_rows=3,
        n_cols=5,
        cells=[
            CellRecord(row=row, col=col, raw_value=value, cached_value=value,
                       is_formula=False, is_number_stored_as_text=False, num_fmt="")
            for row, col, value in values
        ],
        formula_metadata={
            "capability_limited": True,
            "limitations": ["xls: no formula cache/merged-cell topology"],
        },
    )
    fake_result = ParseResult(
        parser="xlrd", status="partial", sheets=[fake_sheet],
        stats={"capability_limited": True, "limitations": ["xls limitation"]},
    )
    monkeypatch.setattr(excel_parser, "parse_file", lambda *_args, **_kwargs: fake_result)

    report, sheet, _period_id = _import_fixture(info, conn, project_dir, src)

    assert report.status == "partial"
    assert report.needs_manual_review is True
    assert sheet["sheet_status"] == "pending"
    assert "xls" in sheet["sheet_status_reason"]


def test_pending_structural_sheet_can_be_reconfirmed_with_explicit_range(imported_project, tmp_path):
    """结构风险页可在人工给出范围后重放抽取，且不会追加重复明细。"""
    info, conn, project_dir = imported_project
    src = tmp_path / "隐藏行人工确认.xlsx"
    _make_workbook(src, amount_formula=200)
    wb = openpyxl.load_workbook(src)
    wb[wb.sheetnames[0]].row_dimensions[3].hidden = True
    wb.save(src)
    report, sheet, _period_id = _import_fixture(info, conn, project_dir, src)
    sheet_id = int(sheet["id"])
    before = conn.execute(
        "SELECT COUNT(*) AS c FROM line_items WHERE sheet_id=?", (sheet_id,)
    ).fetchone()["c"]

    n = settlement_io.confirm_sheet_role_and_extract(
        conn,
        info.project_id,
        sheet_id,
        actor="复核人",
        reason="人工确认隐藏行仍属于本期明细，明确纳入完整数据范围",
        direction="upward",
        confirmed_col_map={
            "code": 1, "name": 2, "quantity": 3, "unit_price": 4, "amount": 5,
        },
        confirmed_header_range=(1, 1),
        confirmed_data_range=(2, 4),
    )

    assert n == before == 3
    assert conn.execute(
        "SELECT sheet_status FROM raw_sheets WHERE id=?", (sheet_id,)
    ).fetchone()["sheet_status"] == "confirmed"
    after = conn.execute(
        "SELECT COUNT(*) AS c FROM line_items WHERE sheet_id=?", (sheet_id,)
    ).fetchone()["c"]
    assert after == n == before

    # 人工明确了隐藏行的纳入范围后，覆盖证明应吸收这一决策；不能既把
    # Sheet 标成已确认，又让当前覆盖证明继续停留在 unproven。
    from jiadun.core.engine import crosscheck

    result = crosscheck.check_period(conn, _period_id)
    assert result.coverage_proof_status == "complete", result.notes
    assert result.range_unproven_sheets == 0, result.notes
    assert result.verification_level == "sufficient", result.notes
    proof = conn.execute(
        "SELECT proof_status, proof_reason FROM sheet_coverage_proofs "
        "WHERE period_id=? ORDER BY id DESC LIMIT 1",
        (_period_id,),
    ).fetchone()
    assert proof["proof_status"] == "complete"
    assert "hidden_rows_or_columns_manually_confirmed" in json.loads(
        proof["proof_reason"]
    )
    header_evidence = conn.execute(
        "SELECT data_range_evidence_json FROM table_headers WHERE sheet_id=?",
        (sheet_id,),
    ).fetchone()
    assert json.loads(header_evidence["data_range_evidence_json"])["actor"] == "复核人"


def test_latest_schema_exposes_structural_gate_metadata(tmp_path: Path):
    db = tmp_path / "project.db"
    migrations.migrate(db, tmp_path / "backups")
    conn = migrations.connect(db)
    try:
        raw_sheet_columns = {row[1] for row in conn.execute("PRAGMA table_info(raw_sheets)")}
        raw_cell_columns = {row[1] for row in conn.execute("PRAGMA table_info(raw_cells)")}
        assert {
            "auto_filter_ref", "filter_conditions_json", "table_ranges_json",
            "filter_state", "formula_metadata_json",
        } <= raw_sheet_columns
        assert "formula_cache_status" in raw_cell_columns
    finally:
        conn.close()


def test_engine_keeps_formula_risk_as_a_recorded_finding(imported_project, tmp_path):
    info, conn, project_dir = imported_project
    src = tmp_path / "公式异常引擎.xlsx"
    _make_workbook(src)
    _import_fixture(info, conn, project_dir, src)

    findings = engine.run_anomalies(conn, info.project_id)
    formula_findings = [finding for finding in findings if finding.rule_id == "formula_no_cache"]
    assert formula_findings
    row = conn.execute(
        "SELECT message FROM anomalies WHERE rule_id='formula_no_cache' ORDER BY id DESC LIMIT 1"
    ).fetchone()
    assert row is not None
    assert "无缓存" in row["message"]


def test_formula_cache_cannot_make_crosscheck_sufficient(imported_project, tmp_path):
    """即使缓存值与数量×单价相等，公式仍未由本程序独立重算。"""
    info, conn, project_dir = imported_project
    src = tmp_path / "缓存看似一致.xlsx"
    _make_workbook(src, amount_formula="=C2*D2", control_formula="=SUM(E2:E3)")
    _set_formula_cache(src, {"E2": "200", "E4": "500"})
    _report, _sheet, period_id = _import_fixture(info, conn, project_dir, src)

    assert period_id is not None
    result = crosscheck.check_period(conn, period_id)

    assert result.path_a_total == result.path_b_total
    assert result.control_status == "match"
    assert result.verification_level == "insufficient"
    assert result.range_unproven_sheets >= 1
    assert any("公式" in note for note in result.notes)


def test_duplicate_header_cannot_be_confirmed_as_complete_range(imported_project, tmp_path):
    """人工确认列映射也不能抹掉跨页重复表头造成的范围缺口。"""
    info, conn, project_dir = imported_project
    src = tmp_path / "重复表头范围.xlsx"
    _make_duplicate_header_workbook(src)
    report = settlement_io.import_settlement_file(
        conn, info.project_id, project_dir, src, direction="upward"
    )
    assert report.status == "partial"
    sheet_id = conn.execute(
        "SELECT id FROM raw_sheets WHERE batch_id=?", (report.batch_id,)
    ).fetchone()["id"]
    settlement_io.confirm_sheet_role_and_extract(
        conn,
        info.project_id,
        int(sheet_id),
        actor="复核人",
        reason="人工指定首个表头，但重复表头仍需保留范围风险",
        direction="upward",
        confirmed_col_map={
            "code": 1, "name": 2, "quantity": 3, "unit_price": 4, "amount": 5,
        },
        confirmed_header_range=(1, 1),
    )
    period_id = conn.execute(
        "SELECT period_id FROM raw_sheets WHERE id=?", (sheet_id,)
    ).fetchone()["period_id"]
    result = crosscheck.check_period(conn, int(period_id))

    assert result.verification_level == "insufficient"
    assert result.coverage_proof_status != "complete"
    assert any("重复表头" in note or "覆盖证明" in note for note in result.notes)


@pytest.mark.parametrize(
    ("filename", "kwargs", "expected_note"),
    [
        ("明细合并校核.xlsx", {"merged_data": True}, "合并"),
        ("筛选校核.xlsx", {"with_filter": True}, "筛选"),
    ],
)
def test_structural_risks_cannot_make_crosscheck_sufficient(
    imported_project, tmp_path, filename, kwargs, expected_note
):
    info, conn, project_dir = imported_project
    src = tmp_path / filename
    _make_workbook(src, **kwargs)
    _report, _sheet, period_id = _import_fixture(info, conn, project_dir, src)

    result = crosscheck.check_period(conn, period_id)

    assert result.verification_level == "insufficient"
    assert result.range_unproven_sheets >= 1
    assert any(expected_note in note for note in result.notes)

"""业务类型门控（监督第八/九轮，先红后绿，全部公开合成样例）。

第九轮口径：
- 删除行数护栏（过拟合）：600 行合法大结算必须解析，20 行汇总小表仍挡——判断不按大小；
- 所有 det.needs_review（含歧义/低置信，无论 confidence）一律不写 canonical，
  未经人工确认前文件级 partial；无歧义强表头（needs_review=False）才可自动解析；
- sheet 名/document 语义门控：汇总/核销/台账类 sheet 名即使强表头也需角色确认；
- 人工确认入口 confirm_sheet_role_and_extract：确认后重放语义层抽取；
- 文件只要还有被挡 sheet，报告列 pending role review；只有无歧义 canonical 计自动成功。
"""
import json
from pathlib import Path

import pytest

from costguard.core.db import migrations
from costguard.core.engine import settlement_io
from costguard.core.models import project as pm


def _wb_sheet(ws, header_row, rows, header):
    for c, v in enumerate(header, start=1):
        ws.cell(row=header_row, column=c, value=v)
    for i, row in enumerate(rows, start=header_row + 1):
        for c, v in enumerate(row, start=1):
            if v is not None:
                ws.cell(row=i, column=c, value=v)


def make_strong_settlement(path: Path) -> None:
    """规范结算清单（主线回归：必须解析并写 canonical）。"""
    import openpyxl

    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "第1期"
    _wb_sheet(ws, 1, [
        ("K1", "混凝土垫层", "m3", 10, 465, 4650, "0.09"),
        ("K2", "钢筋制作安装", "t", 5, 4850, 24250, "0.13"),
        ("K3", "平整场地", "m2", 100, 8.5, 850, "0.09"),
    ], ["清单编码", "清单名称", "单位", "工程量", "综合单价", "合价", "税率"])
    wb.save(path)


def make_weak_ledger(path: Path) -> None:
    """弱表头台账（模拟收入合同台账）：字段弱命中、无数量/合价结构。"""
    import openpyxl

    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "台账明细"
    _wb_sheet(ws, 1, [
        (f"HT-{i}", f"某合同{i}", f"对方单位{i}", 1000 + i * 7) for i in range(1, 9)
    ], ["合同编号", "合同名称", "承包人", "合同金额"])
    wb.save(path)


def make_kv_content_page(path: Path) -> None:
    """键值内容页（模拟项目基本信息表）：键值对为主，含个别数字列。"""
    import openpyxl

    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "项目基本信息"
    rows = [
        ("项目名称：", "某工程（脱敏）"),
        ("项目经理：", "张三"),
        ("合同金额：", "12,000,000.00"),
        ("开工日期：", "2025-03-01"),
        ("建筑面积：", "45,000.00"),
        ("结构形式：", "框架"),
    ] * 4
    for r, row in enumerate(rows, start=1):
        ws.cell(row=r, column=1, value=row[0])
        ws.cell(row=r, column=2, value=row[1])
        ws.cell(row=r, column=3, value=r * 13.5)  # 数值列（诱使弱识别）
        ws.cell(row=r, column=4, value=r * 2)
    wb.save(path)


def make_oversized_resource_summary(path: Path) -> None:
    """超长资源汇总（模拟人材机/核销）：规范列结构但 500+ 行、无期次信号。"""
    import openpyxl

    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "资源汇总"
    rows = [(f"R{i:04d}", f"材料{i}", "kg", i % 90 + 1, round(3.5 + i % 20, 2),
             round((i % 90 + 1) * (3.5 + i % 20), 2)) for i in range(1, 520)]
    _wb_sheet(ws, 1, rows,
              ["编号", "名称规格", "单位", "数量", "单价", "合价"])
    wb.save(path)


def make_mixed_oversized(path: Path) -> None:
    """混合文件：一个大表 + 一个小表（文件级护栏应整体挡下）。"""
    import openpyxl

    wb = openpyxl.Workbook()
    ws1 = wb.active
    ws1.title = "汇总大表"
    rows = [(f"R{i:04d}", f"项{i}", "kg", 1, 2, 2) for i in range(1, 520)]
    _wb_sheet(ws1, 1, rows, ["编号", "名称规格", "单位", "数量", "单价", "合价"])
    ws2 = wb.create_sheet("小清单")
    _wb_sheet(ws2, 1, [("K1", "某清单", "m3", 3, 10, 30)],
              ["清单编码", "清单名称", "单位", "工程量", "综合单价", "合价"])
    wb.save(path)


@pytest.fixture()
def db(tmp_path):
    db_path = tmp_path / "project.db"
    migrations.migrate(db_path, tmp_path / "backups")
    conn = migrations.connect(db_path)
    with conn:
        pid = conn.execute(
            "INSERT INTO projects(name, schema_version, workspace_path, created_at)"
            " VALUES ('g',1,'/t','2026')").lastrowid
    yield conn, pid, tmp_path
    conn.close()


def _import(db, name_maker):
    conn, pid, tmp = db
    src = tmp / "in.xlsx"
    name_maker(src)
    _info = pm.ProjectInfo(pid, "g", str(tmp), 1, "2026")
    report = settlement_io.import_settlement_file(conn, pid, tmp, src)
    return conn, pid, report


def _counts(conn):
    return {t: conn.execute(f"SELECT COUNT(*) c FROM {t}").fetchone()["c"]
            for t in ("settlement_periods", "line_items", "period_totals")}


class TestGate:
    def test_strong_settlement_still_parsed(self, db):
        """主线回归：规范结算表必须照常解析并写 canonical，overall=ok。"""
        conn, pid, report = _import(db, make_strong_settlement)
        assert report.status == "ok"
        assert not report.needs_manual_review
        assert any(s.status == "parsed" for s in report.sheets)
        c = _counts(conn)
        assert c["settlement_periods"] == 1 and c["line_items"] == 3

    def test_weak_ledger_gated(self, db):
        """弱表头台账：needs_role_review，零 canonical 写入，overall 非 ok。"""
        conn, pid, report = _import(db, make_weak_ledger)
        assert report.status == "partial"
        assert report.needs_manual_review
        assert all(s.status == "needs_role_review" for s in report.sheets)
        assert _counts(conn) == {"settlement_periods": 0, "line_items": 0,
                                 "period_totals": 0}
        ev = conn.execute(
            "SELECT COUNT(*) c FROM evidence WHERE kind='sheet_role_candidate'").fetchone()["c"]
        assert ev >= 1, "被挡 sheet 必须留 evidence 候选供人工选择"

    def test_kv_content_page_form_routed(self, db):
        """键值内容页：表单路由优先（即使 det 非 None 误命中）。"""
        conn, pid, report = _import(db, make_kv_content_page)
        assert report.needs_manual_review
        assert all(s.status == "non_settlement_form" for s in report.sheets)
        assert _counts(conn)["line_items"] == 0

    def test_oversized_summary_gated_file_level(self, db):
        """超长资源汇总：文件级 needs_role_review（500+ 行大表护栏）。"""
        conn, pid, report = _import(db, make_oversized_resource_summary)
        assert report.needs_manual_review
        assert all(s.status == "needs_role_review" for s in report.sheets)
        assert _counts(conn)["line_items"] == 0

    def test_mixed_file_semantic_gate(self, db):
        """第九轮口径：'汇总大表'按语义挡、'小清单'强表头无歧义照常解析；
        文件报告必须列出 pending role review。"""
        conn, pid, report = _import(db, make_mixed_oversized)
        by_name = {x.sheet_name: x for x in report.sheets}
        assert by_name["汇总大表"].status == "needs_role_review", \
            "汇总语义 sheet 必须挡下"
        assert by_name["小清单"].status == "parsed", \
            "无歧义强表头小清单照常解析（判断不按行数）"
        # 只要仍有被挡 sheet：overall=partial + needs_manual_review（不得 ok/full_pipeline 掩盖 pending）
        assert report.status == "partial", \
            f"存在 pending sheet 时 overall 必须为 partial: {report.status}"
        assert report.needs_manual_review is True
        assert "pending" in report.message or "needs_manual_review" in report.message
        # pending role review 必须可见（report/audit 双通道）
        assert any(x.status == "needs_role_review" for x in report.sheets)
        assert conn.execute(
            "SELECT COUNT(*) c FROM evidence WHERE kind='sheet_role_candidate'"
        ).fetchone()["c"] >= 1
        assert conn.execute(
            "SELECT COUNT(*) c FROM raw_sheets").fetchone()["c"] == 2  # 保真层保留

    def test_gated_sheets_leave_evidence_and_audit(self, db):
        conn, pid, report = _import(db, make_weak_ledger)
        _ = report
        ev = conn.execute(
            "SELECT summary, sources_json FROM evidence WHERE kind='sheet_role_candidate'"
        ).fetchall()
        assert ev and "待人工" in ev[0]["summary"]
        src = json.loads(ev[0]["sources_json"])[0]
        assert src["sheet_id"] and src["confidence"] is not None
        from costguard.core.evidence import audit as audit_log

        assert any("role" in e.action or "角色" in e.reason
                   for e in audit_log.history_for(conn, pid))


def make_large_legal_settlement(path: Path) -> None:
    """600 行、无歧义强表头、Sheet 名"第1期结算清单"的合法大结算。"""
    import openpyxl

    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "第1期结算清单"
    rows = [(f"K{i:04d}", f"清单项{i}", "m3", i % 50 + 1, round(10 + i % 30, 2),
             round((i % 50 + 1) * (10 + i % 30), 2), "0.09") for i in range(1, 601)]
    _wb_sheet(ws, 1, rows,
              ["清单编码", "清单名称", "单位", "工程量", "综合单价", "合价", "税率"])
    wb.save(path)


def make_small_resource_summary(path: Path) -> None:
    """20 行人材机汇总小表（语义门控：汇总类 sheet 名）。"""
    import openpyxl

    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "人材机汇总"
    rows = [(f"R{i:03d}", f"材料{i}", "kg", i, round(2 + i % 9, 2), i * (2 + i % 9))
            for i in range(1, 21)]
    _wb_sheet(ws, 1, rows, ["编号", "名称规格", "单位", "数量", "单价", "合价"])
    wb.save(path)


def make_ambiguous_header(path: Path) -> None:
    """高置信但歧义表头（两列'综合单价'）→ needs_review，必须挡。"""
    import openpyxl

    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "第1期"
    _wb_sheet(ws, 1, [
        ("K1", "某清单", "m3", 10, 465, 4650, 4850, "0.09"),
        ("K2", "某清单2", "m3", 5, 4850, 24250, 24800, "0.13"),
    ], ["清单编码", "清单名称", "单位", "工程量", "综合单价", "含税合价", "不含税合价", "税率"])
    wb.save(path)


def make_ledger_named_sheet(path: Path) -> None:
    """sheet 名含'台账'且强表头 → 语义门控挡。"""
    import openpyxl

    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "合同台账"
    _wb_sheet(ws, 1, [
        ("K1", "某清单", "m3", 10, 465, 4650),
        ("K2", "某清单2", "m3", 5, 4850, 24250),
    ], ["清单编码", "清单名称", "单位", "工程量", "综合单价", "合价"])
    wb.save(path)


class TestRound9Semantics:
    def test_large_legal_settlement_parsed(self, db):
        """600 行合法大结算必须解析（判断不按大小）。"""
        conn, pid, report = _import(db, make_large_legal_settlement)
        assert report.status == "ok" and not report.needs_manual_review
        assert any(s.status == "parsed" for s in report.sheets)
        assert _counts(conn)["line_items"] == 600

    def test_small_resource_summary_gated(self, db):
        """20 行人材机汇总小表仍需角色复核（判断不按大小）。"""
        conn, pid, report = _import(db, make_small_resource_summary)
        assert report.status == "partial" and report.needs_manual_review
        assert all(s.status in ("needs_role_review", "no_header") for s in report.sheets)
        assert _counts(conn)["line_items"] == 0

    def test_ambiguous_high_confidence_gated(self, db):
        """高置信歧义表头（needs_review）→ 不得写 canonical。"""
        conn, pid, report = _import(db, make_ambiguous_header)
        assert report.status == "partial" and report.needs_manual_review
        assert _counts(conn)["line_items"] == 0

    def test_ledger_named_sheet_gated(self, db):
        """sheet 名含'台账' → 语义门控，即使强表头。"""
        conn, pid, report = _import(db, make_ledger_named_sheet)
        assert report.status == "partial" and report.needs_manual_review
        assert _counts(conn)["line_items"] == 0

    def test_confirm_sheet_role_replays_extraction(self, db):
        """人工确认入口：确认角色为结算清单后重放抽取写 canonical，并留审计。"""
        conn, pid, report = _import(db, make_ledger_named_sheet)
        sheet_id = conn.execute(
            "SELECT id FROM raw_sheets WHERE sheet_name='合同台账'").fetchone()["id"]
        settlement_io.confirm_sheet_role_and_extract(
            conn, pid, sheet_id, actor="复核人",
            reason="经人工核对该 sheet 为结算清单")
        assert _counts(conn)["line_items"] == 2
        from costguard.core.evidence import audit as audit_log

        assert any(e.action == "confirm_sheet_role"
                   for e in audit_log.history_for(conn, pid))

    def test_confirm_requires_reason(self, db):
        conn, pid, report = _import(db, make_ledger_named_sheet)
        sheet_id = conn.execute(
            "SELECT id FROM raw_sheets WHERE sheet_name='合同台账'").fetchone()["id"]
        from costguard.core.evidence import audit as audit_log

        with pytest.raises(audit_log.AuditReasonRequiredError):
            settlement_io.confirm_sheet_role_and_extract(
                conn, pid, sheet_id, actor="复核人", reason="")


def make_mixed_named_summary(path: Path) -> None:
    """文件名含汇总语义 + 混合 sheet（大 gated 表 + 小 canonical 表）
    → 文档级语义门控：整文件需角色审阅。"""
    import openpyxl

    wb = openpyxl.Workbook()
    ws1 = wb.active
    ws1.title = "资源汇总"
    rows = [(f"R{i:04d}", f"项{i}", "kg", 1, 2, 2) for i in range(1, 520)]
    _wb_sheet(ws1, 1, rows, ["编号", "名称规格", "单位", "数量", "单价", "合价"])
    ws2 = wb.create_sheet("钢筋")
    _wb_sheet(ws2, 1, [("K1", "钢筋制作", "t", 3, 4850, 14550)],
              ["清单编码", "清单名称", "单位", "工程量", "综合单价", "合价"])
    wb.save(path)


class TestRound9DocumentSemantics:
    def test_document_level_semantic_gate(self, db):
        """文件名含汇总/核销语义 → 文档级角色审阅，即使存在小 canonical sheet。"""
        conn, pid, tmp = db
        src = tmp / "人材机汇总.xlsx"  # 文件名语义（真实用户文件同样命中）
        make_mixed_named_summary(src)
        report = settlement_io.import_settlement_file(conn, pid, tmp, src)
        assert report.needs_manual_review
        assert all(s.status in ("needs_role_review", "no_header",
                                "non_settlement_form") for s in report.sheets), \
            f"文档级语义门控应挡整文件: {[x.status for x in report.sheets]}"
        assert _counts(conn)["line_items"] == 0


class TestRound9EnglishNames:
    def test_english_summary_filename_gated(self, db):
        """英文文件名语义（resource_summary 等）同样触发文档级门控。"""
        conn, pid, tmp = db
        src = tmp / "R99_resource_summary.xlsx"
        make_mixed_named_summary(src)
        report = settlement_io.import_settlement_file(conn, pid, tmp, src)
        assert report.needs_manual_review
        assert _counts(conn)["line_items"] == 0


class TestExplicitColMapConfirmation:
    def test_ambiguous_requires_explicit_map(self, db):
        """歧义 sheet（两列金额）人工确认必须显式传 confirmed_col_map。"""
        conn, pid, report = _import(db, make_ambiguous_header)
        sheet_id = conn.execute(
            "SELECT id FROM raw_sheets WHERE sheet_name='第1期'").fetchone()["id"]
        # 未传映射 → 拒绝（人工并未真正选择列）
        with pytest.raises(ValueError, match="显式"):
            settlement_io.confirm_sheet_role_and_extract(
                conn, pid, sheet_id, actor="复核人", reason="确认")

    def test_explicit_map_selects_correct_amount_column(self, db):
        """人工明确选择正确金额列（col6 含税合价）后写入，审计记录 old/new mapping。"""
        conn, pid, report = _import(db, make_ambiguous_header)
        sheet_id = conn.execute(
            "SELECT id FROM raw_sheets WHERE sheet_name='第1期'").fetchone()["id"]
        n = settlement_io.confirm_sheet_role_and_extract(
            conn, pid, sheet_id, actor="复核人",
            reason="歧义列经人工核对：金额取含税合价列",
            confirmed_col_map={"code": 1, "name": 2, "unit": 3, "quantity": 4,
                               "unit_price": 5, "amount": 6, "tax_rate": 8})
        assert n == 2
        rows = conn.execute(
            "SELECT code, amount FROM line_items WHERE sheet_id=? ORDER BY id",
            (sheet_id,)).fetchall()
        assert [r["amount"] for r in rows] == ["4650", "24250"], \
            f"金额必须取人工选择的列6: {[tuple(r) for r in rows]}"
        # 审计 old/new mapping
        from costguard.core.evidence import audit as audit_log

        entries = [e for e in audit_log.history_for(conn, pid)
                   if e.action == "confirm_sheet_role"]
        assert entries, "缺少 confirm_sheet_role 审计"
        detail = conn.execute(
            "SELECT before_json, after_json FROM audit_log WHERE action='confirm_sheet_role'").fetchone()
        before = json.loads(detail["before_json"])
        after = json.loads(detail["after_json"])
        # 审计完整记录 old(程序检测) 与 new(人工确认) 两套映射；
        # 人工选对时两者可相等，但 confirmed 必须是人工明确选择的列
        assert after["mapping"]["detected"] and after["mapping"]["confirmed"]
        assert after["mapping"]["confirmed"]["amount"] == 6
        assert before["needs_review"] is True  # 歧义来源留痕（before）

    def test_explicit_map_rejects_invalid(self, db):
        """列映射校验：列号越界/冲突/缺必需字段 → 拒绝。"""
        conn, pid, report = _import(db, make_ambiguous_header)
        sheet_id = conn.execute(
            "SELECT id FROM raw_sheets WHERE sheet_name='第1期'").fetchone()["id"]
        with pytest.raises(ValueError, match="冲突|越界|必需"):
            settlement_io.confirm_sheet_role_and_extract(
                conn, pid, sheet_id, actor="复核人", reason="同列两字段",
                confirmed_col_map={"code": 1, "name": 1})  # 冲突
        with pytest.raises(ValueError, match="冲突|越界|必需"):
            settlement_io.confirm_sheet_role_and_extract(
                conn, pid, sheet_id, actor="复核人", reason="缺名称",
                confirmed_col_map={"code": 1, "unit": 3})  # 缺必需 name
        with pytest.raises(ValueError, match="冲突|越界|必需"):
            settlement_io.confirm_sheet_role_and_extract(
                conn, pid, sheet_id, actor="复核人", reason="越界",
                confirmed_col_map={"name": 99})  # 越界


class TestRound11ConfirmationGates:
    def _gated(self, db):
        conn, pid, report = _import(db, make_ambiguous_header)
        sheet_id = conn.execute(
            "SELECT id FROM raw_sheets WHERE sheet_name='第1期'").fetchone()["id"]
        return conn, pid, sheet_id

    def test_colmap_requires_quantity_and_amount_or_price(self, db):
        """列映射校验收严：name+quantity 必需，amount/unit_price 至少一个。"""
        conn, pid, sheet_id = self._gated(db)
        with pytest.raises(ValueError, match="必需"):
            settlement_io.confirm_sheet_role_and_extract(
                conn, pid, sheet_id, actor="复核人", reason="只有名称",
                confirmed_col_map={"name": 2, "quantity": 4})
        with pytest.raises(ValueError, match="必需"):
            settlement_io.confirm_sheet_role_and_extract(
                conn, pid, sheet_id, actor="复核人", reason="有名称有数量无金额单价",
                confirmed_col_map={"name": 2, "quantity": 4, "code": 1})
        # 合法：name+quantity+unit_price（无 amount 也接受）
        n = settlement_io.confirm_sheet_role_and_extract(
            conn, pid, sheet_id, actor="复核人", reason="取单价路径",
            confirmed_col_map={"code": 1, "name": 2, "unit": 3,
                               "quantity": 4, "unit_price": 5, "tax_rate": 8})
        assert n == 2

    def test_no_empty_period_on_zero_extraction(self, db):
        """抽取 0 行时不得留下空期次。"""
        import openpyxl

        conn, pid, tmp = db
        src = tmp / "header_only.xlsx"
        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = "第1期"
        for c, v in enumerate(["清单编码", "清单名称", "单位", "工程量", "综合单价", "合价"], start=1):
            ws.cell(row=1, column=c, value=v)
        wb.save(src)  # 只有表头无数据行
        settlement_io.import_settlement_file(conn, pid, tmp, src)
        sheet_id = conn.execute("SELECT id FROM raw_sheets LIMIT 1").fetchone()["id"]
        with pytest.raises(ValueError, match="0 行|未抽取到"):
            settlement_io.confirm_sheet_role_and_extract(
                conn, pid, sheet_id, actor="复核人", reason="确认",
                confirmed_col_map={"code": 1, "name": 2, "quantity": 4, "amount": 6})
        assert conn.execute(
            "SELECT COUNT(*) c FROM settlement_periods").fetchone()["c"] == 0, \
            "抽取失败不得留下空期次"

    def test_success_writes_period_link_and_headers(self, db):
        """成功确认后：raw_sheets.period_id 回写、table_headers 记录已确认映射且 needs_review=0。"""
        conn, pid, sheet_id = self._gated(db)
        confirmed = {"code": 1, "name": 2, "unit": 3, "quantity": 4,
                     "unit_price": 5, "amount": 6, "tax_rate": 8}
        settlement_io.confirm_sheet_role_and_extract(
            conn, pid, sheet_id, actor="复核人", reason="人工核对列",
            confirmed_col_map=confirmed)
        rs = conn.execute("SELECT period_id FROM raw_sheets WHERE id=?", (sheet_id,)).fetchone()
        assert rs["period_id"] is not None, "raw_sheets.period_id 未回写"
        th = conn.execute(
            "SELECT col_map_json, needs_review FROM table_headers WHERE sheet_id=?",
            (sheet_id,)).fetchone()
        assert json.loads(th["col_map_json"]) == confirmed
        assert th["needs_review"] == 0

    def test_duplicate_confirm_rejected(self, db):
        """同一 sheet 重复确认 → 明确拒绝，不得重复插入 line_items。"""
        conn, pid, sheet_id = self._gated(db)
        confirmed = {"code": 1, "name": 2, "unit": 3, "quantity": 4,
                     "unit_price": 5, "amount": 6, "tax_rate": 8}
        settlement_io.confirm_sheet_role_and_extract(
            conn, pid, sheet_id, actor="复核人", reason="首次确认",
            confirmed_col_map=confirmed)
        before = conn.execute("SELECT COUNT(*) c FROM line_items WHERE sheet_id=?",
                              (sheet_id,)).fetchone()["c"]
        with pytest.raises(ValueError, match="已确认|重复"):
            settlement_io.confirm_sheet_role_and_extract(
                conn, pid, sheet_id, actor="复核人", reason="再次确认",
                confirmed_col_map=confirmed)
        after = conn.execute("SELECT COUNT(*) c FROM line_items WHERE sheet_id=?",
                             (sheet_id,)).fetchone()["c"]
        assert before == after == 2

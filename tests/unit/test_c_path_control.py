"""C 路径控制值语义测试：分页小计不得与总计累加翻倍。

真实依据（v0.1.5 发布已知限制 + 移动硬盘真实表-08 版式）：标准结算书同时
存在"本页小计"（每页）与"合计"（全表总计），C 路径把全部小计行求和后
控制值翻倍、正常表被误报 diff。设计纪律：
- 合计级行（合计/总计/累计）= 全表控制值；页级小计（小计/本页小计/…）不是；
- 恰好一个合计级行 → C 控制可用；零个 → not_available；多个 → 不唯一，
  not_available 并留待人工（不得擅自挑一个或求和）。
"""
from __future__ import annotations

import sys
from pathlib import Path

import openpyxl

sys.path.insert(0, str(Path(__file__).resolve().parents[2] / "synthetic_test_data"))

from gb_templates import _finalize  # noqa: E402
from openpyxl.styles import Font  # noqa: E402


def _make_two_page(path: Path) -> None:
    """表-08 真实分页版式：第1页 3 行明细+本页小计，第2页 2 行明细+合计（总计）。"""
    wb = openpyxl.Workbook()
    wb.remove(wb.active)
    ws = wb.create_sheet("F.1 分部分项清单(表-08)【合成·两页】")
    ws.cell(row=1, column=1, value="分部分项工程和单价措施项目清单与计价表").font = Font(bold=True)
    for c, t in ((1, "序号"), (2, "项目编码"), (3, "项目名称"), (4, "计量\n单位"),
                 (5, "工程量"), (6, "综合单价"), (7, "合价")):
        ws.cell(row=2, column=c, value=t)
    r = 3
    page1 = [
        ("1", "040101001001", "平整场地", "m2", 100.00, 8.50, 850.00),
        ("2", "040101003001", "挖沟槽土方", "m3", 50.00, 35.20, 1760.00),
        ("3", "040103001001", "填方及运输", "m3", 30.00, 22.00, 660.00),
    ]
    for it in page1:
        for c, v in enumerate(it, 1):
            ws.cell(row=r, column=c, value=v)
        r += 1
    ws.cell(row=r, column=3, value="本页小计")
    ws.cell(row=r, column=7, value=3270.00)  # 第1页小计
    r += 1
    page2 = [
        ("4", "040203006001", "沥青混凝土面层", "m2", 10.00, 128.50, 1285.00),
        ("5", "040203007001", "透层油", "m2", 10.00, 12.00, 120.00),
    ]
    for it in page2:
        for c, v in enumerate(it, 1):
            ws.cell(row=r, column=c, value=v)
        r += 1
    ws.cell(row=r, column=3, value="合计")
    ws.cell(row=r, column=7, value=4675.00)  # 全表总计 = 3270 + 1405
    _finalize(wb, path)


def _import(path: Path):
    from costguard.core.engine import settlement_io
    from costguard.core.models import project as pm

    info = pm.create_project("C路径测试", path.parent)
    info, conn = pm.open_project(info.workspace_path)
    settlement_io.import_settlement_file(
        conn, info.project_id, Path(info.workspace_path), path, direction="upward")
    period_id = conn.execute(
        "SELECT id FROM settlement_periods LIMIT 1").fetchone()["id"]
    return conn, period_id


def test_c_control_uses_grand_total_not_page_subtotal(tmp_path):
    from costguard.core.engine import crosscheck

    src = tmp_path / "two_page.xlsx"
    _make_two_page(src)
    conn, period_id = _import(src)
    try:
        result = crosscheck.check_period(conn, period_id)
        # A = 3270 + 1405 = 4675；C 控制值必须取"合计"行 4675，而非 3270+4675
        assert result.status == "match", f"干净数据 A/B 应一致：{result.notes}"
        assert result.control_status == "match", (
            f"C 控制值应取合计级行 4675，实际 {result.control_status}"
            f"（control_diff={result.control_diff}）——页小计不得累加翻倍")
    finally:
        conn.close()


def test_multiple_grand_totals_make_control_not_available(tmp_path):
    """两个合计级行（如两页各一个"合计"）→ 控制值不唯一，not_available。"""
    from costguard.core.engine import crosscheck

    src = tmp_path / "multi_grand.xlsx"
    _make_two_page(src)
    # 追加第二个"合计"行，制造控制值不唯一
    wb = openpyxl.load_workbook(src)
    ws = wb.worksheets[0]
    r = ws.max_row + 1
    ws.cell(row=r, column=3, value="合计")
    ws.cell(row=r, column=7, value=4675.00)
    _finalize(wb, src)
    conn, period_id = _import(src)
    try:
        result = crosscheck.check_period(conn, period_id)
        assert result.control_status == "not_available", (
            f"多个合计级行时控制值不唯一，应 not_available，实际 {result.control_status}")
        assert any("不唯一" in n for n in result.notes), "必须留注说明原因"
    finally:
        conn.close()


def test_page_subtotals_sum_to_control(tmp_path):
    """无合计级行时，页级小计之和=全表控制值（互斥分页，不与总计混加）。"""
    from costguard.core.engine import crosscheck

    src = tmp_path / "page_only.xlsx"
    _make_two_page(src)
    wb = openpyxl.load_workbook(src)
    ws = wb.worksheets[0]
    # 删除"合计"行，并把"本页小计"改成两行页小计（第1页/第2页）
    ws.delete_rows(ws.max_row)
    ws.cell(row=ws.max_row + 1, column=3, value="第2页小计")
    ws.cell(row=ws.max_row, column=7, value=1405.00)
    _finalize(wb, src)
    conn, period_id = _import(src)
    try:
        result = crosscheck.check_period(conn, period_id)
        # 页小计互斥求和 = 3270 + 1405 = 4675 = A：既不翻倍也不虚缺
        assert result.control_status == "match", (
            f"页级小计之和应为有效控制值，实际 {result.control_status} diff={result.control_diff}")
        assert any("小计行之和" in n for n in result.notes)
    finally:
        conn.close()


def test_partial_page_subtotal_reports_honest_diff(tmp_path):
    """页小计未覆盖全部明细（缺第2页小计/合计）：控制差额如实报 diff。"""
    from costguard.core.engine import crosscheck

    src = tmp_path / "single_sub.xlsx"
    _make_two_page(src)
    wb = openpyxl.load_workbook(src)
    ws = wb.worksheets[0]
    ws.delete_rows(ws.max_row)  # 删"合计"，仅剩一行"本页小计"
    _finalize(wb, src)
    conn, period_id = _import(src)
    try:
        result = crosscheck.check_period(conn, period_id)
        # 源表只有第1页小计（3270），未覆盖第2页明细（1405）：
        # 控制差额 1405 如实呈现为 diff——不调平、不伪装通过
        assert result.control_status == "diff" and result.control_diff == 1405
        assert any("无合计级行" in n for n in result.notes)
    finally:
        conn.close()

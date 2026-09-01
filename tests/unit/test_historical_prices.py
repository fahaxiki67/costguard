"""P2-01 历史综合单价库的证据、Decimal 和不可比闸门测试。"""
from __future__ import annotations

import hashlib
import sqlite3
from decimal import Decimal
from pathlib import Path
from types import SimpleNamespace

import pytest

from jiadun.core.contracts import run_contract
from jiadun.core.evidence import audit as audit_log
from jiadun.core.evidence import evidence as evidence_api
from jiadun.core.models import project as project_model
from jiadun.core.pricing import history
from jiadun.core.versions import create_project_version

D = Decimal


def _version_fixture(tmp_path: Path):
    info = project_model.create_project("历史单价来源项目", tmp_path / "ws")
    info, conn = project_model.open_project(Path(info.workspace_path))
    period_id = conn.execute(
        """INSERT INTO settlement_periods(project_id, period_no, title, direction)
           VALUES (?, 1, '第1期', 'upward') RETURNING id""",
        (info.project_id,),
    ).fetchone()[0]
    stored = tmp_path / "history.xlsx"
    stored.write_bytes(b"synthetic historical-price source")
    stored_digest = hashlib.sha256(stored.read_bytes()).hexdigest()
    file_id = conn.execute(
        """INSERT INTO source_files(
               project_id, original_path, stored_path, original_name, sha256,
               size_bytes, file_type, imported_at)
           VALUES (?, '/history.xlsx', ?, 'history.xlsx', ?, ?, 'xlsx', '2026')
           RETURNING id""",
        (info.project_id, str(stored), stored_digest, stored.stat().st_size),
    ).fetchone()[0]
    batch_id = conn.execute(
        """INSERT INTO parse_batches(file_id, parser, parsed_at, status)
           VALUES (?, 'test', '2026', 'ok') RETURNING id""",
        (file_id,),
    ).fetchone()[0]
    sheet_id = conn.execute(
        """INSERT INTO raw_sheets(batch_id, sheet_index, sheet_name, n_rows, n_cols, period_id)
           VALUES (?, 0, '第1期明细', 4, 5, ?) RETURNING id""",
        (batch_id, period_id),
    ).fetchone()[0]
    conn.executemany(
        """INSERT INTO line_items(
               period_id, sheet_id, code, name, feature, unit, quantity,
               unit_price, amount, flags_json)
           VALUES (?, ?, ?, 'C25 混凝土', '泵送', 'm3', ?, ?, ?, ?)""",
        [
            (period_id, sheet_id, "C-1", "10", "100", "1000", '{"row": 2}'),
            (period_id, sheet_id, "C-2", "20", "150", "3000", '{"row": 3}'),
            (period_id, sheet_id, "C-3", "10", None, None, '{"row": 4}'),
        ],
    )
    initial = create_project_version(
        conn,
        info.project_id,
        "initial_submission",
        "第一次送审",
        created_by="测试人",
        reason="建立历史单价测试基线",
    )
    final = create_project_version(
        conn,
        info.project_id,
        "final_approval",
        "最终审定",
        created_by="测试人",
        reason="完成审定版本",
    )
    return info, conn, final, period_id, sheet_id, file_id, initial


def _manual_closure(conn, project_id: int, final_version_id: int):
    current = run_contract.get_current_contract(conn, project_id)
    assert current is not None
    # 先写入待绑定的关闭快照，再把 closure_id 写入 Evidence/Audit。
    # 这与生产 API 的两阶段绑定顺序一致，也满足 v38 的 target guard。
    closure_id = conn.execute(
        """INSERT INTO project_closures(
               project_id, final_version_id, status, region, project_type,
               observed_at, closed_by, closed_at, reason, run_id, run_signature)
           VALUES (?, ?, 'closed', '北京', '房建', '2026-08', '测试人',
                   '2026-09-01', '测试关闭', ?, ?) RETURNING id""",
        (project_id, final_version_id, current.run_id, current.signature),
    ).fetchone()[0]
    evidence_id = evidence_api.add_evidence(
        conn,
        project_id,
        "project_closure",
        "测试项目关闭",
        sources=[
            {"closure_id": closure_id, "final_version_id": final_version_id}
        ],
        steps=[
            {"closure_id": closure_id, "final_version_id": final_version_id}
        ],
        run_signature=current.signature,
        run_id=current.run_id,
        scope="human",
    )
    audit_id = audit_log.record_audit(
        conn,
        project_id,
        "测试人",
        "close_project_for_history",
        f"project_closure:{closure_id}",
        None,
        {"closure_id": closure_id, "final_version_id": final_version_id},
        "测试关闭",
        run_id=current.run_id,
        run_signature=current.signature,
    )
    conn.execute(
        "UPDATE project_closures SET evidence_id=?, audit_id=? WHERE id=?",
        (evidence_id, audit_id, closure_id),
    )
    conn.commit()
    return closure_id


def test_close_project_for_history_requires_final_version_and_project_gate(tmp_path, monkeypatch):
    info, conn, final, _period_id, _sheet_id, _file_id, initial = _version_fixture(tmp_path)
    try:
        with pytest.raises(history.HistoricalPriceError, match="最终审定版本"):
            history.close_project_for_history(
                conn,
                info.project_id,
                initial.version_id,
                closed_by="测试人",
                reason="不应关闭初始版本",
            )
        monkeypatch.setattr(
            "jiadun.core.reporting.summary.build_project_summary",
            lambda *_args, **_kwargs: SimpleNamespace(
                statuses={"project_status_code": "conditional"},
                verification={"status": "insufficient", "evidence_complete": False},
            ),
        )
        with pytest.raises(history.HistoricalPriceError, match="证据条件"):
            history.close_project_for_history(
                conn,
                info.project_id,
                final.version_id,
                closed_by="测试人",
                reason="证据不完整",
            )
    finally:
        conn.close()


def test_collect_history_uses_final_snapshot_decimal_and_incomparable_hints(tmp_path):
    info, conn, final, _period_id, _sheet_id, _file_id, _initial = _version_fixture(tmp_path)
    try:
        closure_id = _manual_closure(conn, info.project_id, final.version_id)
        collection = history.collect_historical_prices(
            conn,
            closure_id,
            created_by="测试人",
            reason="沉淀已审定项目历史单价",
        )
        assert len(collection.records) == 2
        assert len(collection.pending_items) == 1
        assert all(isinstance(item.unit_price, D) for item in collection.records)
        assert {item.unit_price for item in collection.records} == {D("100"), D("150")}
        assert collection.evidence_id is not None
        evidence = conn.execute(
            "SELECT kind, scope, historical_reason FROM evidence WHERE id=?",
            (collection.evidence_id,),
        ).fetchone()
        assert tuple(evidence) == (
            "historical_unit_price",
            "historical",
            "历史单价资产只作复核提示，不直接认定当前单价错误",
        )

        hint = history.query_historical_price_hint(
            conn,
            name="C25 混凝土",
            feature="泵送",
            unit="m3",
            direction="upward",
            current_unit_price="125",
            region="北京",
            observed_at="2026-08",
            project_type="房建",
        )
        assert hint.status == "available"
        assert hint.review_only is True
        assert hint.sample_count == hint.comparable_count == 2
        assert hint.min_price == D("100")
        assert hint.max_price == D("150")
        assert hint.median_price == D("125")
        assert hint.deviation_from_median == D("0")
        assert hint.deviation_rate == D("0")
        assert len(hint.evidence_ids) == 1

        incomparable = history.query_historical_price_hint(
            conn,
            name="C25 混凝土",
            feature="泵送",
            unit="m3",
            direction="upward",
            current_unit_price="125",
            region="上海",
            observed_at="2026-08",
            project_type="房建",
        )
        assert incomparable.status == "not_comparable"
        assert incomparable.comparable_count == 0
        assert incomparable.incomparable_count == 2
        assert any("地区不一致" in reason for reason in incomparable.reasons)

        missing_dimension = history.query_historical_price_hint(
            conn,
            name="C25 混凝土",
            feature="泵送",
            unit="m3",
            direction="upward",
            current_unit_price="125",
            region="",
            observed_at=None,
            project_type="",
        )
        assert missing_dimension.status == "not_comparable"
        assert missing_dimension.comparable_count == 0
        assert all("不能直接比较" in reason for reason in missing_dimension.reasons)

        assert len(history.list_historical_unit_prices(
            conn, name="C25 混凝土", feature="泵送", unit="m3", direction="upward"
        )) == 2
        with pytest.raises(history.HistoricalPriceError, match="重复写入"):
            history.collect_historical_prices(
                conn,
                closure_id,
                created_by="测试人",
                reason="重复沉淀",
            )
    finally:
        conn.close()


@pytest.mark.parametrize("bad_value", ["not-a-number", "NaN", "Infinity", "-Infinity"])
def test_corrupt_unit_price_is_safe_in_archive_read_and_export(tmp_path, bad_value):
    """损坏的历史单价保留原文并降为无效，不得让档案读取或导出崩溃。"""
    from openpyxl import Workbook, load_workbook

    from jiadun.core.export import excel_export

    info, conn, final, _period_id, _sheet_id, _file_id, _initial = _version_fixture(tmp_path)
    try:
        closure_id = _manual_closure(conn, info.project_id, final.version_id)
        collection = history.collect_historical_prices(
            conn,
            closure_id,
            created_by="测试人",
            reason="建立损坏单价回读测试",
        )
        price_id = collection.records[0].price_id
        conn.execute("DROP TRIGGER trg_historical_unit_prices_immutable_update")
        conn.execute(
            "UPDATE historical_unit_prices SET unit_price=? WHERE id=?",
            (bad_value, price_id),
        )

        archive = history.list_historical_unit_prices(
            conn, source_project_id=info.project_id, include_revoked=True
        )
        record = next(item for item in archive if item.price_id == price_id)
        assert record.status == "invalid"
        assert record.unit_price is None
        assert record.metadata["unit_price_integrity_status"] == "invalid"
        assert record.metadata["unit_price_raw"] == bad_value

        current = history.list_historical_unit_prices(
            conn, source_project_id=info.project_id, include_revoked=False
        )
        assert all(item.price_id != price_id for item in current)

        workbook = Workbook()
        excel_export.export_historical_price_sheet(conn, info.project_id, workbook)
        output = tmp_path / "historical-prices.xlsx"
        workbook.save(output)
        loaded = load_workbook(output, data_only=False)
        sheet = loaded["历史综合单价库"]
        headers = [cell.value for cell in sheet[1]]
        integrity_col = headers.index("单价完整性") + 1
        raw_col = headers.index("单价原始值") + 1
        values = [
            (sheet.cell(row=row, column=integrity_col).value,
             sheet.cell(row=row, column=raw_col).value)
            for row in range(2, sheet.max_row + 1)
        ]
        assert ("invalid", bad_value) in values
    finally:
        conn.close()


def test_project_closure_api_records_evidence_and_history_is_immutable(tmp_path, monkeypatch):
    info, conn, final, _period_id, _sheet_id, _file_id, _initial = _version_fixture(tmp_path)
    try:
        monkeypatch.setattr(
            "jiadun.core.reporting.summary.build_project_summary",
            lambda *_args, **_kwargs: SimpleNamespace(
                statuses={"project_status_code": "can_conclude"},
                verification={"status": "sufficient", "evidence_complete": True},
            ),
        )
        closure = history.close_project_for_history(
            conn,
            info.project_id,
            final.version_id,
            closed_by="测试人",
            reason="完成项目关闭确认",
            region="北京",
            project_type="房建",
            observed_at="2026-08",
        )
        assert closure.status == "closed"
        assert closure.evidence_id is not None and closure.audit_id is not None
        stored = history.get_project_closure(conn, info.project_id)
        assert stored is not None and stored.closure_id == closure.closure_id
        with pytest.raises(sqlite3.IntegrityError, match="project closure immutable"):
            conn.execute(
                "UPDATE project_closures SET region='上海' WHERE id=?",
                (closure.closure_id,),
            )
        with pytest.raises(sqlite3.IntegrityError, match="project closure immutable"):
            conn.execute("DELETE FROM project_closures WHERE id=?", (closure.closure_id,))
    finally:
        conn.close()


def test_historical_price_direct_insert_without_closed_evidence_is_rejected(tmp_path):
    info, conn, final, _period_id, _sheet_id, _file_id, _initial = _version_fixture(tmp_path)
    try:
        item_id = conn.execute(
            "SELECT id FROM project_version_items WHERE version_id=? ORDER BY id LIMIT 1",
            (final.version_id,),
        ).fetchone()[0]
        with pytest.raises(
            sqlite3.IntegrityError,
            match="FOREIGN KEY|closed project evidence|snapshot source mismatch",
        ):
            conn.execute(
                """INSERT INTO historical_unit_prices(
                       source_project_id, closure_id, source_version_id, source_version_item_id,
                       raw_project_name, normalized_project_name, raw_name, normalized_name,
                       unit_price, evidence_id, audit_id, created_by, created_at)
                   VALUES (?, 99999, ?, ?, '项目', '项目', '清单', '清单',
                           '1', 99999, 99999, '测试人', '2026')""",
                (info.project_id, final.version_id, item_id),
            )
    finally:
        conn.close()

"""语言扫描测试：用户可见界面/成果页不得暴露内部英文与开发字段。

扫描两层：
1. 源码层——ui/ 与 core/export/ 中不得残留短方向词字面量（"对上"/"对下"）
   或 addItems 原始键列表（upward/downward 直接显示）；
2. 行为层——导出工作簿中 方向列/规则列/对象列/状态列 全部为业务中文，
   内部代码仅允许出现在"规则代码"列（高级证据）。
"""
from __future__ import annotations

import re
from pathlib import Path

import openpyxl

REPO_ROOT = Path(__file__).resolve().parents[2]
SCAN_DIRS = [
    REPO_ROOT / "src" / "costguard" / "ui",
    REPO_ROOT / "src" / "costguard" / "core" / "export",
]
# 允许的完整业务词（含这些子串不算暴露）
BUSINESS_WORDS = ("对上结算", "对下结算", "未标记", "对上累计", "对下累计")


def _short_direction_exposure(text: str) -> list[str]:
    """查找未拼接成业务词的短方向词字面量。"""
    hits = []
    for m in re.finditer(r'["\'](对上|对下)["\']', text):
        ctx = text[max(0, m.start() - 30):m.end() + 10]
        if not any(w in ctx for w in BUSINESS_WORDS):
            hits.append(ctx.replace("\n", " ")[:70])
    return hits


def test_no_short_direction_literals_in_ui_and_export():
    offenders = []
    for d in SCAN_DIRS:
        for f in d.rglob("*.py"):
            hits = _short_direction_exposure(f.read_text(encoding="utf-8"))
            for h in hits:
                offenders.append(f"{f.relative_to(REPO_ROOT)}: …{h}…")
    assert not offenders, "短方向词字面量残留（应为 对上结算/对下结算）：\n" + "\n".join(offenders)


def test_no_raw_key_list_additems_in_ui():
    for f in (REPO_ROOT / "src" / "costguard" / "ui").rglob("*.py"):
        src = f.read_text(encoding="utf-8")
        assert 'addItems(["", "upward", "downward"])' not in src, (
            f"{f} 仍在用原始键列表做显示项（应 addItem(中文, userData)）")


def test_workbench_direction_combo_uses_userdata(qtapp=None):
    from PySide6.QtWidgets import QApplication

    _app = QApplication.instance() or QApplication([])
    import tempfile

    from costguard.core.engine import settlement_io
    from costguard.core.models import project as pm
    from costguard.ui.workbench import WorkbenchPage

    ws = Path(tempfile.mkdtemp(prefix="cg-scan-")) / "ws"
    info = pm.create_project("扫描组合", ws)
    info, conn = pm.open_project(Path(info.workspace_path))
    src = REPO_ROOT / "examples" / "demo" / "演示-对上结算-第1至3期.xlsx"
    settlement_io.import_settlement_file(
        conn, info.project_id, Path(info.workspace_path), src, direction="upward")
    page = WorkbenchPage(conn, info, info.workspace_path, on_back=lambda: None)
    try:
        texts = [page.dir_combo.itemText(i)
                 for i in range(page.dir_combo.count())]
        datas = [page.dir_combo.itemData(i)
                 for i in range(page.dir_combo.count())]
        assert set(texts) == {"未标记", "对上结算", "对下结算"}
        assert set(datas) == {"unknown", "upward", "downward"}
        # 期次表方向列全部业务词
        page.refresh_periods()
        for i in range(page.period_table.rowCount()):
            assert page.period_table.item(i, 2).text() in {
                "对上结算", "对下结算", "未标记"}
    finally:
        conn.close()


def test_exported_workbook_business_language_only(tmp_path):
    """导出工作簿：方向/规则/对象/状态列全部业务中文；内部代码仅在规则代码列。"""
    import re as _re

    from costguard.core.engine import settlement_io
    from costguard.core.export import excel_export
    from costguard.core.models import project as pm

    demo = REPO_ROOT / "examples" / "demo"
    ws = tmp_path / "CostGuardProjects"
    info = pm.create_project("语言扫描导出", ws)
    info, conn = pm.open_project(Path(info.workspace_path))
    settlement_io.import_settlement_file(
        conn, info.project_id, Path(info.workspace_path),
        demo / "演示-对下结算-附表.xlsx", direction="downward")
    from costguard.core.anomalies import engine as anomaly_engine

    anomaly_engine.run_anomalies(conn, info.project_id)
    out = excel_export.export_workbook(conn, info.project_id, Path(info.workspace_path) / "exports")
    conn.close()
    wb = openpyxl.load_workbook(out)
    raw_direction = _re.compile(r"^(upward|downward|unknown)$")
    raw_rule = _re.compile(r"^[a-z][a-z0-9_]+$")
    violations = []
    for sheet_name in wb.sheetnames:
        ws = wb[sheet_name]
        headers = {ws.cell(row=1, column=c).value: c
                   for c in range(1, ws.max_column + 1)}
        dcol = headers.get("方向")
        rcol = headers.get("规则")
        code_col = headers.get("规则代码")
        ocol = headers.get("对象")
        scol = headers.get("状态")
        for r in range(2, ws.max_row + 1):
            if dcol:
                v = ws.cell(row=r, column=dcol).value
                if v and raw_direction.match(str(v)):
                    violations.append(f"{sheet_name} r{r} 方向={v}")
            if rcol:
                v = str(ws.cell(row=r, column=rcol).value or "")
                if raw_rule.match(v) and v not in ("rule_error",):
                    violations.append(f"{sheet_name} r{r} 规则裸代码={v}")
            if ocol:
                v = str(ws.cell(row=r, column=ocol).value or "")
                if _re.match(r"^[a-z_]+#\d+$", v):
                    violations.append(f"{sheet_name} r{r} 对象裸代码={v}")
            if scol:
                v = str(ws.cell(row=r, column=scol).value or "")
                if v in ("open", "resolved"):
                    violations.append(f"{sheet_name} r{r} 状态裸代码={v}")
        if code_col and rcol:
            for r in range(2, ws.max_row + 1):
                zh_v = str(ws.cell(row=r, column=rcol).value or "")
                code_v = str(ws.cell(row=r, column=code_col).value or "")
                if code_v and zh_v == code_v and raw_rule.match(code_v):
                    violations.append(f"{sheet_name} r{r} 规则列=裸代码（应中文）")
    assert not violations, "成果页暴露内部代码：\n" + "\n".join(violations[:10])

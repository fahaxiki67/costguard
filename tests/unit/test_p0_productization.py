"""v0.1.7 P0 产品化验收测试。

P0-1 清单分页（>2000 行不静默截断）；P0-2/P0-5 校核三档级别与行数统计、
绿色严格化；P0-4 业务界面中文化（方向=对上结算/对下结算、OK/Cancel 中文、
Schema 只进 Tooltip）。全部先红后绿。
"""
from __future__ import annotations

import json
import os
from pathlib import Path

import pytest
from PySide6.QtWidgets import QLabel

os.environ.setdefault("QT_QPA_PLATFORM", "offscreen")

pytest.importorskip("PySide6")

REPO_ROOT = Path(__file__).resolve().parents[2]


def _bulk_project(tmp_path: Path, n_items: int):
    """直接批量插入 n_items 明细（绕过 UI），构造大表项目。"""
    from costguard.core.models import project as pm

    info = pm.create_project("大表项目", tmp_path / "ws")
    info, conn = pm.open_project(Path(info.workspace_path))
    pid = info.project_id
    conn.execute(
        """INSERT INTO settlement_periods(id, project_id, period_no, title,
           source_file_id, direction, contract_party, note)
           VALUES (1, ?, 1, '第1期', NULL, 'upward', '', NULL)""", (pid,))
    rows = [
        (i, 1, None, f"Code{i:06d}", f"清单项{i}", "", "m2",
         str(i % 97 + 1), str(round(1.5 + i % 31, 2)),
         str(round((i % 97 + 1) * (1.5 + i % 31), 2)), "0.09", json.dumps({"row": i}))
        for i in range(1, n_items + 1)
    ]
    conn.executemany(
        """INSERT INTO line_items(id, period_id, sheet_id, code, name, feature, unit,
           quantity, unit_price, amount, tax_rate, flags_json)
           VALUES (?,?,?,?,?,?,?,?,?,?,?,?)""", rows)
    conn.commit()
    return info, conn, pid


class TestP01ItemsPagination:
    """P0-1：清单明细取消静默 LIMIT 2000，分页显示且总数可见。"""

    @pytest.fixture()
    def page(self, tmp_path, monkeypatch):
        from PySide6.QtWidgets import QApplication

        _app = QApplication.instance() or QApplication([])
        from costguard.ui.workbench import WorkbenchPage

        info, conn, pid = _bulk_project(tmp_path, 2500)
        page = WorkbenchPage(conn, info, info.workspace_path, on_back=lambda: None)
        page.setAttribute(
            __import__("PySide6.QtCore", fromlist=["Qt"]).Qt.WA_DontShowOnScreen, True)
        page.show()
        yield page
        page.conn.close()

    def test_total_counter_shows_all_rows(self, page):
        lbl = page.items_total_label.text()
        assert "2500" in lbl and "共" in lbl, f"总数标签应显示共 2500 条：{lbl}"

    def test_first_page_shows_page_size_only(self, page):
        assert page.items_table.rowCount() == page.PAGE_SIZE, (
            "单页只显示 PAGE_SIZE 行，不得一次全量渲染")
        lbl = page.items_total_label.text()
        assert "1-500" in lbl, f"应显示当前区间 1-500：{lbl}"

    def test_pagination_covers_all_rows(self, page):
        seen_first_codes = []
        for p in range(5):
            page.items_page = p
            page.refresh_items()
            first = page.items_table.item(0, 2).text()
            seen_first_codes.append(first)
            rows = page.items_table.rowCount()
            if p < 4:
                assert rows == page.PAGE_SIZE
        assert len(set(seen_first_codes)) == 5, "每页首行必须不同（真实翻页）"
        # 最后一页：第 2500 条可见
        page.items_page = 4
        page.refresh_items()
        last_page_rows = page.items_table.rowCount()
        assert 0 < last_page_rows <= page.PAGE_SIZE
        codes = {page.items_table.item(i, 2).text() for i in range(last_page_rows)}
        assert "Code002500" in codes, "第 2500 条必须可通过翻页看到（不得静默截断）"

    def test_db_full_dataset_intact_for_export(self, page):
        """导出读取完整数据集：DB 行数不受 UI 分页影响。"""
        n = page.conn.execute(
            "SELECT COUNT(*) c FROM line_items").fetchone()["c"]
        assert n == 2500


class TestP02VerificationLevel:
    """P0-2/P0-5：校核三档级别 + 行数统计 + 绿色严格化。"""

    def _seed(self, tmp_path, *, amount=None, grand="合计", subtotal_rows=True,
              pending_sheets=0):
        from costguard.core.engine import settlement_io
        from costguard.core.models import project as pm

        info = pm.create_project("校核级别", tmp_path / "ws")
        info, conn = pm.open_project(info.workspace_path)
        pid = info.project_id
        pdir = Path(info.workspace_path)
        src = tmp_path / "t.xlsx"
        import openpyxl

        wb = openpyxl.Workbook()
        wb.remove(wb.active)
        ws = wb.create_sheet("F.1 明细(表-08)【合成】")
        for c, t in ((1, "序号"), (2, "项目编码"), (3, "项目名称"), (4, "计量\n单位"),
                     (5, "工程量"), (6, "综合单价"), (7, "合价")):
            ws.cell(row=1, column=c, value=t)
        data = [
            ("1", "040101001001", "平整场地", "m2", 100.00, 8.50, 850.00),
            ("2", "040101003001", "挖沟槽土方", "m3", 50.00, 35.20, 1760.00),
        ]
        r = 2
        for it in data:
            for c, v in enumerate(it, 1):
                ws.cell(row=r, column=c, value=v)
            r += 1
        if subtotal_rows:
            ws.cell(row=r, column=3, value="本页小计")
            ws.cell(row=r, column=7, value=2610.00)
            r += 1
        if grand:
            ws.cell(row=r, column=3, value=grand)
            ws.cell(row=r, column=7, value=amount if amount is not None else 2610.00)
        wb.save(src)
        settlement_io.import_settlement_file(conn, pid, pdir, src, direction="upward")
        for _ in range(pending_sheets):
            # 追加一张未确认的语义门控页（待人工工作表）
            wb2 = openpyxl.Workbook()
            wb2.remove(wb2.active)
            ws2 = wb2.create_sheet("人材机汇总")
            ws2["A1"] = "合成"
            p2 = tmp_path / f"extra{pending_sheets}.xlsx"
            wb2.save(p2)
            settlement_io.import_settlement_file(conn, pid, pdir, p2, direction="upward")
        period_id = conn.execute(
            "SELECT id FROM settlement_periods LIMIT 1").fetchone()["id"]
        return conn, period_id

    def test_full_match_is_sufficient(self, tmp_path):
        from costguard.core.engine import crosscheck

        conn, period_id = self._seed(tmp_path, amount=2610.00, grand="合计")
        try:
            r = crosscheck.check_period(conn, period_id)
            assert r.verification_level == "sufficient", r.notes
            assert r.detail_rows == 2
            assert r.excluded_subtotal_rows == 2, "本页小计+合计均为排除行"
        finally:
            conn.close()

    def test_c_unavailable_is_insufficient_not_match(self, tmp_path):
        """C 不可用（无任何小计/合计行）时禁止纯绿色通过（P0-5）。

        前提说明：页级小计之和是合法控制值（PR#13 已定语义），
        故 C 不可用仅当小计/合计行完全缺失。
        """
        from costguard.core.engine import crosscheck

        conn, period_id = self._seed(tmp_path, grand=None, subtotal_rows=False)
        try:
            r = crosscheck.check_period(conn, period_id)
            assert r.status == "match" and r.control_status == "not_available"
            assert r.verification_level == "insufficient", (
                "A/B 一致但 C 不可用 → 校核不充分，不得绿色通过")
        finally:
            conn.close()

    def test_control_diff_is_findings(self, tmp_path):
        from costguard.core.engine import crosscheck

        conn, period_id = self._seed(tmp_path, amount=9999.00, grand="合计")
        try:
            r = crosscheck.check_period(conn, period_id)
            assert r.control_status == "diff"
            assert r.verification_level == "findings"
        finally:
            conn.close()

    def test_pending_sheets_block_green(self, tmp_path):
        """存在待人工工作表 → 不得校核充分（P0-5）。"""
        from costguard.core.engine import crosscheck

        conn, period_id = self._seed(tmp_path, amount=2610.00, grand="合计",
                                     pending_sheets=1)
        try:
            r = crosscheck.check_period(conn, period_id)
            assert r.verification_level == "insufficient"
            assert r.pending_sheets >= 1
        finally:
            conn.close()

    def test_counts_exposed_for_display(self, tmp_path):
        """三行数指标：参与累计明细 / 排除小计 / 排除标题说明行。"""
        from costguard.core.engine import crosscheck

        conn, period_id = self._seed(tmp_path, amount=2610.00, grand="合计",
                                     subtotal_rows=True)
        try:
            r = crosscheck.check_period(conn, period_id)
            assert r.detail_rows == 2
            assert r.excluded_subtotal_rows == 2, "本页小计+合计均为排除行"
            assert r.excluded_title_rows >= 0  # 本夹具无标题行，0 合法
        finally:
            conn.close()


class TestP04BusinessLanguage:
    """P0-4：业务界面中文化。"""

    def test_direction_labels_full_business_words(self):
        from costguard.ui.labels import DIRECTION_ZH

        assert DIRECTION_ZH["upward"] == "对上结算"
        assert DIRECTION_ZH["downward"] == "对下结算"
        assert DIRECTION_ZH["unknown"] == "未标记"

    def test_reason_dialog_buttons_chinese(self, tmp_path):
        from PySide6.QtWidgets import QApplication, QPushButton

        _app = QApplication.instance() or QApplication([])
        from costguard.ui.workbench import ReasonDialog

        dlg = ReasonDialog("处理异常", "将异常标记为已处理")
        texts = [b.text() for b in dlg.findChildren(QPushButton)]
        assert texts, "必须有按钮"
        for t in texts:
            assert t not in ("OK", "Cancel", "Ok"), f"按钮英文未中文化：{t}"
        assert "确定" in texts or "确认" in texts, f"缺少中文确认按钮：{texts}"
        assert "取消" in texts, f"缺少中文取消按钮：{texts}"
        dlg.close()

    def test_no_schema_word_in_visible_status(self, tmp_path, monkeypatch):
        """Schema 字样只进 Tooltip，不得出现在可见状态文本。"""
        from PySide6.QtWidgets import QApplication

        _app = QApplication.instance() or QApplication([])
        from costguard.core.models import project as pm
        from costguard.ui.main_window import MainWindow

        ws = tmp_path / "ws"
        info = pm.create_project("中文化检查", ws)
        monkeypatch.setattr(pm, "workspace_root", lambda: ws)
        win = MainWindow()
        try:
            info2, conn = pm.open_project(Path(info.workspace_path))
            win._enter_workbench(info2, conn)
            msg = win.statusBar().currentMessage()
            assert "schema" not in msg.lower(), f"状态栏暴露 schema：{msg}"
            page = win.stack.widget(1)
            for lbl in page.findChildren(QLabel):
                if "Schema" in lbl.text():
                    raise AssertionError(f"可见标签暴露 Schema：{lbl.text()}")
        finally:
            win.close()

"""P1-01 清单差异雷达：分类、Decimal 影响和当前运行边界。"""

from decimal import Decimal
from pathlib import Path

import pytest

from jiadun.core.db import migrations
from jiadun.core.diff.radar import build_diff_radar, persist_diff_radar, read_current_diff_radar

D = Decimal


@pytest.fixture()
def db(tmp_path: Path):
    db_path = tmp_path / "project.db"
    migrations.migrate(db_path, tmp_path / "backups")
    conn = migrations.connect(db_path)
    with conn:
        project_id = conn.execute(
            "INSERT INTO projects(name, schema_version, workspace_path, created_at) VALUES (?,?,?,?)",
            ("差异雷达测试", migrations.LATEST_SCHEMA_VERSION, str(tmp_path), "2026"),
        ).lastrowid
    yield conn, int(project_id)
    conn.close()


def _period(conn, project_id: int, no: int, title: str, direction: str = "downward") -> int:
    with conn:
        return int(
            conn.execute(
                """INSERT INTO settlement_periods(project_id, period_no, title, direction)
                   VALUES (?,?,?,?)""",
                (project_id, no, title, direction),
            ).lastrowid
        )


def _item(
    conn,
    period_id: int,
    code: str,
    name: str,
    qty: str | None,
    price: str | None,
    amount: str | None,
    *,
    feature: str = "",
    unit: str = "m3",
):
    with conn:
        conn.execute(
            """INSERT INTO line_items(
                   period_id, code, name, feature, unit, quantity, unit_price, amount, flags_json)
               VALUES (?,?,?,?,?,?,?,?,?)""",
            (period_id, code, name, feature, unit, qty, price, amount, "{}"),
        )


def test_radar_classifies_changes_and_excludes_unconfirmed_net(db):
    conn, project_id = db
    baseline = _period(conn, project_id, 1, "基准期")
    current = _period(conn, project_id, 2, "当前期")
    _item(conn, baseline, "K1", "旧名称", "10", "10", "100")
    _item(conn, current, "K1", "新名称", "10", "12", "120")
    _item(conn, baseline, "K2", "删除项", "1", "30", "30")
    _item(conn, current, "K3", "新增项", "1", "50", "50")
    _item(conn, baseline, "K4", "数量项", "2", "10", "20")
    _item(conn, current, "K4", "数量项", "3", "10", "30")
    _item(conn, baseline, "K5", "单位项", "1", "10", "10", unit="m3")
    _item(conn, current, "K5", "单位项", "1", "10", "10", unit="m2")
    _item(conn, baseline, "K6", "编码变化", "1", "10", "10")
    _item(conn, current, "K6-new", "编码变化", "1", "10", "10")
    _item(conn, baseline, "K7", "特征项", "1", "10", "10", feature="普通")
    _item(conn, current, "K7", "特征项", "1", "10", "10", feature="加强")
    _item(conn, baseline, "K8", "待确认金额", "", "", None)
    _item(conn, current, "K8", "待确认金额", "", "", None)

    radar = build_diff_radar(conn, project_id, baseline, current)
    assert radar.status == "conditional"
    by_code = {
        item.current.get("code") or item.baseline.get("code"): item
        for item in radar.items
    }
    assert by_code["K1"].primary_category == "name_changed"
    assert {"name_changed", "unit_price_changed", "amount_changed"} <= set(by_code["K1"].categories)
    assert by_code["K1"].confirmed_amount_impact == D("20")
    assert by_code["K2"].primary_category == "deleted"
    assert by_code["K2"].confirmed_amount_impact == D("-30")
    assert by_code["K3"].primary_category == "added"
    assert by_code["K3"].confirmed_amount_impact == D("50")
    assert by_code["K5"].status == "incomparable"
    assert by_code["K5"].confirmed_amount_impact is None
    assert by_code["K8"].status == "pending"
    assert by_code["K8"].confirmed_amount_impact is None
    # 20 - 30 + 50 + 10 = 50；单位变化和待确认项均不进入确认净影响。
    assert radar.confirmed_net_amount_impact == D("50")
    assert radar.category_summary["added"]["count"] == 1
    assert radar.category_summary["added"]["amount_impact"] == "50"
    assert radar.status_counts["pending"] == 1
    assert radar.status_counts["incomparable"] == 1


def test_cross_direction_comparison_requires_explicit_opt_in(db):
    conn, project_id = db
    upward = _period(conn, project_id, 1, "对上第1期", "upward")
    downward = _period(conn, project_id, 1, "对下第1期", "downward")
    _item(conn, upward, "K1", "项目", "1", "10", "10")
    _item(conn, downward, "K1", "项目", "1", "12", "12")
    with pytest.raises(ValueError, match="不能默认互比"):
        build_diff_radar(conn, project_id, upward, downward)
    radar = build_diff_radar(
        conn, project_id, upward, downward, allow_cross_direction=True
    )
    assert "cross_direction_explicit" in radar.reason_codes
    assert radar.confirmed_net_amount_impact == D("2")


def test_persisted_radar_is_evidence_bound_and_immutable(db):
    conn, project_id = db
    baseline = _period(conn, project_id, 1, "基准期")
    current = _period(conn, project_id, 2, "当前期")
    _item(conn, baseline, "K1", "项目", "1", "10", "10")
    _item(conn, current, "K1", "项目", "2", "10", "20")
    radar = persist_diff_radar(conn, build_diff_radar(conn, project_id, baseline, current))
    assert radar.evidence_id is not None
    assert radar.diff_run_id is not None
    evidence = conn.execute(
        "SELECT kind, scope, run_id, run_signature FROM evidence WHERE id=?",
        (radar.evidence_id,),
    ).fetchone()
    assert tuple(evidence) == ("diff_radar", "current", radar.run_id, radar.run_signature)
    current_rows = read_current_diff_radar(conn, project_id)
    assert len(current_rows) == 1
    assert current_rows[0]["id"] == radar.diff_run_id
    assert current_rows[0]["items"][0]["confirmed_amount_impact"] == "10"
    with pytest.raises(Exception, match="diff run immutable"):
        conn.execute("UPDATE diff_runs SET status='available' WHERE id=?", (radar.diff_run_id,))
    with pytest.raises(Exception, match="diff item immutable"):
        conn.execute("DELETE FROM diff_items WHERE diff_run_id=?", (radar.diff_run_id,))

    # 期次标题变化使输入签名变化；旧比较快照保留，但不得进入当前读取面。
    with conn:
        conn.execute("UPDATE settlement_periods SET title='当前期补充资料' WHERE id=?", (current,))
    assert read_current_diff_radar(conn, project_id) == []


def test_export_radar_contains_category_impact_and_sources(db):
    from openpyxl import Workbook

    from jiadun.core.export.excel_export import export_diff_radar_sheet

    conn, project_id = db
    baseline = _period(conn, project_id, 1, "基准期")
    current = _period(conn, project_id, 2, "当前期")
    _item(conn, baseline, "K1", "项目", "1", "10", "10")
    _item(conn, current, "K1", "项目", "2", "10", "20")
    workbook = Workbook()
    workbook.remove(workbook.active)
    export_diff_radar_sheet(conn, project_id, workbook)

    summary = workbook["清单差异雷达"]
    rows = list(summary.iter_rows(min_row=1, values_only=True))
    assert rows[0][:5] == ("比较范围", "基准期", "当前期", "方向", "类别")
    assert rows[1][4] == "确认净金额影响"
    assert rows[1][6] == D("10")
    assert any(row[4] == "工程量变化" and row[5] == 1 and row[6] == D("10") for row in rows)

    detail = workbook["差异雷达明细"]
    headers = [detail.cell(1, column).value for column in range(1, detail.max_column + 1)]
    assert "Evidence ID" in headers
    detail_values = [
        [detail.cell(row, column).value for column in range(1, detail.max_column + 1)]
        for row in range(2, detail.max_row + 1)
    ]
    assert len(detail_values) == 1
    assert detail_values[0][12] == D("10")
    assert detail_values[0][13] == D("10")
    assert detail_values[0][15] is not None


def test_mirror_comparison_uses_decimal_differences_and_both_directions(db):
    from jiadun.core.contracts import run_contract
    from jiadun.core.matching.mirror import build_mirror_comparison

    conn, project_id = db
    upward = _period(conn, project_id, 1, "对上第1期", "upward")
    downward = _period(conn, project_id, 1, "对下第1期", "downward")
    _item(conn, upward, "K1", "对上名称", "10", "10", "100", feature="普通")
    _item(conn, downward, "K1", "对下名称", "12", "12", "140", feature="加强")
    contract = run_contract.ensure_run_contract(conn, project_id)
    with conn:
        match_id = conn.execute(
            """INSERT INTO matches(
                   project_id, group_key, item_ids_json, level, method, score,
                   status, run_signature, run_id)
               VALUES (?,?,?,?,?,?, 'pending', ?, ?)""",
            (project_id, "downward:code:K1", "[]", "probable", "code_exact", 0.8,
             contract.signature, contract.run_id),
        ).lastrowid
    result = build_mirror_comparison(conn, project_id, int(match_id))
    assert result.match_id == match_id
    assert result.match_status == "pending"
    assert result.quantity_difference == D("2")
    assert result.unit_price_difference == D("2")
    assert result.amount_difference == D("40")
    assert result.amount_difference_rate == D("0.4")
    by_field = {field.field: field for field in result.fields}
    assert by_field["name"].status == "名称存在差异"
    assert by_field["feature"].status == "存在差异"
    assert by_field["unit"].status == "完全一致"
    assert result.upward_sources and result.downward_sources


def test_mirror_comparison_missing_side_is_pending_not_zero(db):
    from jiadun.core.contracts import run_contract
    from jiadun.core.matching.mirror import build_mirror_comparison

    conn, project_id = db
    downward = _period(conn, project_id, 1, "对下第1期", "downward")
    _item(conn, downward, "K1", "只有对下", "2", "10", "20")
    contract = run_contract.ensure_run_contract(conn, project_id)
    with conn:
        match_id = conn.execute(
            """INSERT INTO matches(
                   project_id, group_key, item_ids_json, level, method, score,
                   status, run_signature, run_id)
               VALUES (?,?,?,?,?,?, 'pending', ?, ?)""",
            (project_id, "downward:code:K1", "[]", "probable", "code_exact", 0.8,
             contract.signature, contract.run_id),
        ).lastrowid
    result = build_mirror_comparison(conn, project_id, int(match_id))
    assert result.quantity_difference is None
    assert result.amount_difference is None
    assert all(field.status == "待补资料/缺一侧" for field in result.fields)

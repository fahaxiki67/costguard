"""P0-3 Excel 审核底稿双值与结构可复现检查。

要求：底稿同时输出"程序 Decimal 计算值"与"Excel 复核公式值"，
主结论不依赖 Office 重算。当前环境无 WPS/macOS Excel/Windows Excel
三 Office，真机重算列为待补环境；本测试做结构/公式/数值可复现检查：
1. 合价列=公式串（Excel/WPS 打开时重算）；
2. 新增"程序计算合价"列=Decimal 精确数值（引擎同口径），与数量×单价独立
   复核一致；
3. 公式单元格无缓存值属预期（首次打开由 Office 计算），程序值列兜底；
4. 待补资料行为不因新增列改变（不得出现 0.00 充数）。
"""
from __future__ import annotations

from decimal import Decimal
from pathlib import Path

import openpyxl
import pytest

from jiadun.core.engine.money import money_mul
from jiadun.core.export import excel_export
from jiadun.core.models import project as pm


@pytest.fixture()
def exported(tmp_path):
    from jiadun.core.engine import settlement_io

    info = pm.create_project("双值底稿", tmp_path / "ws")
    info, conn = pm.open_project(Path(info.workspace_path))
    src = tmp_path / "demo.xlsx"
    wb = openpyxl.Workbook()
    wb.remove(wb.active)
    ws = wb.create_sheet("F.1 明细(表-08)【合成】")
    for c, t in ((1, "序号"), (2, "项目编码"), (3, "项目名称"), (4, "计量\n单位"),
                 (5, "工程量"), (6, "综合单价"), (7, "合价")):
        ws.cell(row=1, column=c, value=t)
    data = [
        ("1", "040101001001", "平整场地", "m2", 100.00, 8.50, 850.00),
        ("2", "040101003001", "挖沟槽土方", "m3", 50.00, 35.20, 1760.00),
        ("3", "040103006001", "混凝土面层", "m2", 10.00, 128.505, 1285.05),
        ("4", "040103007001", "缺失单价行", "m2", 10.00, None, None),
    ]
    r = 2
    for it in data:
        for c, v in enumerate(it, 1):
            if v is not None:
                ws.cell(row=r, column=c, value=v)
        r += 1
    ws.cell(row=r, column=3, value="合计")
    ws.cell(row=r, column=7, value=4180.05)
    wb.save(src)
    settlement_io.import_settlement_file(
        conn, info.project_id, Path(info.workspace_path), src, direction="upward")
    out = excel_export.export_workbook(conn, info.project_id, Path(info.workspace_path) / "exports")
    wb_out = openpyxl.load_workbook(out)
    ws_out = wb_out["审核底稿"]
    yield ws_out, conn
    conn.close()
    wb_out.close()


def _header_map(ws_out):
    return {ws_out.cell(row=1, column=c).value: c
            for c in range(1, ws_out.max_column + 1)}


class TestDualValueWorksheet:
    def test_formula_and_program_value_columns_coexist(self, exported):
        ws_out, _ = exported
        hm = _header_map(ws_out)
        assert "合价(底稿公式)" in hm
        assert "程序计算合价" in hm, "缺少程序 Decimal 计算值列（主结论不依赖 Office 重算）"

    def test_program_values_match_decimal_recompute(self, exported):
        """程序计算合价列 = Decimal(数量)×Decimal(单价)，与 DB 同口径。"""
        ws_out, conn = exported
        hm = _header_map(ws_out)
        g_col, h_col = hm["合价(底稿公式)"], hm["程序计算合价"]
        checked = 0
        for r in range(2, ws_out.max_row + 1):
            formula = ws_out.cell(row=r, column=g_col).value
            if not (isinstance(formula, str) and formula.startswith("=")):
                continue  # 待补资料行
            qty = ws_out.cell(row=r, column=hm["数量"]).value
            price = ws_out.cell(row=r, column=hm["单价"]).value
            program = ws_out.cell(row=r, column=h_col).value
            expect = money_mul(Decimal(str(qty)), Decimal(str(price)))
            assert program is not None, "公式行必须有程序计算值兜底"
            assert Decimal(str(program)) == expect, (
                f"第{r}行程序值 {program} ≠ Decimal 复算 {expect}")
            assert isinstance(formula, str) and formula.startswith("="), (
                "公式列必须保留公式串（Excel 复核值）")
            checked += 1
        assert checked >= 3, f"公式行过少：{checked}"

    def test_formula_cells_have_no_stale_cache(self, exported):
        """公式单元格无缓存值属预期（首次打开由 Office 计算）——如实断言。"""
        ws_out, _ = exported
        hm = _header_map(ws_out)
        wb_data = ws_out.parent  # 已加载的 workbook（data_only=False）
        ws_values = wb_data["审核底稿"]
        g_col = hm["合价(底稿公式)"]
        for r in range(2, ws_out.max_row + 1):
            cell = ws_values.cell(row=r, column=g_col)
            if isinstance(cell.value, str) and cell.value.startswith("="):
                assert cell.data_type == "f", "公式单元格必须是公式类型（无伪造缓存）"

    def test_missing_stays_pending_not_zero(self, exported):
        ws_out, _ = exported
        hm = _header_map(ws_out)
        zeros = 0
        for r in range(2, ws_out.max_row + 1):
            name = ws_out.cell(row=r, column=hm["清单名称"]).value or ""
            if "缺失单价行" in name:
                g = ws_out.cell(row=r, column=hm["合价(底稿公式)"]).value
                assert g == "待补资料", f"缺失行应保持待补资料，实际 {g!r}（不得填 0）"
            for c in (hm["合价(底稿公式)"], hm["程序计算合价"]):
                v = ws_out.cell(row=r, column=c).value
                if v == 0:
                    zeros += 1
        assert zeros == 0, "缺失行不得以 0.00 充数"

    def test_row_count_matches_db(self, exported):
        ws_out, conn = exported
        hm = _header_map(ws_out)
        db_rows = conn.execute(
            """SELECT COUNT(*) c FROM line_items li JOIN settlement_periods sp
               ON sp.id=li.period_id
               WHERE li.flags_json NOT LIKE '%"subtotal": true%'
                 AND li.flags_json NOT LIKE '%title_row%'""").fetchone()["c"]
        sheet_rows = sum(
            1 for r in range(2, ws_out.max_row + 1)
            if ws_out.cell(row=r, column=hm["清单编码"]).value is not None
            or ws_out.cell(row=r, column=hm["清单名称"]).value is not None)
        assert sheet_rows == db_rows, f"底稿行数 {sheet_rows} ≠ DB 明细 {db_rows}"

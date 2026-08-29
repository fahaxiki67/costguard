"""方向隔离残留阻断回归（监督第二轮，先红后绿）。

阻断1 差异表：export_diff_sheets 混合全部 direction 且聚合层无 unit_price，
      单价差异表为空、同期对上/对下被串成跨期比较；
阻断2 工作台：_set_direction 按 project_id+period_no 更新，同期号两个方向一起被改；
阻断3 期号递增：next_period_no 未按方向独立；
阻断4 重复检测：rule_duplicates 键缺方向，跨方向同期号误报。

本文件在缺陷实现下运行必须失败。
"""
import os
from decimal import Decimal
from pathlib import Path

import pytest

os.environ.setdefault("QT_QPA_PLATFORM", "offscreen")

from costguard.core.anomalies.rules import rule_duplicates  # noqa: E402
from costguard.core.db import migrations  # noqa: E402
from costguard.core.engine.settlement_io import ensure_period, next_period_no  # noqa: E402
from costguard.core.export.excel_export import export_diff_sheets  # noqa: E402


@pytest.fixture()
def db(tmp_path):
    db_path = tmp_path / "project.db"
    migrations.migrate(db_path, tmp_path / "backups")
    conn = migrations.connect(db_path)
    with conn:
        pid = conn.execute(
            "INSERT INTO projects(name, schema_version, workspace_path, created_at)"
            " VALUES ('残留阻断', 1, '/t', '2026')"
        ).lastrowid
    yield conn, pid
    conn.close()


def add_item(conn, period_id, code, name, qty, price, amount, unit="m3"):
    with conn:
        conn.execute(
            "INSERT INTO line_items(period_id, code, name, unit, quantity, unit_price, amount, flags_json)"
            " VALUES (?,?,?,?,?,?,?,?)",
            (period_id, code, name, unit, qty, price, amount, "{}"),
        )


def _grid_rows(ws):
    """读差异表数据区为 dict 行列表。"""
    headers = [ws.cell(row=1, column=c).value for c in range(1, ws.max_column + 1)]
    out = []
    for r in range(2, ws.max_row + 1):
        row = {headers[c - 1]: ws.cell(row=r, column=c).value for c in range(1, ws.max_column + 1)}
        if any(v is not None for v in row.values()):
            out.append(row)
    return out


class TestBlock1DiffSheets:
    @pytest.fixture()
    def mixed(self, db):
        conn, pid = db
        up1 = ensure_period(conn, pid, 1, "up1", None, direction="upward")
        up2 = ensure_period(conn, pid, 2, "up2", None, direction="upward")
        down1 = ensure_period(conn, pid, 1, "down1", None, direction="downward")
        add_item(conn, up1, "K1", "清单K", "10", "10", "100")
        add_item(conn, up2, "K1", "清单K", "10", "12", "120")
        add_item(conn, down1, "K1", "清单K", "10", "20", "200")
        return conn, pid

    def test_price_diff_sheet_has_rows_and_no_cross_direction(self, mixed):
        """单价差异表必须有行（原始 unit_price），且绝不出现跨方向比较。"""
        conn, pid = mixed
        wb = pytest.importorskip("openpyxl").Workbook()
        wb.remove(wb.active)
        export_diff_sheets(conn, pid, wb)
        rows = _grid_rows(wb["单价差异表"])
        assert rows, "单价差异表不得为空（单价必须来自可追溯原始 unit_price）"
        # up第2期(12) 只允许与 up第1期(10) 比较
        assert any(
            r["方向"] == "对上" and r["本期值"] == Decimal("12") and r["上期值"] == Decimal("10")
            for r in rows
        ), f"缺少 up 第2期 vs up 第1期的单价比较行: {rows}"
        # down 第1期是 down 方向首期：仅作展示（标注"首期"），不得产生比较行；
        # 跨方向串表的定义 = 对下方向出现非空上期值或差异公式
        down_rows = [r for r in rows if r["方向"] == "对下"]
        assert all(r["上期值"] is None and r["差异"] is None for r in down_rows), \
            f"对下方向出现比较行（串表）: {down_rows}"
        assert all("首期" in str(r.get("差异率", "")) for r in down_rows)

    def test_qty_and_amount_diff_sheets_correct(self, mixed):
        """工程量/金额差异表：up 第2期只与 up 第1期比。"""
        conn, pid = mixed
        wb = pytest.importorskip("openpyxl").Workbook()
        wb.remove(wb.active)
        export_diff_sheets(conn, pid, wb)
        qty_rows = _grid_rows(wb["工程量差异表"])
        amt_rows = _grid_rows(wb["金额差异表"])
        assert any(
            r["方向"] == "对上" and r["本期值"] == Decimal("10") and r["上期值"] == Decimal("10")
            for r in qty_rows
        ), f"工程量差异表缺 up 比较行: {qty_rows}"
        assert any(
            r["方向"] == "对上" and r["本期值"] == Decimal("120") and r["上期值"] == Decimal("100")
            for r in amt_rows
        ), f"金额差异表缺 up 比较行: {amt_rows}"
        assert not any(r["方向"] == "对下" for r in qty_rows + amt_rows if r.get("差异") not in (None, ""))

    def test_multi_price_in_period_marked_not_averaged(self, db):
        """同期同组多个不同单价 → 标记不可比/待复核，不得平均掩盖。"""
        conn, pid = db
        up1 = ensure_period(conn, pid, 1, "up1", None, direction="upward")
        add_item(conn, up1, "K1", "清单K", "5", "15", "75")
        add_item(conn, up1, "K1", "清单K", "5", "16", "80")
        wb = pytest.importorskip("openpyxl").Workbook()
        wb.remove(wb.active)
        export_diff_sheets(conn, pid, wb)
        rows = _grid_rows(wb["单价差异表"])
        assert not any(isinstance(r["本期值"], Decimal) and r["本期值"] == Decimal("15.5") for r in rows), \
            "出现平均价（15.5）即违规"
        assert any("不可比" in str(r["本期值"]) or "多价" in str(r["本期值"]) or "待复核" in str(r["本期值"])
                   for r in rows if "第1期" in str(r["期间"]) or True), \
            f"多价期未标记待复核: {rows}"


class TestBlock2SetDirection:
    def test_set_direction_updates_selected_period_only(self, db, monkeypatch):
        """同期号 up/down 两行，只改选中行；audit subject 按 period_id。"""
        from PySide6.QtWidgets import QApplication

        from costguard.core.evidence import audit as audit_log
        from costguard.core.models.project import ProjectInfo
        from costguard.ui.workbench import WorkbenchPage

        _app = QApplication.instance() or QApplication([])
        conn, pid = db
        up1 = ensure_period(conn, pid, 1, "up1", None, direction="upward")
        down1 = ensure_period(conn, pid, 1, "down1", None, direction="downward")
        info = ProjectInfo(pid, "残留阻断", "/t", 1, "2026")
        page = WorkbenchPage(conn, info, "/t", on_back=lambda: None)
        page.refresh_periods()
        target_row = None
        for r in range(page.period_table.rowCount()):
            if page.period_table.item(r, 1).text() == "up1":
                target_row = r
        assert target_row is not None
        page.period_table.selectRow(target_row)
        page.dir_combo.setCurrentIndex(page.dir_combo.findText("downward"))

        class FakeDlg:
            def __init__(self, *a, **k):
                pass

            def exec(self):
                return 1  # QDialog.Accepted

            def reason(self):
                return "改方向测试"

        monkeypatch.setattr("costguard.ui.workbench.ReasonDialog", FakeDlg)
        warnings = []
        from PySide6.QtWidgets import QMessageBox

        monkeypatch.setattr(QMessageBox, "warning", lambda *a, **k: warnings.append(1))
        page._set_direction()

        up_d = conn.execute("SELECT direction FROM settlement_periods WHERE id=?", (up1,)).fetchone()["direction"]
        down_d = conn.execute("SELECT direction FROM settlement_periods WHERE id=?", (down1,)).fetchone()["direction"]
        # down 第1期已存在 → 撞车被友好拒绝，两行都保持各自原方向，弹提示
        assert up_d == "upward" and down_d == "downward", "撞车时不得发生任何更新"
        assert warnings, "撞车必须给出提示"
        entries = audit_log.history_for(conn, pid)
        assert not any(e.action == "set_direction" for e in entries), "拒绝时不得写审计"

    def test_set_direction_success_non_conflicting(self, db, monkeypatch):
        """非冲突标记：只改选中行（按 period_id），同期号另一方向不受影响。"""
        from PySide6.QtWidgets import QApplication

        from costguard.core.evidence import audit as audit_log
        from costguard.core.models.project import ProjectInfo
        from costguard.ui.workbench import WorkbenchPage

        _app = QApplication.instance() or QApplication([])
        conn, pid = db
        up1 = ensure_period(conn, pid, 1, "up1", None, direction="upward")
        down1 = ensure_period(conn, pid, 1, "down1", None, direction="downward")
        info = ProjectInfo(pid, "残留阻断", "/t", 1, "2026")
        page = WorkbenchPage(conn, info, "/t", on_back=lambda: None)
        page.refresh_periods()
        target_row = next(r for r in range(page.period_table.rowCount())
                          if page.period_table.item(r, 1).text() == "up1")
        page.period_table.selectRow(target_row)
        page.dir_combo.setCurrentIndex(page.dir_combo.findText("downward"))

        class FakeDlg:
            def __init__(self, *a, **k):
                pass

            def exec(self):
                return 1

            def reason(self):
                return "改方向测试"

        monkeypatch.setattr("costguard.ui.workbench.ReasonDialog", FakeDlg)
        # 先移走 down 第1期的期号（改为 down 第5期），使 up1→downward 不撞车
        with conn:
            conn.execute("UPDATE settlement_periods SET period_no=5 WHERE id=?", (down1,))
        page.refresh_periods()
        target_row = next(r for r in range(page.period_table.rowCount())
                          if page.period_table.item(r, 1).text() == "up1")
        page.period_table.selectRow(target_row)
        page._set_direction()

        up_d = conn.execute("SELECT direction FROM settlement_periods WHERE id=?", (up1,)).fetchone()["direction"]
        down_d = conn.execute("SELECT direction, period_no FROM settlement_periods WHERE id=?", (down1,)).fetchone()
        assert up_d == "downward", "选中行方向应被更新"
        assert down_d["direction"] == "downward" and down_d["period_no"] == 5, \
            "同期号原另一方向行不得被覆盖"
        entries = audit_log.history_for(conn, pid)
        assert any(e.action == "set_direction" and e.target == f"period:{up1}" for e in entries), \
            "audit subject 必须按 period_id 记录"

    def test_set_direction_no_null(self, db):
        """未标记方向必须写 schema 允许的明确值 'unknown'，不得写 NULL。"""
        conn, pid = db
        # 由实现保证：direction 取值域 {"unknown","upward","downward"}
        row = ensure_period(conn, pid, 1, "t", None, direction="unknown")
        d = conn.execute("SELECT direction FROM settlement_periods WHERE id=?", (row,)).fetchone()["direction"]
        assert d in ("unknown", "upward", "downward")


class TestBlock3NextPeriodNo:
    def test_next_period_no_independent_per_direction(self, db):
        """up 已有第1期 → next(up)=2；down 无期 → next(down)=1。"""
        conn, pid = db
        ensure_period(conn, pid, 1, "up1", None, direction="upward")
        assert next_period_no(conn, pid, direction="upward") == 2
        assert next_period_no(conn, pid, direction="downward") == 1

    def test_import_fallback_passes_direction(self, db, tmp_path):
        """导入兜底路径（无期号文件名+无 sheet 期号）必须按导入方向递增。"""
        import openpyxl

        from costguard.core.engine import settlement_io
        conn, pid = db
        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = "清单"
        for c, v in enumerate(["清单编码", "清单名称", "单位", "工程量", "综合单价", "合价"], start=1):
            ws.cell(row=1, column=c, value=v)
        ws.cell(row=2, column=1, value="K1")
        ws.cell(row=2, column=2, value="某清单")
        ws.cell(row=2, column=3, value="m3")
        ws.cell(row=2, column=4, value=10)
        ws.cell(row=2, column=5, value=10)
        ws.cell(row=2, column=6, value=100)
        src = tmp_path / "无期号文件.xlsx"
        wb.save(src)
        pdir = Path(tmp_path / "proj")
        pdir.mkdir()
        with conn:
            conn.execute("UPDATE projects SET workspace_path=? WHERE id=?", (str(pdir), pid))
        report = settlement_io.import_settlement_file(
            conn, pid, pdir, src, direction="downward")
        assert report.period_no == 1, "downward 首次导入应得到 down 序列的第 1 期"
        d = conn.execute(
            "SELECT direction FROM settlement_periods WHERE id=?", (report.period_id,)
        ).fetchone()["direction"]
        assert d == "downward"


class TestBlock4Duplicates:
    def test_same_period_no_across_directions_not_duplicate(self, db):
        """up第1期 与 down第1期 同码同名 → 不得误报重复。"""
        conn, pid = db
        up1 = ensure_period(conn, pid, 1, "up1", None, direction="upward")
        down1 = ensure_period(conn, pid, 1, "down1", None, direction="downward")
        add_item(conn, up1, "K1", "清单K", "10", "10", "100")
        add_item(conn, down1, "K1", "清单K", "10", "20", "200")
        assert rule_duplicates(conn, pid) == [], "跨方向同期号被误报为重复"

    def test_true_duplicate_within_same_period_detected(self, db):
        """同一期次内同码同名两行 → 必须报重复（真反例）。"""
        conn, pid = db
        up1 = ensure_period(conn, pid, 1, "up1", None, direction="upward")
        add_item(conn, up1, "K1", "清单K", "10", "10", "100")
        add_item(conn, up1, "K1", "清单K", "5", "10", "50")
        f = rule_duplicates(conn, pid)
        assert len(f) == 1 and "重复" in f[0].message

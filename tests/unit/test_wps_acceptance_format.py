from __future__ import annotations

import importlib.util
from pathlib import Path

from openpyxl import load_workbook


def _load_generator():
    root = Path(__file__).resolve().parents[2]
    path = root / "scripts" / "generate_wps_acceptance.py"
    spec = importlib.util.spec_from_file_location("generate_wps_acceptance", path)
    assert spec and spec.loader
    module = importlib.util.module_from_spec(spec)
    spec.loader.exec_module(module)
    return module


def test_wps_acceptance_workbook_has_borders_and_centered_cells(tmp_path):
    generator = _load_generator()
    xlsx_path, _ = generator.generate(tmp_path)
    wb = load_workbook(xlsx_path, data_only=False)

    for sheet_name, max_row, max_column in (("审核底稿", 7, 9), ("汇总校验", 4, 2)):
        ws = wb[sheet_name]
        for row in ws.iter_rows(min_row=1, max_row=max_row, min_col=1, max_col=max_column):
            for cell in row:
                assert cell.border.left.style == "thin"
                assert cell.border.right.style == "thin"
                assert cell.border.top.style == "thin"
                assert cell.border.bottom.style == "thin"
                assert cell.alignment.horizontal == "center"
                assert cell.alignment.vertical == "center"

    assert wb["审核底稿"]["F2"].value == "=D2*E2"
    assert wb["审核底稿"]["H2"].value == "=ROUND(F2-G2,2)"
    assert wb["汇总校验"]["B2"].value == "=SUM(审核底稿!G2:G7)"

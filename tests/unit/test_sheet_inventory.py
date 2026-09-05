"""全 Sheet 清单与类型建议测试（任务书任务 B1-B6）。"""

import pytest

from jiadun.core.db import migrations
from jiadun.core.engine import sheet_inventory


@pytest.fixture()
def project_db(tmp_path):
    db_path = tmp_path / "project.db"
    migrations.migrate(db_path, tmp_path / "backups")
    conn = migrations.connect(db_path)
    with conn:
        pid = conn.execute(
            """INSERT INTO projects(name, schema_version, workspace_path, created_at)
               VALUES (?,?,?,?)""",
            ("Sheet清单测试", migrations.LATEST_SCHEMA_VERSION, str(tmp_path), "2026"),
        ).lastrowid
    yield conn, int(pid), tmp_path
    conn.close()


def _add_file_with_sheets(conn, pid, tmp_path, name, sheets):
    """登记文件与批次并写入若干 Sheet（name, status, reason, n_rows）。"""
    from jiadun.core.models.source_file import import_file

    src = tmp_path / name
    src.write_bytes(b"placeholder")
    sf = import_file(conn, pid, tmp_path, src)
    now = "2026-09-05T00:00:00"
    with conn:
        cur = conn.execute(
            """INSERT INTO parse_batches(file_id, parser, parsed_at, status, stats_json)
               VALUES (?,?,?,?,?)""",
            (sf.file_id, "pipeline", now, "ok", "{}"),
        )
        batch_id = cur.lastrowid
        for index, (sheet_name, status, reason, n_rows) in enumerate(sheets, start=1):
            conn.execute(
                """INSERT INTO raw_sheets(batch_id, sheet_index, sheet_name,
                   n_rows, n_cols, sheet_status, sheet_status_reason)
                   VALUES (?,?,?,?,?,?,?)""",
                (batch_id, index, sheet_name, n_rows, 8, status, reason),
            )
    return sf.file_id


class TestListWorkbookSheets:
    def test_lists_all_sheets_including_confirmed_and_non_pending(self, project_db):
        conn, pid, tmp = project_db
        _add_file_with_sheets(
            conn, pid, tmp, "多Sheet工作簿.xlsx",
            [
                ("分部分项清单", "pending", "表头歧义待人工确认", 120),
                ("总价措施项目清单与计价表", "pending", "汇总样式", 20),
                ("封面", "confirmed", "非业务页已确认", 3),
                ("汇总表", "confirmed", "汇总页已确认", 10),
            ],
        )
        sheets = sheet_inventory.list_workbook_sheets(conn, pid)
        assert len(sheets) == 4  # 全列表：pending + confirmed 都在
        by_name = {s["sheet_name"]: s for s in sheets}
        assert by_name["分部分项清单"]["suggested_kind"] == "boq_detail"
        assert "分部分项" in by_name["分部分项清单"]["suggest_reason"]
        assert by_name["总价措施项目清单与计价表"]["suggested_kind"] == "measure_total"
        assert "费率计取" in by_name["总价措施项目清单与计价表"]["suggest_reason"]
        assert by_name["封面"]["suggested_kind"] == "non_business"

    def test_confirmed_sheet_keeps_human_kind(self, project_db):
        conn, pid, tmp = project_db
        _add_file_with_sheets(
            conn, pid, tmp, "工作簿.xlsx", [("汇总表", "confirmed", "已确认", 10)]
        )
        conn.execute(
            "UPDATE raw_sheets SET list_kind='measure_unit' WHERE sheet_name='汇总表'"
        )
        sheets = sheet_inventory.list_workbook_sheets(conn, pid)
        assert sheets[0]["suggested_kind"] == "measure_unit"

    def test_filters(self, project_db):
        conn, pid, tmp = project_db
        fid = _add_file_with_sheets(
            conn, pid, tmp, "工作簿.xlsx",
            [
                ("分部分项清单", "pending", "待确认", 50),
                ("封面", "confirmed", "已确认", 2),
            ],
        )
        by_status = sheet_inventory.list_workbook_sheets(conn, pid, status="pending")
        assert [s["sheet_name"] for s in by_status] == ["分部分项清单"]
        by_kw = sheet_inventory.list_workbook_sheets(conn, pid, keyword="封面")
        assert [s["sheet_name"] for s in by_kw] == ["封面"]
        by_file = sheet_inventory.list_workbook_sheets(conn, pid, file_id=fid)
        assert len(by_file) == 2

    def test_unknown_when_no_feature(self, project_db):
        conn, pid, tmp = project_db
        _add_file_with_sheets(conn, pid, tmp, "工作簿.xlsx", [("Sheet1", "pending", "", 5)])
        sheets = sheet_inventory.list_workbook_sheets(conn, pid)
        assert sheets[0]["suggested_kind"] == "unknown"


class TestSetSheetListKind:
    def test_human_annotation_requires_reason(self, project_db):
        conn, pid, tmp = project_db
        _add_file_with_sheets(
            conn, pid, tmp, "工作簿.xlsx", [("清单", "pending", "待确认", 30)]
        )
        sheet_id = sheet_inventory.list_workbook_sheets(conn, pid)[0]["sheet_id"]
        with pytest.raises(ValueError, match="理由"):
            sheet_inventory.set_sheet_list_kind(conn, pid, sheet_id, "boq_detail")
        result = sheet_inventory.set_sheet_list_kind(
            conn, pid, sheet_id, "boq_detail", reason="表头含 编码/名称/工程量/单价/合价"
        )
        assert result["after"] == "boq_detail"
        kinds = {
            r["kind"] for r in conn.execute(
                "SELECT DISTINCT kind FROM evidence WHERE project_id=?", (pid,)
            ).fetchall()
        }
        assert "sheet_list_kind" in kinds

    def test_invalid_kind_and_foreign_sheet(self, project_db):
        conn, pid, tmp = project_db
        _add_file_with_sheets(conn, pid, tmp, "工作簿.xlsx", [("清单", "pending", "", 30)])
        sheet_id = sheet_inventory.list_workbook_sheets(conn, pid)[0]["sheet_id"]
        with pytest.raises(ValueError, match="未知的清单类型"):
            sheet_inventory.set_sheet_list_kind(conn, pid, sheet_id, "golden")
        with pytest.raises(ValueError, match="不存在或不属于当前项目"):
            sheet_inventory.set_sheet_list_kind(conn, pid, 999999, "boq_detail", reason="x")


class TestRoleEnumExtension:
    """B3：角色枚举扩展（对上明细/对下明细/其他费用）与置信度分级。"""

    def test_new_kinds_accepted_and_suggested(self, project_db):
        conn, pid, tmp = project_db
        _add_file_with_sheets(
            conn, pid, tmp, "工作簿.xlsx",
            [
                ("对上结算台账", "pending", "", 40),
                ("对下分包结算明细", "pending", "", 60),
                ("其他费用清单", "pending", "", 12),
            ],
        )
        sheets = sheet_inventory.list_workbook_sheets(conn, pid)
        by_name = {s["sheet_name"]: s for s in sheets}
        assert by_name["对上结算台账"]["suggested_kind"] == "upstream_detail"
        assert by_name["对下分包结算明细"]["suggested_kind"] == "downstream_detail"
        assert by_name["其他费用清单"]["suggested_kind"] == "other_fee"
        sheet_id = by_name["对上结算台账"]["sheet_id"]
        result = sheet_inventory.set_sheet_list_kind(
            conn, pid, sheet_id, "other_fee", reason="人工改判为其他费用"
        )
        assert result["after"] == "other_fee"

    def test_confidence_levels(self, project_db):
        conn, pid, tmp = project_db
        _add_file_with_sheets(
            conn, pid, tmp, "工作簿.xlsx",
            [
                ("分部分项工程清单计价表", "pending", "", 100),
                ("对上结算台账", "pending", "", 40),
                ("Sheet1", "pending", "", 5),
            ],
        )
        by_name = {s["sheet_name"]: s for s in sheet_inventory.list_workbook_sheets(conn, pid)}
        # GB50500 特征词命中 Sheet 名 → 高；弱特征（对上/其他费用）→ 中；无特征 → 低
        assert by_name["分部分项工程清单计价表"]["suggest_confidence"] == "高"
        assert by_name["对上结算台账"]["suggest_confidence"] == "中"
        assert by_name["Sheet1"]["suggest_confidence"] == "低"
        assert by_name["Sheet1"]["suggested_kind"] == "unknown"

    def test_invalid_new_filters_and_modes(self, project_db):
        conn, pid, tmp = project_db
        with pytest.raises(ValueError, match="未知的过滤模式"):
            sheet_inventory.list_workbook_sheets(conn, pid, filter_mode="bogus")


class TestFilterModes:
    """B2：仅待确认 / 仅建议参与分析 / 全部。"""

    def _seed(self, conn, pid, tmp):
        _add_file_with_sheets(
            conn, pid, tmp, "工作簿.xlsx",
            [
                ("分部分项清单", "pending", "待确认", 50),
                ("总价措施项目清单", "confirmed", "已确认", 20),
                ("封面", "non_business_ok", "已确认非业务", 2),
                ("随便一个名字", "pending", "", 5),
            ],
        )

    def test_filter_pending(self, project_db):
        conn, pid, tmp = project_db
        self._seed(conn, pid, tmp)
        pending = sheet_inventory.list_workbook_sheets(conn, pid, filter_mode="pending")
        assert {s["sheet_name"] for s in pending} == {"分部分项清单", "随便一个名字"}

    def test_filter_suggested_detail_only(self, project_db):
        conn, pid, tmp = project_db
        self._seed(conn, pid, tmp)
        suggested = sheet_inventory.list_workbook_sheets(conn, pid, filter_mode="suggested")
        # 建议参与分析 = 明细类建议（分部分项/措施费），不含未知特征页与非业务页
        assert {s["sheet_name"] for s in suggested} == {"分部分项清单", "总价措施项目清单"}

    def test_filter_all_default(self, project_db):
        conn, pid, tmp = project_db
        self._seed(conn, pid, tmp)
        everything = sheet_inventory.list_workbook_sheets(conn, pid)
        assert len(everything) == 4


class TestVisibleState:
    """B1：工作簿级可见状态——解析器捕获、清单展示、历史批次未知。"""

    def test_old_batch_reports_unknown(self, project_db):
        conn, pid, tmp = project_db
        _add_file_with_sheets(conn, pid, tmp, "工作簿.xlsx", [("清单", "pending", "", 30)])
        sheet = sheet_inventory.list_workbook_sheets(conn, pid)[0]
        assert sheet["visible_state"] is None  # 展示层据此显示「未知」，不得默认可见

    def test_parser_captures_hidden_sheet(self, tmp_path):
        import openpyxl

        from jiadun.core.parsing import excel_parser

        src = tmp_path / "含隐藏页.xlsx"
        wb = openpyxl.Workbook()
        ws1 = wb.active
        ws1["A1"] = "明细"
        ws2 = wb.create_sheet("隐藏的辅助表")
        ws2["A1"] = "x"
        ws2.sheet_state = "hidden"
        wb.save(src)
        result = excel_parser.parse_file(src, "xlsx")
        by_name = {s.sheet_name: s for s in result.sheets}
        assert by_name["Sheet"].visible_state == "visible"
        assert by_name["隐藏的辅助表"].visible_state == "hidden"

    def test_persist_and_list_report_visible_state(self, project_db, tmp_path):
        import openpyxl

        from jiadun.core.models.source_file import import_file
        from jiadun.core.parsing import excel_parser

        conn, pid, _ = project_db
        src = tmp_path / "真实可见性.xlsx"
        wb = openpyxl.Workbook()
        wb.active["A1"] = "数据"
        hidden = wb.create_sheet("汇总-隐藏")
        hidden["A1"] = "y"
        hidden.sheet_state = "hidden"
        wb.save(src)
        sf = import_file(conn, pid, tmp_path, src)
        result = excel_parser.parse_file(src, "xlsx")
        batch_id = excel_parser.persist_parse_result(conn, sf.file_id, result)
        sheets = sheet_inventory.list_workbook_sheets(conn, pid, file_id=sf.file_id)
        by_name = {s["sheet_name"]: s for s in sheets}
        assert by_name["Sheet"]["visible_state"] == "visible"
        assert by_name["汇总-隐藏"]["visible_state"] == "hidden"
        assert batch_id > 0


class TestThirtySheets:
    """B6 场景 2：30 个 Sheet 全量列出。"""

    def test_lists_thirty_sheets(self, project_db):
        conn, pid, tmp = project_db
        sheets = [(f"清单{i:02d}", "pending", "", 10) for i in range(1, 31)]
        _add_file_with_sheets(conn, pid, tmp, "大工作簿.xlsx", sheets)
        result = sheet_inventory.list_workbook_sheets(conn, pid)
        assert len(result) == 30
        assert all(r["sheet_status"] == "pending" for r in result)


class TestCarryForward:
    """B4：重解析人工决策结转（同名 + 单元格摘要一致才结转）。"""

    def _setup_two_batches(self, conn, pid, tmp_path, second_content=(("a", 1),),
                           extra_new_sheets=()):
        """建文件与两个批次；批次1 含已确认+人工标注的 Sheet（a=已确认,
        changed=待确认）；批次2 按 second_content 建带单元格的新页。"""
        from jiadun.core.models.source_file import import_file

        src = tmp_path / "重导工作簿.xlsx"
        src.write_bytes(b"placeholder")
        sf = import_file(conn, pid, tmp_path, src)
        with conn:
            for batch_no in (1, 2):
                cur = conn.execute(
                    """INSERT INTO parse_batches(file_id, parser, parsed_at, status, stats_json)
                       VALUES (?,?,?,?,?)""",
                    (sf.file_id, "pipeline", f"2026-09-05T0{batch_no}:00:00", "ok", "{}"),
                )
                batch_id = cur.lastrowid
                if batch_no == 2:
                    entries = list(second_content) + [
                        (name, 0) for name in extra_new_sheets
                    ]
                    for row_off, (val, num) in enumerate(entries, start=1):
                        sheet_cur = conn.execute(
                            """INSERT INTO raw_sheets(batch_id, sheet_index, sheet_name,
                               n_rows, n_cols) VALUES (?,?,?,?,?)""",
                            (batch_id, row_off, val, 5, 3),
                        )
                        sheet_id = sheet_cur.lastrowid
                        for col in range(1, num + 1):
                            conn.execute(
                                """INSERT INTO raw_cells(sheet_id, row, col, raw_value,
                                   cached_value, is_formula, is_number_stored_as_text, num_fmt)
                                   VALUES (?,?,?,?,?,?,?,?)""",
                                (sheet_id, 1, col, f"v{col}", f"c{col}", 0, 0, ""),
                            )
                else:
                    conn.execute(
                        """INSERT INTO raw_sheets(batch_id, sheet_index, sheet_name,
                           n_rows, n_cols, sheet_status, sheet_status_reason, list_kind)
                           VALUES (?,?,?,?,?,?,?,?)""",
                        (batch_id, 1, "a", 5, 3, "confirmed", "已人工确认", "boq_detail"),
                    )
                    conn.execute(
                        """INSERT INTO raw_sheets(batch_id, sheet_index, sheet_name,
                           n_rows, n_cols, sheet_status, sheet_status_reason)
                           VALUES (?,?,?,?,?,?,?)""",
                        (batch_id, 2, "changed", 4, 3, "pending", "待确认"),
                    )
        return sf.file_id

    def test_identical_content_carries_decisions(self, project_db):
        conn, pid, tmp = project_db
        fid = self._setup_two_batches(conn, pid, tmp)
        old_sheet = conn.execute(
            "SELECT id FROM raw_sheets WHERE sheet_name='a' ORDER BY batch_id LIMIT 1"
        ).fetchone()["id"]
        # 旧批次 a 无 raw_cells（空摘要）；给旧 a 也补相同单元格
        conn.execute(
            """INSERT INTO raw_cells(sheet_id, row, col, raw_value, cached_value,
               is_formula, is_number_stored_as_text, num_fmt)
               VALUES (?,?,?,?,?,?,?,?)""",
            (old_sheet, 1, 1, "v1", "c1", 0, 0, ""),
        )
        new_batch = conn.execute(
            "SELECT MAX(id) AS m FROM parse_batches WHERE file_id=?", (fid,)
        ).fetchone()["m"]
        summary = sheet_inventory.carry_forward_sheet_decisions(conn, pid, fid, int(new_batch))
        assert summary["carried"] == 1
        new_sheet = conn.execute(
            "SELECT sheet_status, list_kind FROM raw_sheets WHERE sheet_name='a' AND batch_id=?",
            (int(new_batch),),
        ).fetchone()
        assert new_sheet["sheet_status"] == "confirmed"
        assert new_sheet["list_kind"] == "boq_detail"
        kinds = {
            r["kind"] for r in conn.execute(
                "SELECT DISTINCT kind FROM evidence WHERE project_id=?", (pid,)
            ).fetchall()
        }
        assert "sheet_decision_carry_forward" in kinds

    def test_content_change_not_carried(self, project_db):
        conn, pid, tmp = project_db
        fid = self._setup_two_batches(conn, pid, tmp)
        old_sheet = conn.execute(
            "SELECT id FROM raw_sheets WHERE sheet_name='a' ORDER BY batch_id LIMIT 1"
        ).fetchone()["id"]
        conn.execute(
            """INSERT INTO raw_cells(sheet_id, row, col, raw_value, cached_value,
               is_formula, is_number_stored_as_text, num_fmt)
               VALUES (?,?,?,?,?,?,?,?)""",
            (old_sheet, 1, 1, "不同内容", "c1", 0, 0, ""),
        )
        new_batch = conn.execute(
            "SELECT MAX(id) AS m FROM parse_batches WHERE file_id=?", (fid,)
        ).fetchone()["m"]
        summary = sheet_inventory.carry_forward_sheet_decisions(conn, pid, fid, int(new_batch))
        assert summary["carried"] == 0
        assert summary["skipped_content_changed"] == 1
        new_status = conn.execute(
            "SELECT sheet_status FROM raw_sheets WHERE sheet_name='a' AND batch_id=?",
            (int(new_batch),),
        ).fetchone()["sheet_status"]
        assert new_status == "pending"  # 未结转，保持机器默认的待确认

    def test_unmatched_name_skipped(self, project_db):
        conn, pid, tmp = project_db
        fid = self._setup_two_batches(
            conn, pid, tmp, extra_new_sheets=("全新无旧页的表",)
        )
        new_batch = conn.execute(
            "SELECT MAX(id) AS m FROM parse_batches WHERE file_id=?", (fid,)
        ).fetchone()["m"]
        summary = sheet_inventory.carry_forward_sheet_decisions(conn, pid, fid, int(new_batch))
        # a：旧批次无 cells、新批次有 cells → 内容不同；全新页无同名旧页 → unmatched
        assert summary["skipped_content_changed"] == 1
        assert summary["skipped_unmatched"] == 1
        assert summary["carried"] == 0


class TestListKindInvalidatesRunContract:
    """B5：清单类型（角色）变化必须使 Run Contract 签名变化、旧运行失效。"""

    def test_kind_change_changes_signature(self, project_db):
        from jiadun.core.contracts import run_contract

        conn, pid, tmp = project_db
        _add_file_with_sheets(conn, pid, tmp, "工作簿.xlsx", [("清单", "pending", "", 30)])
        first = run_contract.ensure_run_contract(conn, pid)
        sheet_id = sheet_inventory.list_workbook_sheets(conn, pid)[0]["sheet_id"]
        sheet_inventory.set_sheet_list_kind(conn, pid, sheet_id, "boq_detail", reason="人工确认")
        second = run_contract.ensure_run_contract(conn, pid)
        assert first.signature != second.signature
        assert first.run_id != second.run_id
        old_row = conn.execute(
            "SELECT invalidated_at FROM run_contracts WHERE run_id=?", (first.run_id,)
        ).fetchone()
        assert old_row["invalidated_at"] is not None


class TestSheetDigestSharedImplementation:
    """摘要算法唯一实现：engine.sheet_digest 与缓存表行为。"""

    def test_digest_stable_and_cached(self, project_db):
        from jiadun.core.engine.sheet_digest import sheet_cell_digest

        conn, pid, tmp = project_db
        _add_file_with_sheets(conn, pid, tmp, "工作簿.xlsx", [("清单", "pending", "", 3)])
        sheet_id = sheet_inventory.list_workbook_sheets(conn, pid)[0]["sheet_id"]
        conn.execute(
            """INSERT INTO raw_cells(sheet_id, row, col, raw_value, cached_value,
               is_formula, is_number_stored_as_text, num_fmt)
               VALUES (?,?,?,?,?,?,?,?)""",
            (sheet_id, 1, 1, "x", "y", 0, 0, ""),
        )
        first = sheet_cell_digest(conn, sheet_id)
        cached = conn.execute(
            "SELECT digest FROM sheet_cell_digests WHERE sheet_id=?", (sheet_id,)
        ).fetchone()
        assert cached["digest"] == first
        conn.execute("DELETE FROM sheet_cell_digests WHERE sheet_id=?", (sheet_id,))
        second = sheet_cell_digest(conn, sheet_id)
        assert first == second  # 缓存清空重建后摘要不变（算法确定性）


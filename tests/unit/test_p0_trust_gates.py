"""P0/P1 可信度闸门回归。

这些测试刻意构造数值可以彼此对上的输入，验证系统不会把同源复算、重复
明细或负数调整误报为项目级绿色通过；同时检查每条校核路径都留下可回读
的来源范围。
"""
from __future__ import annotations

import json
from decimal import Decimal
from pathlib import Path

import openpyxl
from tests.unit.test_review_gates import _import_project, _make_book


def _duplicate_book(path: Path) -> None:
    _make_book(path)
    wb = openpyxl.load_workbook(path)
    ws = wb[wb.sheetnames[0]]
    # 在两个控制行之前复制一条明细；同步控制值使 A/B/C 数值仍能碰巧相等。
    ws.insert_rows(4, 1)
    for col, value in enumerate(("A1", "项目A（重复）", 2, 100, 200), 1):
        ws.cell(4, col, value=value)
    ws.cell(5, 5, value=700)
    ws.cell(6, 5, value=700)
    wb.save(path)


def _negative_adjustment_book(path: Path) -> None:
    _make_book(path)
    wb = openpyxl.load_workbook(path)
    ws = wb[wb.sheetnames[0]]
    # 合法与否不能由程序臆断；即便控制值同步为 -100，也必须进入人工复核。
    ws.cell(3, 3, value=-1)
    ws.cell(3, 5, value=-300)
    ws.cell(4, 5, value=-100)
    ws.cell(5, 5, value=-100)
    wb.save(path)


def _all_missing_amount_book(path: Path) -> None:
    _make_book(path)
    wb = openpyxl.load_workbook(path)
    ws = wb[wb.sheetnames[0]]
    ws.cell(2, 5).value = None
    ws.cell(3, 5).value = None
    # 控制行也不提供可冒充的“0”金额。
    ws.cell(4, 5).value = None
    ws.cell(5, 5).value = None
    wb.save(path)


def _internal_blank_title_book(path: Path) -> None:
    """数据区内部的空行/分部标题不能被默认当作已证明的连续明细。"""
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "第1期明细"
    for col, value in enumerate(("项目编码", "项目名称", "工程量", "综合单价", "合价"), 1):
        ws.cell(1, col, value)
    ws.append(("A1", "项目A", 1, 100, 100))
    ws.append((None, None, None, None, None))
    ws.append((None, "分部二", None, None, None))
    ws.append(("A2", "项目B", 2, 200, 400))
    ws.append((None, "合计", None, None, 500))
    wb.save(path)


def test_clean_paths_record_independent_scope_and_evidence(tmp_path: Path):
    from jiadun.core.engine import crosscheck

    _info, conn, _report, period_id = _import_project(tmp_path)
    try:
        result = crosscheck.check_period(conn, period_id)
        assert result.verification_level == "sufficient", result.notes
        assert result.path_scopes["path_a"]["source_kind"] == "line_items"
        assert result.path_scopes["path_b"]["source_kind"] == "raw_cells"
        assert result.path_scopes["path_a"]["sheets"]
        assert result.path_scopes["path_b"]["sheets"]
        payload = json.loads(crosscheck.make_evidence(1, result))
        scope_step = next(step for step in payload["steps"] if step["step"] == "路径来源范围")
        assert scope_step["result"]["path_a"]["filters"]
        assert scope_step["result"]["path_b"]["filters"]
    finally:
        conn.close()


def test_line_item_mutation_cannot_contaminate_raw_b_path_or_green_status(
    tmp_path: Path,
):
    """清洗明细被污染时，B 仍须读取原始网格，项目级结果不得标绿。"""
    from jiadun.core.engine import crosscheck

    _info, conn, _report, period_id = _import_project(tmp_path)
    try:
        detail = conn.execute(
            """SELECT id, amount FROM line_items
               WHERE period_id=? AND json_extract(flags_json, '$.subtotal') IS NOT 1
               ORDER BY id LIMIT 1""",
            (period_id,),
        ).fetchone()
        assert detail is not None
        original_amount = Decimal(str(detail["amount"]))
        conn.execute(
            "UPDATE line_items SET amount='9999.99' WHERE id=?", (detail["id"],)
        )

        result = crosscheck.check_period(conn, period_id)

        assert result.path_a_total == Decimal("500") - original_amount + Decimal("9999.99")
        assert result.path_b_total == Decimal("500")
        assert result.path_a_total != result.path_b_total
        assert result.status == "diff"
        assert result.verification_level in {"findings", "insufficient"}
        assert result.path_scopes["path_b"]["source_kind"] == "raw_cells"
        assert any("A/B" in note or "差异" in note for note in result.notes)
    finally:
        conn.close()


def test_raw_layer_synchronous_omission_cannot_become_sufficient(tmp_path: Path):
    """原始网格与清洗明细同步漏行时，范围证明仍必须阻断绿色结论。"""
    from jiadun.core.engine import crosscheck

    _info, conn, _report, period_id = _import_project(tmp_path)
    try:
        detail = conn.execute(
            """SELECT id, sheet_id FROM line_items
               WHERE period_id=? AND json_extract(flags_json, '$.subtotal') IS NOT 1
               ORDER BY id LIMIT 1""",
            (period_id,),
        ).fetchone()
        assert detail is not None
        # 仅在隔离测试库中模拟旧库/外部 SQL 绕过不可变触发器的解析层
        # 同步漏行；生产数据库仍由触发器保护原始证据。
        conn.execute("DROP TRIGGER IF EXISTS trg_raw_cells_immutable_delete")
        conn.execute("DELETE FROM line_items WHERE id=?", (detail["id"],))
        conn.execute(
            "DELETE FROM raw_cells WHERE sheet_id=? AND row=2",
            (detail["sheet_id"],),
        )

        result = crosscheck.check_period(conn, period_id)

        assert result.verification_level in {"insufficient", "findings"}
        assert result.verification_level != "sufficient"
        assert result.path_a_total == result.path_b_total == Decimal("300")
        assert result.raw_subtotal == Decimal("500")
        assert result.control_status == "diff"
    finally:
        conn.close()


def test_parser_source_sheet_census_blocks_common_sheet_omission(tmp_path: Path, monkeypatch):
    """解析器共同漏掉整张 Sheet 时，入口范围证明必须降级而不能三路同绿。"""
    from jiadun.core.engine import crosscheck, settlement_io
    from jiadun.core.models import project as pm
    from jiadun.core.parsing import excel_parser

    src = tmp_path / "两张表.xlsx"
    wb = openpyxl.Workbook()
    first = wb.active
    first.title = "第1期"
    second = wb.create_sheet("第2期")
    for ws, amount in ((first, 500), (second, 30)):
        for col, value in enumerate(("项目编码", "项目名称", "工程量", "综合单价", "合价"), 1):
            ws.cell(1, col, value=value)
        ws.append((f"{ws.title}-A", "项目", 1, amount, amount))
        ws.append((None, "合计", None, None, amount))
    wb.save(src)

    real_parse = excel_parser.parse_xlsx

    def omit_last_sheet(path: Path):
        result = real_parse(path)
        result.sheets.pop()
        # 保留解析器自己的统计值，模拟“入口统计声称看到了两张表、落库只剩一张”。
        return result

    monkeypatch.setattr(excel_parser, "parse_xlsx", omit_last_sheet)
    info = pm.create_project("解析范围证明", tmp_path / "ws")
    info, conn = pm.open_project(Path(info.workspace_path))
    try:
        report = settlement_io.import_settlement_file(
            conn, info.project_id, Path(info.workspace_path), src, direction="upward"
        )
        batch = conn.execute(
            "SELECT status, stats_json FROM parse_batches WHERE id=?", (report.batch_id,)
        ).fetchone()
        assert batch is not None
        stats = json.loads(batch["stats_json"])
        assert stats["source_census"]["sheet_count"] == 2
        assert stats["source_census_status"] == "mismatch"
        assert batch["status"] == "partial"
        period_id = conn.execute(
            "SELECT id FROM settlement_periods WHERE project_id=?", (info.project_id,)
        ).fetchone()["id"]
        result = crosscheck.check_period(conn, period_id)
        assert result.verification_level != "sufficient"
        assert result.range_unproven_sheets >= 1
        assert any("源文件" in note or "Sheet" in note for note in result.notes)
    finally:
        conn.close()


def test_parser_source_dimension_mismatch_blocks_common_range_drift(tmp_path: Path, monkeypatch):
    """解析器若篡改 Sheet 行列范围元数据，入口也必须降级。"""
    from jiadun.core.engine import settlement_io
    from jiadun.core.models import project as pm
    from jiadun.core.parsing import excel_parser

    src = tmp_path / "范围漂移.xlsx"
    _make_book(src)
    real_parse = excel_parser.parse_xlsx

    def alter_range(path: Path):
        result = real_parse(path)
        # 保留所有 raw cell 均在声明网格内，单独模拟解析器把范围扩大的漂移。
        result.sheets[0].n_rows += 1
        return result

    monkeypatch.setattr(excel_parser, "parse_xlsx", alter_range)
    info = pm.create_project("解析范围漂移", tmp_path / "ws")
    info, conn = pm.open_project(Path(info.workspace_path))
    try:
        report = settlement_io.import_settlement_file(
            conn, info.project_id, Path(info.workspace_path), src, direction="upward"
        )
        stats = json.loads(
            conn.execute(
                "SELECT stats_json FROM parse_batches WHERE id=?", (report.batch_id,)
            ).fetchone()["stats_json"]
        )
        assert stats["source_census_status"] == "mismatch"
        assert report.status == "partial"
    finally:
        conn.close()


def test_parser_in_sheet_row_omission_blocks_common_row_range(tmp_path: Path, monkeypatch):
    """同一 Sheet 内漏掉一整行且保留尺寸时，源内容指纹也必须降级。"""
    from jiadun.core.engine import settlement_io
    from jiadun.core.models import project as pm
    from jiadun.core.parsing import excel_parser

    src = tmp_path / "Sheet内漏行.xlsx"
    _make_book(src)
    real_parse = excel_parser.parse_xlsx

    def omit_middle_row(path: Path):
        result = real_parse(path)
        sheet = result.sheets[0]
        # 保留 n_rows/n_cols，模拟解析器只漏掉物理第 3 行的单元格；
        # 仅比较 Sheet 尺寸无法发现这种共同漏行。
        sheet.cells = [cell for cell in sheet.cells if cell.row != 3]
        result.stats["n_cells"] = sum(len(item.cells) for item in result.sheets)
        return result

    monkeypatch.setattr(excel_parser, "parse_xlsx", omit_middle_row)
    info = pm.create_project("Sheet内漏行", tmp_path / "ws")
    info, conn = pm.open_project(Path(info.workspace_path))
    try:
        report = settlement_io.import_settlement_file(
            conn, info.project_id, Path(info.workspace_path), src, direction="upward"
        )
        batch = conn.execute(
            "SELECT status, stats_json FROM parse_batches WHERE id=?",
            (report.batch_id,),
        ).fetchone()
        assert batch is not None
        stats = json.loads(batch["stats_json"])
        assert stats["source_census_status"] == "mismatch"
        assert any("内容" in difference or "单元格" in difference
                   for difference in stats["source_census_differences"])
        assert report.status == "partial"
    finally:
        conn.close()


def test_source_census_rejects_duplicate_and_noncontiguous_sheet_indices():
    """Sheet 索引必须是源目录同序的 0-based 连续身份。"""
    from jiadun.core.parsing.excel_parser import (
        SheetRecord,
        _compare_source_census,
    )

    census = {
        "status": "available",
        "sheets": [
            {"sheet_index": 0, "sheet_name": "A", "range_status": "available", "n_rows": 3, "n_cols": 5},
            {"sheet_index": 1, "sheet_name": "B", "range_status": "available", "n_rows": 3, "n_cols": 5},
        ],
    }
    for bad_sheets in (
        [SheetRecord(0, "A", 3, 5), SheetRecord(0, "B", 3, 5)],
        [SheetRecord(1, "A", 3, 5), SheetRecord(2, "B", 3, 5)],
    ):
        status, differences = _compare_source_census(census, bad_sheets)
        assert status == "mismatch"
        assert any("索引" in difference for difference in differences)


def test_duplicate_sheet_index_is_rejected_before_persist(tmp_path: Path, monkeypatch):
    """坏索引不得先物化再在 coverage proof 末端留下错误来源现场。"""
    from jiadun.core.engine import settlement_io
    from jiadun.core.models import project as pm
    from jiadun.core.parsing import excel_parser

    src = tmp_path / "重复索引.xlsx"
    wb = openpyxl.Workbook()
    first = wb.active
    first.title = "第1期"
    second = wb.create_sheet("第2期")
    for ws, amount in ((first, 500), (second, 30)):
        for col, value in enumerate(("项目编码", "项目名称", "工程量", "综合单价", "合价"), 1):
            ws.cell(1, col, value=value)
        ws.append((f"{ws.title}-A", "项目", 1, amount, amount))
        ws.append((None, "合计", None, None, amount))
    wb.save(src)

    real_parse = excel_parser.parse_xlsx

    def duplicate_indices(path: Path):
        result = real_parse(path)
        result.sheets[1].sheet_index = 0
        return result

    monkeypatch.setattr(excel_parser, "parse_xlsx", duplicate_indices)
    info = pm.create_project("重复 Sheet 索引", tmp_path / "ws")
    info, conn = pm.open_project(Path(info.workspace_path))
    try:
        report = settlement_io.import_settlement_file(
            conn, info.project_id, Path(info.workspace_path), src, direction="upward"
        )
        assert report.status == "failed"
        assert "索引" in report.message
        batch = conn.execute(
            "SELECT status FROM parse_batches WHERE id=?", (report.batch_id,)
        ).fetchone()
        assert batch["status"] == "failed"
        assert conn.execute("SELECT COUNT(*) AS c FROM raw_sheets").fetchone()["c"] == 0
        assert conn.execute("SELECT COUNT(*) AS c FROM raw_cells").fetchone()["c"] == 0
        assert conn.execute("SELECT COUNT(*) AS c FROM settlement_periods").fetchone()["c"] == 0
        assert conn.execute("SELECT COUNT(*) AS c FROM line_items").fetchone()["c"] == 0
    finally:
        conn.close()


def test_import_rolls_back_materialization_when_coverage_proof_fails(
    tmp_path: Path, monkeypatch
):
    """覆盖证明末端失败时，不能留下半套 raw/period/line_items 现场。"""
    from jiadun.core.engine import settlement_io
    from jiadun.core.models import project as pm

    src = tmp_path / "覆盖证明失败.xlsx"
    _make_book(src)
    info = pm.create_project("覆盖证明原子性", tmp_path / "ws")
    info, conn = pm.open_project(Path(info.workspace_path))

    def fail_after_materialization(*_args, **_kwargs):
        raise RuntimeError("injected coverage proof failure")

    monkeypatch.setattr(
        settlement_io, "_persist_coverage_proof", fail_after_materialization
    )
    try:
        report = settlement_io.import_settlement_file(
            conn, info.project_id, Path(info.workspace_path), src, direction="upward"
        )
        assert report.status == "failed"
        assert "injected coverage proof failure" in report.message
        assert conn.execute("SELECT COUNT(*) AS c FROM source_files").fetchone()["c"] == 1
        assert conn.execute("SELECT COUNT(*) AS c FROM raw_sheets").fetchone()["c"] == 0
        assert conn.execute("SELECT COUNT(*) AS c FROM raw_cells").fetchone()["c"] == 0
        assert conn.execute("SELECT COUNT(*) AS c FROM settlement_periods").fetchone()["c"] == 0
        assert conn.execute("SELECT COUNT(*) AS c FROM line_items").fetchone()["c"] == 0
        batches = conn.execute(
            "SELECT status FROM parse_batches ORDER BY id"
        ).fetchall()
        assert [row["status"] for row in batches] == ["failed"]
    finally:
        conn.close()


def test_c_control_wraps_source_evidence_for_current_run_without_mutating_source(
    tmp_path: Path,
):
    """C 控制值保留原始来源，同时用当前运行包装后再进入摘要。"""
    from jiadun.core.contracts import run_contract
    from jiadun.core.engine import aggregate, crosscheck
    from jiadun.core.reporting import build_project_summary

    info, conn, _report, period_id = _import_project(tmp_path)
    try:
        aggregate.persist_period_totals(
            conn,
            info.project_id,
            aggregate.aggregate_project(conn, info.project_id, direction="upward"),
        )
        grand = conn.execute(
            """SELECT id, flags_json FROM line_items
               WHERE period_id=? AND json_extract(flags_json, '$.grand_total')=1""",
            (period_id,),
        ).fetchone()
        assert grand is not None
        source_id = int(json.loads(grand["flags_json"])["source_evidence_id"])
        source_before = dict(
            conn.execute(
                """SELECT scope, run_id, run_signature, sources_json, steps_json
                   FROM evidence WHERE id=?""",
                (source_id,),
            ).fetchone()
        )
        assert source_before["scope"] == "source"
        assert source_before["run_id"] is None
        assert source_before["run_signature"] is None

        result = crosscheck.run_crosscheck(
            conn, info.project_id, [1], direction="upward"
        )[0]
        current = run_contract.get_current_contract(conn, info.project_id)
        assert current is not None
        assert result.c_control_evidence_id is not None
        assert result.c_control_evidence_id != source_id

        wrapper = conn.execute(
            """SELECT kind, scope, run_id, run_signature, sources_json, steps_json
               FROM evidence WHERE id=?""",
            (result.c_control_evidence_id,),
        ).fetchone()
        assert wrapper is not None
        assert (wrapper["kind"], wrapper["scope"]) == ("line_item_source", "current")
        assert (wrapper["run_id"], wrapper["run_signature"]) == (
            current.run_id,
            current.signature,
        )
        wrapper_sources = json.loads(wrapper["sources_json"])
        control_source = next(
            item for item in wrapper_sources if item.get("line_item_id") == grand["id"]
        )
        assert control_source["source_evidence_id"] == source_id
        assert control_source["file_id"] is not None
        assert control_source["sheet_id"] is not None
        assert control_source["row"] == 5
        assert control_source["col"] == 5
        assert control_source["raw_value"] == "500"
        assert control_source["raw_cell_value"] == "500"
        assert control_source["raw_cell_is_formula"] is False
        wrapper_steps = json.loads(wrapper["steps_json"])
        assert wrapper_steps[0]["step"] == "C 控制来源包装"
        assert source_id in wrapper_steps[0]["source_evidence_ids"]

        source_after = dict(
            conn.execute(
                """SELECT scope, run_id, run_signature, sources_json, steps_json
                   FROM evidence WHERE id=?""",
                (source_id,),
            ).fetchone()
        )
        assert source_after == source_before

        summary = build_project_summary(conn, info.project_id)
        assert summary.verification["evidence_complete"] is True
        assert summary.verification["status"] == "sufficient"
    finally:
        conn.close()


def test_c_control_blocks_when_line_item_locator_cannot_reach_raw_cells(
    tmp_path: Path,
):
    """C 不能用清洗金额替代缺失的原始控制单元格。"""
    from jiadun.core.engine import crosscheck

    _info, conn, _report, period_id = _import_project(tmp_path)
    try:
        grand = conn.execute(
            """SELECT id FROM line_items
               WHERE period_id=? AND json_extract(flags_json, '$.grand_total')=1""",
            (period_id,),
        ).fetchone()
        assert grand is not None
        conn.execute(
            "UPDATE line_items SET amount_evid=? WHERE id=?",
            (json.dumps({"raw": "500", "value": "500", "row": 5, "col": 99}), grand["id"]),
        )

        result = crosscheck.check_period(conn, period_id)

        assert result.raw_subtotal is None
        assert result.control_status == "not_available"
        assert result.c_independence_level == "not_proven"
        assert result.verification_level == "insufficient"
        assert any("无法独立定位" in note for note in result.notes)
    finally:
        conn.close()


def test_c_control_marks_multiple_grand_candidates_not_proven(tmp_path: Path):
    """多个合计候选不能通过清洗投影或求和强行形成 C 控制值。"""
    from jiadun.core.engine import crosscheck

    _info, conn, _report, period_id = _import_project(
        tmp_path, two_grand_totals=True
    )
    try:
        result = crosscheck.check_period(conn, period_id)

        assert result.control_status == "not_available"
        assert result.c_independence_level == "not_proven"
        assert result.verification_level == "insufficient"
        assert any("控制值不唯一" in note for note in result.notes)
    finally:
        conn.close()


def test_duplicate_detail_blocks_green_even_when_totals_are_balanced(tmp_path: Path):
    from jiadun.core.engine import crosscheck, settlement_io
    from jiadun.core.models import project as pm

    src = tmp_path / "duplicate.xlsx"
    _duplicate_book(src)
    info = pm.create_project("重复明细闸门", tmp_path / "ws")
    info, conn = pm.open_project(info.workspace_path)
    try:
        settlement_io.import_settlement_file(
            conn, info.project_id, Path(info.workspace_path), src, direction="upward"
        )
        period_id = conn.execute(
            "SELECT id FROM settlement_periods WHERE project_id=?", (info.project_id,)
        ).fetchone()["id"]
        result = crosscheck.check_period(conn, period_id)
        assert result.status == "match"
        assert result.control_status == "match"
        assert result.verification_level != "sufficient"
        assert any("重复" in note for note in result.notes)
    finally:
        conn.close()


def test_negative_adjustment_requires_human_review_even_when_totals_are_balanced(tmp_path: Path):
    from jiadun.core.engine import crosscheck, settlement_io
    from jiadun.core.models import project as pm

    src = tmp_path / "negative-adjustment.xlsx"
    _negative_adjustment_book(src)
    info = pm.create_project("负数调整闸门", tmp_path / "ws")
    info, conn = pm.open_project(info.workspace_path)
    try:
        settlement_io.import_settlement_file(
            conn, info.project_id, Path(info.workspace_path), src, direction="upward"
        )
        period_id = conn.execute(
            "SELECT id FROM settlement_periods WHERE project_id=?", (info.project_id,)
        ).fetchone()["id"]
        result = crosscheck.check_period(conn, period_id)
        assert result.status == "match"
        assert result.control_status == "match"
        assert result.verification_level != "sufficient"
        assert any("负数" in note or "调整" in note for note in result.notes)
    finally:
        conn.close()


def test_reimporting_same_source_does_not_duplicate_canonical_rows(tmp_path: Path):
    """相同文件重复选择不能把同一批明细再次追加到当前期次。"""
    from jiadun.core.engine import aggregate, crosscheck, settlement_io

    info, conn, _report, period_id = _import_project(tmp_path)
    source = tmp_path / "第1期结算.xlsx"
    workspace = Path(info.workspace_path)
    try:
        before = {
            "items": conn.execute(
                "SELECT COUNT(*) AS c FROM line_items WHERE period_id=?", (period_id,)
            ).fetchone()["c"],
            "sheets": conn.execute(
                "SELECT COUNT(*) AS c FROM raw_sheets WHERE period_id=?", (period_id,)
            ).fetchone()["c"],
        }
        settlement_io.import_settlement_file(
            conn, info.project_id, workspace, source, direction="upward"
        )
        after = {
            "items": conn.execute(
                "SELECT COUNT(*) AS c FROM line_items WHERE period_id=?", (period_id,)
            ).fetchone()["c"],
            "sheets": conn.execute(
                "SELECT COUNT(*) AS c FROM raw_sheets WHERE period_id=?", (period_id,)
            ).fetchone()["c"],
        }
        assert after == before
        result = crosscheck.check_period(conn, period_id)
        assert result.path_a_total == Decimal("500")
        assert result.control_status == "match"
        aggs = aggregate.aggregate_project(conn, info.project_id, direction="upward")
        assert sum((item.cum_amount or Decimal("0") for item in aggs), Decimal("0")) == Decimal("500")
    finally:
        conn.close()


def test_missing_raw_amount_is_not_persisted_as_zero_in_coverage_proof(tmp_path: Path):
    from jiadun.core.engine import settlement_io
    from jiadun.core.models import project as pm

    src = tmp_path / "missing-amount.xlsx"
    _all_missing_amount_book(src)
    info = pm.create_project("缺失金额闸门", tmp_path / "ws")
    info, conn = pm.open_project(info.workspace_path)
    try:
        settlement_io.import_settlement_file(
            conn, info.project_id, Path(info.workspace_path), src, direction="upward"
        )
        proof = conn.execute(
            "SELECT raw_amount_total FROM sheet_coverage_proofs WHERE project_id=?",
            (info.project_id,),
        ).fetchone()
        assert proof is not None
        assert proof["raw_amount_total"] is None
    finally:
        conn.close()


def test_internal_blank_or_title_rows_block_green_even_when_totals_match(tmp_path):
    """内部空白/标题行的排除必须进入复核状态，不能仅凭 A/B/C 数字放行。"""
    from jiadun.core.engine import crosscheck, settlement_io
    from jiadun.core.models import project as pm

    src = tmp_path / "内部空白标题.xlsx"
    _internal_blank_title_book(src)
    info = pm.create_project("内部空白标题闸门", tmp_path / "ws")
    info, conn = pm.open_project(info.workspace_path)
    try:
        report = settlement_io.import_settlement_file(
            conn, info.project_id, Path(info.workspace_path), src, direction="upward"
        )
        period_id = conn.execute(
            "SELECT id FROM settlement_periods WHERE project_id=?", (info.project_id,)
        ).fetchone()["id"]
        result = crosscheck.check_period(conn, period_id)

        assert report.status == "partial"
        assert result.status == "match"
        assert result.control_status == "match"
        assert result.verification_level == "insufficient"
        assert result.excluded_blank_rows >= 1
        assert result.excluded_title_rows >= 1
        assert any("空白" in note or "标题" in note for note in result.notes)
    finally:
        conn.close()


def test_anonymized_golden_registry_is_present_and_explicit(tmp_path: Path):
    # 真实案例尚未由用户提供时也必须有受控入口，不能用不存在的目录冒充已覆盖。
    from scripts.golden_regression import load_registry

    registry = Path(__file__).resolve().parents[1] / "anonymized_golden_cases" / "cases.json"
    data = load_registry(registry)
    assert data["schema_version"] == 1
    assert isinstance(data["cases"], list) and data["cases"]
    assert all(case.get("case_kind") == "sanitized_real" for case in data["cases"])
    assert all(case.get("availability") in {"available", "not_available"} for case in data["cases"])

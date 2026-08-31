"""监督门槛回归测试（Phase 7 提交前五项补充）。

1) Decimal("0") 必须保留，缺失仅按 is None；
2) 对上/对下累计按 direction 独立聚合，period_id 取值防同期号串表；
3) 导出文件第二路径 Decimal 勾稽（独立于 LibreOffice 重算）；
   WPS 无 headless CLI → 明确标注未自动验证；
4) 序列化边界 _num：Decimal 直写保真，float 仅在常量分支且先 round2；
5) 摘要声明范围/期次/方向/税口径/证据数；Word 不声称无入口的"全可追溯"。
"""
import sys
from decimal import Decimal
from pathlib import Path

import pytest

sys.path.insert(0, str(Path(__file__).parents[2] / "synthetic_test_data"))

from generator import make_messy, make_multi_period  # noqa: E402

from jiadun.core.anomalies import engine  # noqa: E402
from jiadun.core.engine import settlement_io  # noqa: E402
from jiadun.core.export import excel_export  # noqa: E402


@pytest.fixture()
def project(tmp_path):
    from jiadun.core.models import project as pm

    info = pm.create_project("监督门槛", tmp_path / "ws")
    info, conn = pm.open_project(Path(info.workspace_path))
    yield info, conn, Path(info.workspace_path)
    conn.close()


class TestZeroPreserved:
    """门槛 1：数量/单价/合价为 0 是有效值，不得被当作缺失。"""

    def test_zero_not_treated_as_missing(self, project):
        info, conn, pdir = project
        src = pdir.parent / "第1期.xlsx"
        make_messy(src, seed=7)
        settlement_io.import_settlement_file(conn, info.project_id, pdir, src)
        # 直接落库一行：合价 0、数量 0（有效值），一行：合价缺失
        with conn:
            p1 = conn.execute(
                "SELECT id FROM settlement_periods WHERE project_id=? AND period_no=1",
                (info.project_id,),
            ).fetchone()
            conn.execute(
                "INSERT INTO line_items(period_id, code, name, unit, quantity, unit_price, amount, flags_json)"
                " VALUES (?, 'Z1', '零值清单', 'm3', '0', '100', '0', '{}')",
                (p1["id"],),
            )
            conn.execute(
                "INSERT INTO line_items(period_id, code, name, unit, quantity, flags_json)"
                " VALUES (?, 'Z2', '缺失金额清单', 'm3', '5', '{}')",
                (p1["id"],),
            )
        path = excel_export.export_workbook(conn, info.project_id, pdir / "exports")
        import openpyxl

        wb = openpyxl.load_workbook(path, data_only=False)
        ws = wb["审核底稿"]
        found_zero = found_missing = False
        for r in range(2, ws.max_row + 1):
            name = ws.cell(row=r, column=3).value
            if name == "零值清单":
                # 0 必须以数值 0 写入（缺失才会写 "待补资料"/None）
                assert ws.cell(row=r, column=5).value == 0
                assert ws.cell(row=r, column=9).value == 0   # 原表合价 = 0
                assert ws.cell(row=r, column=8).value == 0   # 程序计算合价 = 0
                assert ws.cell(row=r, column=7).value == f"=E{r}*F{r}"  # 公式保留
                found_zero = True
            if name == "缺失金额清单":
                assert ws.cell(row=r, column=9).value is None  # 缺失 = None（原表合价）
                assert ws.cell(row=r, column=7).value == "待补资料"
                assert ws.cell(row=r, column=8).value == "待补资料"  # 程序值同标
                found_missing = True
        assert found_zero and found_missing


class TestDirectionSeparation:
    """门槛 2：对上/对下同期号不得串表，累计不得混入另一方向。"""

    @pytest.fixture()
    def mixed_direction_project(self, project):
        info, conn, pdir = project
        # 对上第1期 100 元、对下第1期 200 元（同期号不同方向）

        with conn:
            cur = conn.execute(
                "INSERT INTO settlement_periods(project_id, period_no, title, direction)"
                " VALUES (?, 1, '对上第1期', 'upward')", (info.project_id,))
            up_pid = cur.lastrowid
            conn.execute(
                "INSERT INTO settlement_periods(project_id, period_no, title, direction)"
                " VALUES (?, 1, '对下第1期', 'downward')", (info.project_id,))
            down_pid = conn.execute(
                "SELECT id FROM settlement_periods WHERE project_id=? AND period_no=1 AND direction='downward'",
                (info.project_id,),
            ).fetchone()["id"]
            conn.execute(
                "INSERT INTO line_items(period_id, code, name, unit, quantity, amount, flags_json)"
                " VALUES (?, 'K1', '同一清单', 'm3', '10', '100', '{}')", (up_pid,))
            conn.execute(
                "INSERT INTO line_items(period_id, code, name, unit, quantity, amount, flags_json)"
                " VALUES (?, 'K1', '同一清单', 'm3', '20', '200', '{}')", (down_pid,))
        return info, conn, pdir, up_pid, down_pid

    def test_aggregate_by_direction_not_crossed(self, mixed_direction_project):
        info, conn, pdir, up_pid, down_pid = mixed_direction_project
        ups = {a.item_key: a for a in excel_export.aggregate_project(conn, info.project_id, direction="upward")}
        downs = {a.item_key: a for a in excel_export.aggregate_project(conn, info.project_id, direction="downward")}
        assert ups["code:K1"].cum_amount == Decimal("100")
        assert downs["code:K1"].cum_amount == Decimal("200")
        # per_period 键为 period_id，两方向各自一期，互不串
        assert list(ups["code:K1"].per_period) == [up_pid]
        assert list(downs["code:K1"].per_period) == [down_pid]
        # 只有调用方明确声明项目跨方向展示时，才允许全项目 = 300
        with pytest.raises(ValueError, match="direction"):
            excel_export.aggregate_project(conn, info.project_id)
        alls = {
            a.item_key: a for a in excel_export.aggregate_project(
                conn, info.project_id, include_all_directions=True
            )
        }
        assert alls["code:K1"].cum_amount == Decimal("300")

    def test_summary_sheets_separated(self, mixed_direction_project):
        info, conn, pdir, *_ = mixed_direction_project
        wb = excel_export.Workbook()
        wb.remove(wb.active)
        excel_export.export_settlement_summary(conn, info.project_id, wb, direction="upward")
        excel_export.export_settlement_summary(conn, info.project_id, wb, direction="downward")
        import io

        buf = io.BytesIO()
        wb.save(buf)
        buf.seek(0)
        import openpyxl

        wb2 = openpyxl.load_workbook(buf)
        up = wb2["对上结算累计表"]
        down = wb2["对下结算累计表"]
        # 对上表第1期金额列 = 100（不含对下的 200）
        assert up.cell(row=2, column=4).value == Decimal("100")
        assert down.cell(row=2, column=4).value == Decimal("200")

    def test_full_export_has_no_ambiguous_cross_direction_summary(self, mixed_direction_project):
        info, conn, pdir, up_pid, down_pid = mixed_direction_project

        with conn:
            up_item = conn.execute(
                "SELECT id FROM line_items WHERE period_id=?", (up_pid,)
            ).fetchone()["id"]
            down_item = conn.execute(
                "SELECT id FROM line_items WHERE period_id=?", (down_pid,)
            ).fetchone()["id"]
            conn.executemany(
                """INSERT INTO anomalies(project_id, rule_id, severity, subject_type,
                   subject_id, message, status, created_at)
                   VALUES (?, 'direction_probe', 'medium', 'line_item', ?, ?, 'open', '2026')""",
                [
                    (info.project_id, up_item, "对上方向反例"),
                    (info.project_id, down_item, "对下方向反例"),
                ],
            )

        path = excel_export.export_workbook(conn, info.project_id, pdir / "exports")
        import openpyxl

        wb = openpyxl.load_workbook(path, data_only=False)
        assert "结算累计表" not in wb.sheetnames
        assert "对上结算累计表" in wb.sheetnames
        assert "对下结算累计表" in wb.sheetnames
        audit = wb["审核底稿"]
        assert audit.cell(row=1, column=15).value == "方向"
        assert {audit.cell(row=r, column=15).value for r in range(2, audit.max_row + 1)} == {
            "对上结算", "对下结算"
        }
        summary_text = " | ".join(
            str(wb["管理层摘要"].cell(row=r, column=1).value)
            for r in range(1, wb["管理层摘要"].max_row + 1)
        )
        assert "对上结算累计金额（可用部分）" in summary_text
        assert "对下结算累计金额（可用部分）" in summary_text
        assert "累计金额合计（可用部分）" not in summary_text
        anomaly = wb["异常清单"]
        assert [anomaly.cell(row=1, column=c).value for c in range(1, 9)] == [
            "编号", "方向", "规则", "级别", "对象", "说明", "证据ID", "状态",
        ]
        anomaly_directions = {
            anomaly.cell(row=r, column=2).value
            for r in range(2, anomaly.max_row + 1)
            if anomaly.cell(row=r, column=9).value == "direction_probe"
        }
        assert anomaly_directions == {"对上结算", "对下结算"}
        assert all(
            anomaly.cell(row=r, column=3).value == "其他审核问题"
            for r in range(2, anomaly.max_row + 1)
            if anomaly.cell(row=r, column=9).value == "direction_probe"
        )
        pending = wb["待核实事项清单"]
        pending_directions = {
            pending.cell(row=r, column=2).value
            for r in range(2, pending.max_row + 1)
            if pending.cell(row=r, column=6).value == "direction_probe"
        }
        assert pending_directions == {"对上结算", "对下结算"}


class TestSecondPathRecompute:
    """门槛 3：第二路径 —— 从导出文件读回，Decimal 复算明细→汇总关系。"""

    def test_exported_workbook_internal_reconcile(self, project):
        info, conn, pdir = project
        src = pdir.parent / "multi.xlsx"
        make_multi_period(src, periods=3)
        settlement_io.import_settlement_file(conn, info.project_id, pdir, src)
        path = excel_export.export_workbook(conn, info.project_id, pdir / "exports")

        import openpyxl

        wb = openpyxl.load_workbook(path, data_only=False)
        ws_detail = wb["审核底稿"]
        ws_summary = wb["未标记累计表"]

        # 第二路径：明细按归组键 Decimal 累计（读回的是写入值，非 DB）。
        # 归组口径必须与累计表一致：code 优先，无码用名称（同码异名属同组）。
        detail_sum: dict[tuple, Decimal] = {}
        for r in range(2, ws_detail.max_row + 1):
            amount = ws_detail.cell(row=r, column=8).value
            if amount is None or isinstance(amount, str):
                continue
            code = ws_detail.cell(row=r, column=2).value
            name = ws_detail.cell(row=r, column=3).value
            key = ("code", code) if code else ("name", name)
            detail_sum[key] = detail_sum.get(key, Decimal("0")) + Decimal(repr(amount))
        # 与累计表"累计金额"列逐一对比（累计表为 round2 展示口径，容差 0.01）
        checked = 0
        for r in range(2, ws_summary.max_row + 1):
            code = ws_summary.cell(row=r, column=1).value
            name = ws_summary.cell(row=r, column=2).value
            key = ("code", code) if code else ("name", name)
            cum = ws_summary.cell(row=r, column=8).value  # 累计金额
            if key not in detail_sum or cum is None:
                continue
            assert abs(detail_sum[key] - Decimal(repr(cum))) <= Decimal("0.01"), \
                f"summary mismatch for {key}: detail={detail_sum[key]} summary={cum}"
            checked += 1
        assert checked > 0

class TestSummaryScope:
    """门槛 5：摘要声明范围/期次/方向/税口径/证据数；Word 不声称无入口的追溯。"""

    def test_summary_sheet_declares_scope(self, project):
        info, conn, pdir = project
        src = pdir.parent / "multi.xlsx"
        make_multi_period(src, periods=2)
        settlement_io.import_settlement_file(conn, info.project_id, pdir, src)
        engine.run_anomalies(conn, info.project_id)
        wb = excel_export.Workbook()
        wb.remove(wb.active)
        excel_export.export_management_summary(conn, info.project_id, wb)
        import io

        buf = io.BytesIO()
        wb.save(buf)
        buf.seek(0)
        import openpyxl

        ws = openpyxl.load_workbook(buf)["管理层摘要"]
        text = " | ".join(str(ws.cell(row=r, column=1).value) for r in range(1, ws.max_row + 1))
        for required in ("其中：对上结算", "其中：对下结算", "其中：未标记", "单价税口径", "证据记录数", "已导入原始文件数"):
            assert required in text, f"summary missing scope item: {required}"

    def test_docx_does_not_claim_traceability_without_entry(self, project):
        info, conn, pdir = project
        src = pdir.parent / "multi.xlsx"
        make_multi_period(src, periods=2)
        settlement_io.import_settlement_file(conn, info.project_id, pdir, src)
        path = excel_export.export_management_summary_docx(conn, info.project_id, pdir / "exports")
        import docx as docx_lib

        doc = docx_lib.Document(str(path))
        full_text = "\n".join(p.text for p in doc.paragraphs)
        # 不得声称"所有结论可追溯"（Word 无证据入口）
        assert "所有结论可在证据索引中追溯" not in full_text
        assert "不包含证据入口" in full_text
        # 必须指明证据位置与合成数据声明
        assert "证据索引" in full_text and "审核底稿" in full_text
        assert "不构成任何真实业务结论" in full_text

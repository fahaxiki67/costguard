"""合成结算文件生成器（synthetic_test_data）。

每个生成函数返回 (xlsx_path, ground_truth)：ground_truth 记录"应该解析出什么"，
测试断言解析结果与真值一致。文件只写入指定目录（合成数据可入库，无敏感性）。

覆盖场景（AGENTS.md 六 / 任务书九）：
表头位置不同 | 多层表头 | 合并单元格 | 隐藏行 | 隐藏列 | 空白行 | 文本数字
千分位 | ¥符号 | 负数(含括号) | 公式错误 | 公式无缓存 | 多Sheet | Sheet名随机
重复数据 | 漏项 | 单位转换 | 名称轻微差异 | 名称语义一致但字面不同
名称相似但实际不同 | 税率变化 | 舍入差异
"""
from __future__ import annotations

import random
from dataclasses import dataclass, field
from pathlib import Path

import openpyxl
from openpyxl.styles import Font


@dataclass
class Item:
    code: str
    name: str
    feature: str = ""
    unit: str = "m3"
    quantity: str = ""
    unit_price: str = ""
    amount: str = ""
    tax_rate: str = "0.09"


@dataclass
class GroundTruth:
    file_name: str
    sheet_name: str
    header_rows: tuple[int, int]  # 表头行区间（1-based）
    items: list[Item] = field(default_factory=list)
    hidden_rows: list[int] = field(default_factory=list)
    hidden_cols: list[int] = field(default_factory=list)
    notes: list[str] = field(default_factory=list)


STANDARD_ITEMS = [
    Item("010101001001", "平整场地", "三类土", "m2", "1250.50", "8.50", "10629.25", "0.09"),
    Item("010101003001", "挖沟槽土方", "深度2m以内 三类土", "m3", "860.00", "35.20", "30272.00", "0.09"),
    Item("010401001001", "砖基础", "MU10机制砖 M7.5水泥砂浆", "m3", "152.30", "420.00", "63966.00", "0.09"),
    Item("010501001001", "C25混凝土垫层", "商品混凝土 C25", "m3", "48.20", "465.00", "22413.00", "0.09"),
    Item("010515001001", "现浇构件钢筋", "HRB400 直径10mm以内", "t", "25.80", "4850.00", "125130.00", "0.13"),
    Item("011001001001", "水泥砂浆楼地面", "1:3水泥砂浆 20mm厚", "m2", "680.00", "45.80", "31144.00", "0.09"),
]


def _write_sheet(wb, sheet_name: str, items: list[Item], gt: GroundTruth,
                 header_start: int = 1, title_rows: int = 0,
                 two_row_header: bool = False, merge_header: bool = False,
                 merge_data_unit: bool = False, blank_row_after: int | None = None,
                 text_numbers: set[int] | None = None, thousands: set[int] | None = None,
                 yuan_sign: set[int] | None = None, negatives: set[int] | None = None,
                 formula_errors: set[int] | None = None, formulas: set[int] | None = None,
                 duplicates: list[int] | None = None, unit_alias: set[int] | None = None,
                 name_tweaks: dict[int, str] | None = None, round_diff: set[int] | None = None,
                 tax_change: dict[int, str] | None = None, hidden_row_in_data: int | None = None,
                 hidden_col_note: bool = False, currency_fmt: bool = False) -> None:
    ws = wb.create_sheet(sheet_name)
    r = 1
    for i in range(title_rows):
        ws.cell(row=r, column=1, value="××工程结算书（合成测试数据）" if i == 0 else f"编制单位：合成数据生成器  第1期 说明行{i}").font = Font(bold=True, size=14)
        r += 1
    header_row = header_start

    if two_row_header:
        # 第一行：合并"合价"组表头；第二行：具体字段
        ws.cell(row=r, column=1, value="清单编码")
        ws.cell(row=r, column=2, value="清单名称")
        ws.cell(row=r, column=3, value="项目特征")
        ws.cell(row=r, column=4, value="单位")
        ws.cell(row=r, column=5, value="工程量")
        ws.cell(row=r, column=6, value="综合单价")
        ws.cell(row=r, column=7, value="金额")
        ws.merge_cells(start_row=r, start_column=7, end_row=r, end_column=8)
        ws.cell(row=r + 1, column=7, value="合价")
        ws.cell(row=r + 1, column=8, value="税率")
        gt.notes.append("two-row header with merged group cell")
        header_row = r + 1
        r += 2
    else:
        cols = ["清单编码", "清单名称", "项目特征", "单位", "工程量", "综合单价", "合价", "税率"]
        for c, v in enumerate(cols, start=1):
            ws.cell(row=header_row, column=c, value=v)
        if merge_header:
            # 表头区上方的组表头合并（不覆盖有效字段表头）
            ws.cell(row=header_row - 1, column=1, value="清单信息")
            ws.merge_cells(start_row=header_row - 1, start_column=1, end_row=header_row - 1, end_column=4)
        r = header_row + 1

    text_numbers = text_numbers or set()
    thousands = thousands or set()
    yuan_sign = yuan_sign or set()
    negatives = negatives or set()
    formula_errors = formula_errors or set()
    formulas = formulas or set()
    unit_alias = unit_alias or set()
    name_tweaks = name_tweaks or {}
    round_diff = round_diff or set()
    tax_change = tax_change or {}

    all_items = list(items)
    for d in (duplicates or []):
        all_items.insert(d, items[min(d, len(items) - 1)])

    for i, it in enumerate(all_items, start=1):
        row = r
        name = name_tweaks.get(i, it.name)
        ws.cell(row=row, column=1, value=it.code)
        ws.cell(row=row, column=2, value=name)
        ws.cell(row=row, column=3, value=it.feature)
        unit = "立方米" if i in unit_alias else it.unit
        ws.cell(row=row, column=4, value=unit)
        qty, price, amount = it.quantity, it.unit_price, it.amount
        if i in negatives:
            qty, amount = f"-{qty}", f"-{amount}"
        if i in negatives and i % 2 == 0:
            qty = f"({qty.lstrip('-')})"
        if i in text_numbers:
            ws.cell(row=row, column=5, value=qty)  # 文本直接写
            ws.cell(row=row, column=7, value=amount)
        elif i in yuan_sign:
            ws.cell(row=row, column=5, value=f"¥{qty}")
            ws.cell(row=row, column=6, value=f"¥{price}")
            ws.cell(row=row, column=7, value=f"¥{amount}")
        elif i in thousands:
            def thou(s):
                if "." in s:
                    ip, fp = s.split(".")
                    return f"{int(ip):,}.{fp}"
                return f"{int(s):,}"
            ws.cell(row=row, column=5, value=thou(qty))
            ws.cell(row=row, column=7, value=thou(amount))
        elif i in formulas:
            ws.cell(row=row, column=5, value=float(qty))
            ws.cell(row=row, column=6, value=float(price))
            ws.cell(row=row, column=7, value=f"=E{row}*F{row}")
        elif i in formula_errors:
            ws.cell(row=row, column=5, value=float(qty))
            ws.cell(row=row, column=6, value=0)
            ws.cell(row=row, column=7, value=f"=E{row}/F{row}")  # #DIV/0!
        else:
            ws.cell(row=row, column=5, value=float(qty))
            ws.cell(row=row, column=6, value=float(price))
            ws.cell(row=row, column=7, value=float(amount))
        tax = tax_change.get(i, it.tax_rate)
        if i in round_diff:
            ws.cell(row=row, column=7, value=float(amount) + 0.01)
            gt.notes.append(f"row {i}: amount off by 0.01")
        if i in text_numbers:
            ws.cell(row=row, column=8, value=str(tax))  # 税率以文本存储
        else:
            ws.cell(row=row, column=8, value=float(tax))
        if currency_fmt and i % 2 == 0:
            ws.cell(row=row, column=7).number_format = "¥#,##0.00"
        if merge_data_unit and i > 1 and i % 3 == 0:
            ws.merge_cells(start_row=row - 1, start_column=4, end_row=row, end_column=4)
        # ground truth 始终记录规范化后的真值
        gt_qty, gt_amount = qty, amount
        if i in negatives:
            gt_qty, gt_amount = f"-{it.quantity}", f"-{it.amount}"
        gt.items.append(
            Item(it.code, name, it.feature, unit, gt_qty, price,
                 gt_amount if i not in round_diff else f"{float(amount) + 0.01:.2f}", tax)
        )
        if blank_row_after == i:
            r += 1  # 空行
        r += 1

    # 小计行
    ws.cell(row=r, column=2, value="小计")
    try:
        total = sum(float(x.amount) for x in all_items)
    except ValueError:
        total = 0.0
    ws.cell(row=r, column=7, value=round(total, 2))
    gt.notes.append("subtotal row present (must be flagged, not counted)")

    if hidden_row_in_data:
        ws.row_dimensions[header_row + hidden_row_in_data].hidden = True
        gt.hidden_rows = [header_row + hidden_row_in_data]
    if hidden_col_note:
        ws.column_dimensions["H"].hidden = True
        gt.hidden_cols = [8]

    gt.header_rows = (header_row, header_row)
    gt.sheet_name = ws.title


def make_clean(path: Path) -> GroundTruth:
    """基准：规范单期结算表。"""
    wb = openpyxl.Workbook()
    wb.remove(wb.active)
    gt = GroundTruth("clean_1期.xlsx", "第1期", (1, 1))
    _write_sheet(wb, "第1期", STANDARD_ITEMS, gt)
    wb.save(path)
    return gt


def make_messy(path: Path, seed: int = 42) -> list[GroundTruth]:
    """高难度：叠加大量畸形的文件（3 个 sheet，随机 sheet 名）。"""
    rng = random.Random(seed)
    wb = openpyxl.Workbook()
    wb.remove(wb.active)
    gts: list[GroundTruth] = []

    gt1 = GroundTruth(path.name, rng.choice(["Sheet1", "结算表(1)", "data"]), (4, 4))
    _write_sheet(
        wb, gt1.sheet_name, STANDARD_ITEMS, gt1,
        header_start=4, title_rows=3, merge_header=True, merge_data_unit=True,
        blank_row_after=3, text_numbers={2}, thousands={3}, yuan_sign={4},
        negatives={5}, hidden_row_in_data=7, hidden_col_note=True, currency_fmt=True,
    )
    gts.append(gt1)

    gt2 = GroundTruth(gt1.file_name, rng.choice(["补充-清单", "zz"]), (1, 2))
    wb_title = gt2.sheet_name
    _write_sheet(
        wb, wb_title, STANDARD_ITEMS[:4], gt2,
        two_row_header=True, formulas={2}, formula_errors={3},
        unit_alias={2}, name_tweaks={1: "平整场地（人工）"},
        round_diff={4}, tax_change={4: "0.13"},
    )
    gts.append(gt2)

    gt3 = GroundTruth(gt1.file_name, "封面", (0, 0))
    ws = wb.create_sheet("封面")
    ws.cell(row=1, column=1, value="××项目结算汇总")
    ws.cell(row=2, column=1, value="2026年8月")
    gts.append(gt3)  # 封面：应被识别为非清单表

    wb.save(path)
    return gts


def make_multi_period(path: Path, periods: int = 3) -> list[GroundTruth]:
    """多期累计场景：第2期漏项、第3期单价变化、名称轻微差异。"""
    wb = openpyxl.Workbook()
    wb.remove(wb.active)
    gts = []
    for p in range(1, periods + 1):
        gt = GroundTruth(Path(path).name, f"第{p}期", (1, 1))
        items = [Item(x.code, x.name, x.feature, x.unit,
                      str(float(x.quantity) / periods), x.unit_price,
                      str(float(x.quantity) / periods * float(x.unit_price)),
                      x.tax_rate) for x in STANDARD_ITEMS]
        if p == 2:
            items = items[:-1]  # 漏项：最后一项缺失
        tweaks = {}
        if p == 3:
            tweaks = {2: " 挖沟槽土方 ", 4: "C25砼垫层"}  # 轻微差异 / 语义相同字面不同
            for idx, it in enumerate(items):
                if idx == 3:
                    it.unit_price = "510.00"  # 单价异常变化
                    it.amount = str(float(it.quantity) * 510.0)
        _write_sheet(wb, f"第{p}期", items, gt, title_rows=1, name_tweaks=tweaks)
        gts.append(gt)
    wb.save(path)
    return gts

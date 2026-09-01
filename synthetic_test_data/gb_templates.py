"""国标表式合成生成器（GB50500 常见表式，全部合成数据）。

按真实结算书普遍版式构造（不复制任何真实项目内容）：
- make_table_08：表-08 分部分项工程和单价措施项目清单与计价表
  ——三层表头：第3行 序号/编码/名称/特征/计量单位/工程量/金额（元，横跨G:H）；
    第4行 综合单价/合价/其中（横跨I:K）；第5行 定额人工费/定额机械费/暂估价；
    数据区含"分部标题行"（仅名称无数值）与小计/合计行。
- make_e6_summary：表-07-1 单位工程竣工结算汇总表
  ——序号/汇总内容/金额 三列资金汇总表 + 尾部 编制人/复核人 签字行
   （历史上被 weak 表单误判）。
"""
from __future__ import annotations

from pathlib import Path

import openpyxl
from openpyxl.styles import Border, Font, Side

_THIN = Side(style="thin", color="9AA5B1")
_BORDER = Border(left=_THIN, right=_THIN, top=_THIN, bottom=_THIN)


def _finalize(wb, path: Path) -> None:
    wb.properties.creator = "Jiadun 国标表式合成生成器"
    wb.properties.created = wb.properties.modified = __import__(
        "datetime").datetime(2026, 1, 1)
    wb.save(path)


def make_table_08(path: Path) -> None:
    """表-08：三层表头 + 分部标题行 + 明细 + 小计/合计。返回 (明细行数, 小计行数)。"""
    wb = openpyxl.Workbook()
    wb.remove(wb.active)
    # Excel/WPS 工作表名上限为 31 个字符；保留表号、业务语义和合成标记，
    # 避免测试夹具自身制造兼容性警告。
    ws = wb.create_sheet("F.1 表-08 分部分项清单（合成单位工程）")
    ws.cell(row=1, column=1, value="分部分项工程和单价措施项目清单与计价表").font = Font(bold=True)
    ws.merge_cells(start_row=1, start_column=1, end_row=1, end_column=11)
    ws.cell(row=2, column=1, value="工程名称：合成市政道路工程")
    ws.merge_cells(start_row=2, start_column=1, end_row=2, end_column=6)
    ws.cell(row=2, column=7, value="标段：/")
    # ---- 三层表头（第3-5行）----
    ws.cell(row=3, column=1, value="序号")
    ws.cell(row=3, column=2, value="项目编码")
    ws.cell(row=3, column=3, value="项目名称")
    ws.cell(row=3, column=4, value="项目特征描述")
    ws.cell(row=3, column=5, value="计量\n单位")
    ws.cell(row=3, column=6, value="工程量")
    ws.cell(row=3, column=7, value="金额（元）")
    ws.merge_cells(start_row=3, start_column=7, end_row=3, end_column=8)  # 金额横跨 综合单价+合价
    ws.cell(row=3, column=9, value="其中")
    ws.merge_cells(start_row=3, start_column=9, end_row=3, end_column=11)
    for c, t in ((7, "综合单价"), (8, "合价")):
        ws.cell(row=4, column=c, value=t)
    ws.cell(row=4, column=9, value="定额人工费")
    ws.cell(row=4, column=10, value="定额机械费")
    ws.cell(row=4, column=11, value="暂估价")
    # ---- 数据区 ----
    ws.cell(row=6, column=3, value="路基处理")  # 分部标题行（仅名称，无数值）
    items = [
        ("1", "040202001001", "道路床碾压整型", "1.压实度符合规范", "m2", 7194.06, 1.55, 11150.79),
        ("2", "040202011001", "级配碎石垫层", "1.15cm 厚", "m2", 7194.06, 41.02, 295100.34),
        ("3", "040203006001", "沥青混凝土面层", "1.AC-20C 中面层", "m2", 6480.00, 128.50, 832680.00),
    ]
    r = 7
    for it in items:
        for c, v in enumerate(it, 1):
            ws.cell(row=r, column=c, value=v)
        r += 1
    subtotal_amount = sum(i[7] for i in items)
    # 仅一行"合计"（单页表）：真实分页表的"本页小计+合计"并存会令 C 路径
    # 累加翻倍——作为已知局限记录，见测试与报告。
    ws.cell(row=r, column=2, value="合计")
    ws.cell(row=r, column=8, value=subtotal_amount)
    for row in ws.iter_rows(min_row=3, max_row=r, max_col=11):
        for c in row:
            c.border = _BORDER
    for letter, width in (("A", 6), ("B", 14), ("C", 22), ("D", 24), ("E", 8),
                          ("F", 10), ("G", 10), ("H", 12), ("I", 12), ("J", 12), ("K", 10)):
        ws.column_dimensions[letter].width = width
    _finalize(wb, path)


def make_e6_summary(path: Path) -> None:
    """表-07-1 单位工程竣工结算汇总表：三列资金汇总 + 尾部签字行。"""
    wb = openpyxl.Workbook()
    wb.remove(wb.active)
    # 同样遵守 31 字符限制，表号和“竣工结算汇总”语义保持不变。
    ws = wb.create_sheet("E.6 表-07-1 竣工结算汇总（合成单位工程）")
    ws.cell(row=1, column=1, value="单位工程竣工结算汇总表").font = Font(bold=True)
    ws.merge_cells(start_row=1, start_column=1, end_row=1, end_column=4)
    ws.cell(row=2, column=1, value="（适用于一般计税方法）")
    ws.cell(row=3, column=1, value="工程名称：合成市政道路工程")
    ws.merge_cells(start_row=3, start_column=1, end_row=3, end_column=2)
    ws.cell(row=3, column=3, value="标段：/")
    ws.cell(row=4, column=1, value="序号")
    ws.cell(row=4, column=2, value="汇总内容")
    ws.merge_cells(start_row=4, start_column=2, end_row=4, end_column=3)
    ws.cell(row=4, column=4, value="金  额（元）")
    rows = [
        ("1", "分部分项及单价措施项目", 50585822.35),
        ("1.1", "道路路面结构", 3204625.38),
        ("2", "措施项目", 2153000.00),
        ("3", "其他项目", 890000.00),
        ("4", "规费", 1450230.11),
        ("5", "税金", 4860000.00),
    ]
    r = 5
    for no, name, amount in rows:
        ws.cell(row=r, column=1, value=no)
        ws.cell(row=r, column=2, value=name)
        ws.merge_cells(start_row=r, start_column=2, end_row=r, end_column=3)
        ws.cell(row=r, column=4, value=amount)
        r += 1
    for c, t in ((1, "编制人：合成人员"), (3, "复核人：合成人员"), (4, "法定代表人：合成人员")):
        ws.cell(row=r, column=c, value=t)
    for row in ws.iter_rows(min_row=4, max_row=r, max_col=4):
        for c in row:
            c.border = _BORDER
    for letter, width in (("A", 8), ("B", 26), ("C", 10), ("D", 16)):
        ws.column_dimensions[letter].width = width
    _finalize(wb, path)

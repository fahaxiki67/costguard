"""Excel 解析保真层（ADR-008 第一段）。

原则：
- 只"读"，不"解释"：逐格落库 raw_cells，保留公式/缓存值/文本数字/合并拓扑/隐藏行列；
- 永远只读只读副本（originals/），绝不写回；
- .xlsx 用 openpyxl 双读（公式视图 + 缓存值视图）；
- .xls 用 xlrd（无公式缓存与合并单元格信息，作为已知限制记录）；
- .csv 单文件即单表；
- pdf/docx/image 在本 Phase 只登记"解析器未实现"，不假装解析。
"""
from __future__ import annotations

import csv
import hashlib
import posixpath
import re
import sqlite3
import zipfile
from dataclasses import dataclass, field
from datetime import datetime
from pathlib import Path
from xml.etree import ElementTree

from jiadun.core.engine.money import NotANumberError, to_decimal

# 前多少行参与表头识别
HEADER_SCAN_ROWS = 15
SHEET_MAX_EMPTY_TAIL = 200


@dataclass
class CellRecord:
    row: int  # 1-based
    col: int  # 1-based
    raw_value: str | None
    cached_value: str | None
    is_formula: bool
    is_number_stored_as_text: bool
    num_fmt: str
    # 公式缓存是源文件的结构事实，不把缺失/错误缓存伪装成普通文本。
    # ``cached`` 表示文件里存在缓存值，但不代表已经由本程序重算。
    formula_cache_status: str = "not_formula"


@dataclass
class SheetRecord:
    sheet_index: int
    sheet_name: str
    n_rows: int
    n_cols: int
    # 工作簿级可见状态：openpyxl 的 sheet_state（visible/hidden/veryHidden）。
    # 无法取得时保持 None，展示层必须作「未知」，不得默认可见。
    visible_state: str | None = None
    merged_ranges: list[str] = field(default_factory=list)
    hidden_rows: list[int] = field(default_factory=list)
    hidden_cols: list[int] = field(default_factory=list)
    cells: list[CellRecord] = field(default_factory=list)
    # 筛选可见性由 Office 应用维护，openpyxl 只能读到范围/条件，不能证明
    # 当前文件中哪些行在用户最后一次筛选时可见；因此存在筛选元数据时默认
    # 记为 unknown，由导入/异常层保守门控。
    auto_filter_ref: str | None = None
    filter_conditions: list[dict] = field(default_factory=list)
    table_ranges: list[dict] = field(default_factory=list)
    filter_state: str = "none"
    formula_metadata: dict = field(default_factory=dict)


@dataclass
class ParseResult:
    parser: str
    # ``partial`` 表示已保留可预览的原始网格，但解析器能力不足以证明
    # 结构完整（例如旧式 .xls 无法保留公式/合并拓扑）。调用方必须继续
    # 落库以便人工复核，不能把它当作 ``ok``。
    status: str  # 'ok' | 'partial' | 'unsupported' | 'failed'
    sheets: list[SheetRecord] = field(default_factory=list)
    stats: dict = field(default_factory=dict)
    error: str = ""


class ParseResultValidationError(ValueError):
    """解析结果元数据不满足可安全落库的结构合同。"""


def _validate_parse_result_for_persistence(result: ParseResult) -> None:
    """在任何 INSERT 之前校验 Sheet 身份、网格范围和逐格坐标。"""
    errors: list[str] = []
    indices: list[int] = []
    for position, sheet in enumerate(result.sheets):
        index = sheet.sheet_index
        if isinstance(index, bool) or not isinstance(index, int) or index < 0:
            errors.append(f"第{position + 1}个 Sheet 的 sheet_index 无效：{index!r}")
            continue
        indices.append(index)
        valid_grid = (
            isinstance(sheet.n_rows, int)
            and not isinstance(sheet.n_rows, bool)
            and sheet.n_rows >= 0
            and isinstance(sheet.n_cols, int)
            and not isinstance(sheet.n_cols, bool)
            and sheet.n_cols >= 0
        )
        if not valid_grid:
            errors.append(
                f"Sheet「{sheet.sheet_name}」声明网格无效："
                f"{sheet.n_rows!r}×{sheet.n_cols!r}"
            )
            continue
        for cell in sheet.cells:
            if (
                isinstance(cell.row, bool)
                or not isinstance(cell.row, int)
                or cell.row < 1
                or cell.row > int(sheet.n_rows)
                or isinstance(cell.col, bool)
                or not isinstance(cell.col, int)
                or cell.col < 1
                or cell.col > int(sheet.n_cols)
            ):
                errors.append(
                    f"Sheet「{sheet.sheet_name}」存在超出声明网格的原始单元格："
                    f"行{cell.row}列{cell.col}（网格 {sheet.n_rows}×{sheet.n_cols}）"
                )
                # 记录每个 Sheet 的第一处越界即可，避免异常文本膨胀。
                break
    expected = list(range(len(result.sheets)))
    if indices != expected or len(indices) != len(set(indices)):
        errors.append(
            f"解析结果 Sheet 索引必须从 0 开始连续且唯一：实际={indices}，期望={expected}"
        )
    if errors:
        raise ParseResultValidationError("；".join(dict.fromkeys(errors)))


def _num_stored_as_text(value) -> bool:
    """值是字符串、且看起来是数字（含千分位/¥/括号负数）。"""
    if not isinstance(value, str):
        return False
    try:
        to_decimal(value)
        return True
    except NotANumberError:
        return False


_OPAQUE_FORMULA_MARKERS = (
    "INDIRECT(", "OFFSET(", "[", "]", "TABLE(", "@",
)


def _formula_cache_status(formula: str | None, cached) -> str:
    """返回可供 fail-closed 规则消费的公式缓存状态。

    这里只做保守的结构识别，不尝试在 Python 中重算 Excel 公式。跨工作簿、
    动态引用、结构化引用等公式即使携带缓存值，也不能把该缓存当作本程序
    已验证的金额依据。
    """
    if not formula:
        return "not_formula"
    text = str(formula).strip()
    cached_text = _cached_str(cached)
    if cached_text is None:
        return "missing"
    if "#REF!" in text.upper() or cached_text.lstrip().startswith("#"):
        return "error"
    upper = text.upper()
    if any(marker in upper for marker in _OPAQUE_FORMULA_MARKERS):
        return "opaque"
    return "cached"


def _filter_object_conditions(auto_filter, source: str) -> list[dict]:
    """将 worksheet/table AutoFilter 的范围与条件转为稳定 JSON。

    openpyxl 的 FilterColumn 是 Serialisable 对象，直接 ``__dict__`` 会混入
    不稳定的对象表示；这里仅取可审计的列号、筛选值和条件属性。
    """
    if auto_filter is None:
        return []
    out: list[dict] = []
    ref = getattr(auto_filter, "ref", None)
    filter_columns = getattr(auto_filter, "filterColumn", None) or []
    for column in filter_columns:
        entry: dict[str, object] = {
            "source": source,
            "kind": "filter_column",
            "ref": str(ref) if ref else None,
            # OOXML colId is zero based;保留两个口径，避免后续证据把列号错位。
            "col_id": int(column.colId) if column.colId is not None else None,
            "column": (int(column.colId) + 1) if column.colId is not None else None,
        }
        filters = getattr(column, "filters", None)
        if filters is not None:
            values = getattr(filters, "filter", None) or []
            dates = getattr(filters, "dateGroupItem", None) or []
            entry["values"] = [str(value) for value in values]
            entry["blank"] = bool(getattr(filters, "blank", False))
            if dates:
                entry["date_group_items"] = [
                    {
                        key: getattr(item, key)
                        for key in (
                            "year", "month", "day", "hour", "minute", "second", "dateTimeGrouping"
                        )
                        if getattr(item, key, None) is not None
                    }
                    for item in dates
                ]
        custom = getattr(column, "customFilters", None)
        if custom is not None:
            entry["custom_filters"] = [
                {
                    "operator": getattr(item, "operator", None),
                    "value": getattr(item, "val", None),
                }
                for item in (getattr(custom, "customFilter", None) or [])
            ]
        dynamic = getattr(column, "dynamicFilter", None)
        if dynamic is not None:
            entry["dynamic_filter"] = {
                key: getattr(dynamic, key, None)
                for key in ("type", "val", "maxVal")
                if getattr(dynamic, key, None) is not None
            }
        top10 = getattr(column, "top10", None)
        if top10 is not None:
            entry["top10"] = {
                key: getattr(top10, key, None)
                for key in ("top", "percent", "val")
                if getattr(top10, key, None) is not None
            }
        out.append(entry)
    return out


def _sheet_filter_metadata(ws) -> tuple[str | None, list[dict], list[dict], str]:
    """抽取 worksheet 与表格级筛选的范围、条件及未知可见性状态。"""
    worksheet_filter = getattr(ws, "auto_filter", None)
    worksheet_ref = getattr(worksheet_filter, "ref", None)
    conditions = _filter_object_conditions(worksheet_filter, "worksheet.auto_filter")
    active_filter_condition = any(item.get("kind") == "filter_column" for item in conditions)
    if worksheet_ref and not conditions:
        conditions.append({
            "source": "worksheet.auto_filter",
            "kind": "range_only",
            "ref": str(worksheet_ref),
        })
    tables: list[dict] = []
    table_has_filter = False
    for table_name in getattr(ws, "tables", {}).keys():
        table = ws.tables[table_name]
        table_filter = getattr(table, "autoFilter", None)
        table_ref = getattr(table, "ref", None)
        table_filter_ref = getattr(table_filter, "ref", None) if table_filter is not None else None
        tables.append({
            "name": str(getattr(table, "displayName", None) or table_name),
            "ref": str(table_ref) if table_ref else None,
            "auto_filter_ref": str(table_filter_ref) if table_filter_ref else None,
        })
        if table_filter is not None:
            table_has_filter = True
            table_conditions = _filter_object_conditions(table_filter, "table.auto_filter")
            if table_filter_ref and not table_conditions:
                table_conditions.append({
                    "source": "table.auto_filter",
                    "kind": "range_only",
                    "ref": str(table_filter_ref),
                    "table": str(getattr(table, "displayName", None) or table_name),
                })
            for item in table_conditions:
                item.setdefault("table", str(getattr(table, "displayName", None) or table_name))
            conditions.extend(table_conditions)
            active_filter_condition = active_filter_condition or any(
                item.get("kind") == "filter_column" for item in table_conditions
            )
    if active_filter_condition:
        filter_state = "unknown"
    elif worksheet_ref or table_has_filter:
        # 只有范围/下拉按钮、没有实际条件时，文件内没有可见性变化需要猜测；
        # 范围仍完整保存，供上层报告和复核使用。
        filter_state = "range_only"
    else:
        filter_state = "none"
    return (
        str(worksheet_ref) if worksheet_ref else None,
        conditions,
        tables,
        filter_state,
    )


def parse_xlsx(path: Path) -> ParseResult:
    import openpyxl

    wb_formula = openpyxl.load_workbook(path, data_only=False, read_only=False)
    # 缓存值只用于逐格证据比对，不需要第二份可编辑工作簿。只读缓存必须
    # 与公式视图按行并行迭代；禁止对 ReadOnlyWorksheet 逐格调用 ``cell``，
    # 那会退化为大文件上的 O(n²) 扫描。
    wb_cached = openpyxl.load_workbook(path, data_only=True, read_only=True)
    result = ParseResult(parser="openpyxl", status="ok")
    stats = {
        "n_sheets": 0,
        "n_cells": 0,
        "n_formulas": 0,
        "n_formula_unverified": 0,
        "n_filter_sheets": 0,
        "n_text_numbers": 0,
        "error_cells": [],
    }

    try:
        for idx, ws in enumerate(wb_formula.worksheets):
            ws_c = wb_cached.worksheets[idx]
            auto_filter_ref, filter_conditions, table_ranges, filter_state = (
                _sheet_filter_metadata(ws)
            )
            sheet = SheetRecord(
                sheet_index=idx,
                sheet_name=ws.title,
                n_rows=ws.max_row or 0,
                n_cols=ws.max_column or 0,
                visible_state=str(getattr(ws, "sheet_state", None) or "") or None,
                merged_ranges=[str(r) for r in ws.merged_cells.ranges],
                hidden_rows=sorted(r for r, dim in ws.row_dimensions.items() if dim.hidden),
                hidden_cols=sorted(
                    (c.column if isinstance(c, int) else openpyxl.utils.column_index_from_string(c))
                    for c, dim in ws.column_dimensions.items()
                    if dim.hidden
                ),
                auto_filter_ref=auto_filter_ref,
                filter_conditions=filter_conditions,
                table_ranges=table_ranges,
                filter_state=filter_state,
            )
            formula_metadata = {
                "formula_count": 0,
                "cached_count": 0,
                "uncached_count": 0,
                "error_count": 0,
                "opaque_count": 0,
                "unverified_count": 0,
                "unverified_cells": [],
            }
            for row, cached_row in zip(ws.iter_rows(), ws_c.iter_rows(), strict=False):
                for cell in row:
                    v = cell.value
                    if v is None and cell.number_format in ("General", ""):
                        continue
                    cached = (
                        cached_row[cell.column - 1].value
                        if cell.column <= len(cached_row) else None
                    )
                    is_formula = cell.data_type == "f"
                    formula_cache_status = _formula_cache_status(v if is_formula else None, cached)
                    rec = CellRecord(
                        row=cell.row,
                        col=cell.column,
                        raw_value=str(v) if v is not None else None,
                        cached_value=_cached_str(cached),
                        is_formula=is_formula,
                        is_number_stored_as_text=_num_stored_as_text(v),
                        num_fmt=cell.number_format or "",
                        formula_cache_status=formula_cache_status,
                    )
                    sheet.cells.append(rec)
                    if is_formula:
                        stats["n_formulas"] += 1
                        formula_metadata["formula_count"] += 1
                        if formula_cache_status == "cached":
                            formula_metadata["cached_count"] += 1
                        else:
                            formula_metadata["unverified_count"] += 1
                            metadata_key = (
                                "uncached_count" if formula_cache_status == "missing"
                                else f"{formula_cache_status}_count"
                            )
                            formula_metadata[metadata_key] += 1
                            formula_metadata["unverified_cells"].append({
                                "row": cell.row,
                                "col": cell.column,
                                "coordinate": cell.coordinate,
                                "formula": str(v),
                                "cached_value": _cached_str(cached),
                                "cache_status": formula_cache_status,
                            })
                            stats["n_formula_unverified"] += 1
                    if rec.is_number_stored_as_text:
                        stats["n_text_numbers"] += 1
                    if isinstance(v, str) and v.startswith("#"):
                        stats["error_cells"].append(f"{ws.title}!{cell.coordinate}:{v}")
            sheet.formula_metadata = formula_metadata
            if filter_state != "none":
                stats["n_filter_sheets"] += 1
            result.sheets.append(sheet)
            stats["n_sheets"] += 1
            stats["n_cells"] += len(sheet.cells)
    finally:
        wb_formula.close()
        wb_cached.close()
    result.stats = stats
    return result


def _cached_str(v) -> str | None:
    if v is None:
        return None
    return str(v)


def parse_xls(path: Path) -> ParseResult:
    import xlrd

    book = xlrd.open_workbook(str(path))
    # xlrd 2.x 可以读取单元格文本和值，但不会保留 OOXML 那样的公式
    # 缓存、合并单元格拓扑或筛选可见性。即使数值碰巧相等，也不能让
    # 下游把 .xls 当作结构完整的输入；保留网格预览，同时显式降级。
    limitations = [
        "xls: xlrd 不保留公式文本/缓存状态",
        "xls: xlrd 不保留合并单元格拓扑",
        "xls: 筛选后的实际可见行无法由解析器证明",
    ]
    result = ParseResult(parser="xlrd", status="partial")
    stats = {"n_sheets": 0, "n_cells": 0, "n_formulas": 0, "n_text_numbers": 0,
             "limitations": limitations, "capability_limited": True}
    visibility_values = getattr(book, "sheet_visibility", None)
    if visibility_values is None:
        # xlrd 2.0.x exposes the BIFF sheet state only through this private
        # compatibility attribute. If neither source exists, keep visibility
        # unknown so the import layer can fail closed.
        visibility_values = getattr(book, "_sheet_visibility", None)
    if visibility_values is None:
        limitations.append("xls: 工作表 hidden/veryHidden 可见性无法取得")
    for idx in range(book.nsheets):
        sh = book.sheet_by_index(idx)
        # xlrd Book.sheet_visibility：0=visible、1=hidden、2=veryHidden；
        # 越界/异常时保持 None（未知），不得默认可见。
        _vis_map = {0: "visible", 1: "hidden", 2: "veryHidden"}
        try:
            _vis = _vis_map.get(int(visibility_values[idx]))
        except (AttributeError, IndexError, TypeError, ValueError):
            _vis = None
        sheet = SheetRecord(
            sheet_index=idx,
            sheet_name=sh.name,
            n_rows=sh.nrows,
            n_cols=sh.ncols,
            visible_state=_vis,
            hidden_rows=sorted(r for r, info in getattr(sh, "rowinfo_map", {}).items() if info.hidden),
            hidden_cols=sorted(c for c, info in getattr(sh, "colinfo_map", {}).items() if info.hidden),
            formula_metadata={
                "capability_limited": True,
                "limitations": limitations,
                "formula_support": "unavailable",
                "merged_range_support": "unavailable",
                "filter_visibility_support": "unavailable",
            },
        )
        for r in range(sh.nrows):
            for c in range(sh.ncols):
                cell = sh.cell(r, c)
                if cell.ctype == 0:  # EMPTY
                    continue
                text = _xls_cell_text(cell)
                sheet.cells.append(
                    CellRecord(
                        row=r + 1, col=c + 1,
                        raw_value=text, cached_value=text,
                        is_formula=False,
                        is_number_stored_as_text=_num_stored_as_text(text),
                        num_fmt="",
                    )
                )
        result.sheets.append(sheet)
        stats["n_sheets"] += 1
        stats["n_cells"] += len(sheet.cells)
    book.release_resources()
    result.stats = stats
    return result


def _xls_cell_text(cell) -> str:
    if cell.ctype == 2:  # NUMBER
        if float(cell.value).is_integer():
            return str(int(cell.value))
        return repr(cell.value)
    if cell.ctype == 3:  # DATE
        try:
            import xlrd

            dt = xlrd.xldate_as_datetime(cell.value, 0)
            return dt.strftime("%Y-%m-%d")
        except Exception:
            return str(cell.value)
    return str(cell.value)


def parse_csv(path: Path) -> ParseResult:
    result = ParseResult(parser="csv", status="ok")
    stats = {"n_sheets": 1, "n_cells": 0, "n_formulas": 0, "n_text_numbers": 0}
    sheet = SheetRecord(
        sheet_index=0, sheet_name=path.stem, n_rows=0, n_cols=0, visible_state="visible"
    )
    n_rows = 0
    with open(path, encoding="utf-8-sig", newline="") as f:
        for row in csv.reader(f):
            n_rows += 1
            for c, v in enumerate(row, start=1):
                if v == "":
                    continue
                sheet.cells.append(
                    CellRecord(row=n_rows, col=c, raw_value=v, cached_value=v,
                               is_formula=False, is_number_stored_as_text=_num_stored_as_text(v), num_fmt="")
                )
    sheet.n_rows = n_rows
    sheet.n_cols = max((rec.col for rec in sheet.cells), default=0)
    result.sheets = [sheet]
    stats["n_cells"] = len(sheet.cells)
    result.stats = stats
    return result


UNSUPPORTED = {"pdf", "doc", "docx", "image"}


_OOXML_MAIN_NS = "http://schemas.openxmlformats.org/spreadsheetml/2006/main"
_OOXML_REL_NS = "http://schemas.openxmlformats.org/officeDocument/2006/relationships"
_OOXML_PACKAGE_REL_NS = "http://schemas.openxmlformats.org/package/2006/relationships"
_CELL_REF_RE = re.compile(r"^\$?([A-Za-z]+)\$?([0-9]+)$")


def _column_number(label: str) -> int:
    value = 0
    for char in label.upper():
        value = value * 26 + ord(char) - ord("A") + 1
    return value


def _dimension_shape(ref: str | None) -> tuple[int, int] | None:
    """将 OOXML dimension 的末端单元格转换为最大行/列。"""
    if not ref:
        return None
    endpoint = str(ref).split(":")[-1].strip()
    match = _CELL_REF_RE.fullmatch(endpoint)
    if match is None:
        return None
    return int(match.group(2)), _column_number(match.group(1))


def _column_label(number: int) -> str:
    """把 1-based 列号转回 OOXML 使用范围可读的列名。"""
    if number <= 0:
        return "A"
    chars: list[str] = []
    while number:
        number, remainder = divmod(number - 1, 26)
        chars.append(chr(ord("A") + remainder))
    return "".join(reversed(chars))


def _worksheet_shape(worksheet_root: ElementTree.Element) -> tuple[str | None, tuple[int, int] | None, str]:
    """读取 worksheet 的使用范围；没有 dimension 时独立扫描 row/c。"""
    dimension = worksheet_root.find(f"{{{_OOXML_MAIN_NS}}}dimension")
    dimension_ref = str(dimension.attrib.get("ref")) if dimension is not None else None
    shape = _dimension_shape(dimension_ref)
    if shape is not None:
        return dimension_ref, shape, "dimension"

    min_row: int | None = None
    max_row: int | None = None
    min_col: int | None = None
    max_col: int | None = None
    sheet_data = worksheet_root.find(f"{{{_OOXML_MAIN_NS}}}sheetData")
    for row in list(sheet_data) if sheet_data is not None else []:
        try:
            row_number = int(row.attrib.get("r", ""))
        except (TypeError, ValueError):
            continue
        for cell in row.findall(f"{{{_OOXML_MAIN_NS}}}c"):
            ref = cell.attrib.get("r")
            match = _CELL_REF_RE.fullmatch(str(ref or ""))
            if match is None:
                continue
            col_number = _column_number(match.group(1))
            cell_row = int(match.group(2))
            min_row = cell_row if min_row is None else min(min_row, cell_row)
            max_row = cell_row if max_row is None else max(max_row, cell_row)
            min_col = col_number if min_col is None else min(min_col, col_number)
            max_col = col_number if max_col is None else max(max_col, col_number)
        # A row without cells still proves that this row exists, but not a
        # business column. Keep it only when at least one cell is found in the
        # worksheet; an entirely empty sheet remains range-unproven.
        if min_row is not None:
            min_row = min(min_row, row_number)
            max_row = max(max_row or row_number, row_number)
    if min_row is None or max_row is None or min_col is None or max_col is None:
        return None, None, "unavailable"
    fallback_ref = (
        f"{_column_label(min_col)}{min_row}:{_column_label(max_col)}{max_row}"
    )
    return fallback_ref, (max_row, max_col), "row_cell_scan"


def _worksheet_content_fingerprint(
    worksheet_root: ElementTree.Element,
) -> tuple[int, int, str]:
    """独立统计源 Sheet 的有值单元格与物理行坐标。

    ``dimension`` 只描述最大范围，无法发现“尺寸不变但中间漏掉一行”的
    解析错误。这里只依据 OOXML 中带值/公式的 ``c`` 节点生成坐标指纹，
    不解析业务字段、公式结果或 shared string 内容，避免与主解析器共享
    行分类逻辑；解析结果落库前再用同一坐标口径对照。
    """
    sheet_data = worksheet_root.find(f"{{{_OOXML_MAIN_NS}}}sheetData")
    coordinates: list[str] = []
    row_numbers: set[int] = set()
    for row in list(sheet_data) if sheet_data is not None else []:
        for cell in row.findall(f"{{{_OOXML_MAIN_NS}}}c"):
            ref = cell.attrib.get("r")
            match = _CELL_REF_RE.fullmatch(str(ref or ""))
            if match is None:
                continue
            formula = cell.find(f"{{{_OOXML_MAIN_NS}}}f")
            value_nodes = (
                cell.find(f"{{{_OOXML_MAIN_NS}}}v"),
                cell.find(f"{{{_OOXML_MAIN_NS}}}is"),
            )
            has_value = any(
                node is not None
                and any(str(text or "").strip() for text in node.itertext())
                for node in value_nodes
            )
            if formula is None and not has_value:
                continue
            row_number = int(match.group(2))
            coordinate = f"{row_number}:{_column_number(match.group(1))}"
            coordinates.append(coordinate)
            row_numbers.add(row_number)
    coordinates.sort()
    digest = hashlib.sha256("\n".join(coordinates).encode("utf-8")).hexdigest()
    return len(coordinates), len(row_numbers), digest


def _ooxml_part_target(target: str) -> str:
    """将 workbook 关系目标规范化为 ZIP 内部的绝对部件路径。"""
    target = str(target or "").replace("\\", "/")
    if target.startswith("/"):
        return target.lstrip("/")
    return posixpath.normpath(posixpath.join("xl", target))


def _source_xlsx_census(path: Path) -> dict[str, object]:
    """从 OOXML 包目录独立盘点源工作簿的 Sheet 顺序与使用范围。

    该过程不调用 ``parse_xlsx``、不读取解析后的 ``SheetRecord``，只读取
    源文件中的 workbook.xml、关系文件和各 worksheet 的 dimension；若某些
    生成器省略 dimension，则由 worksheet 的 row/c 坐标独立推导使用范围。
    这样即使解析器共同漏掉整张 Sheet，入口仍能以源文件清单发现范围不一致。
    """
    workbook_part = "xl/workbook.xml"
    rels_part = "xl/_rels/workbook.xml.rels"
    with zipfile.ZipFile(path) as package:
        workbook_root = ElementTree.fromstring(package.read(workbook_part))
        rels_root = ElementTree.fromstring(package.read(rels_part))
        relationships = {
            rel.attrib.get("Id", ""): _ooxml_part_target(rel.attrib.get("Target", ""))
            for rel in rels_root.findall(f"{{{_OOXML_PACKAGE_REL_NS}}}Relationship")
        }
        sheets: list[dict[str, object]] = []
        sheets_root = workbook_root.find(f"{{{_OOXML_MAIN_NS}}}sheets")
        for index, sheet in enumerate(
            list(sheets_root) if sheets_root is not None else [], start=0
        ):
            rel_id = sheet.attrib.get(f"{{{_OOXML_REL_NS}}}id", "")
            target = relationships.get(rel_id)
            if not target or target not in package.namelist():
                raise ValueError(f"Sheet 关系缺失：{sheet.attrib.get('name', '')}")
            worksheet_root = ElementTree.fromstring(package.read(target))
            dimension_ref, shape, range_method = _worksheet_shape(worksheet_root)
            populated_cell_count, populated_row_count, content_digest = (
                _worksheet_content_fingerprint(worksheet_root)
            )
            sheets.append({
                "sheet_index": index,
                "sheet_name": str(sheet.attrib.get("name", "")),
                "relationship_id": rel_id,
                "part": target,
                "dimension": dimension_ref,
                "range_status": "available" if shape is not None else "unavailable",
                "range_method": range_method,
                "n_rows": shape[0] if shape is not None else None,
                "n_cols": shape[1] if shape is not None else None,
                "populated_cell_count": populated_cell_count,
                "populated_row_count": populated_row_count,
                "populated_coordinate_sha256": content_digest,
            })
    return {
        "method": "ooxml_zip_directory",
        "status": "available",
        "file_type": "xlsx",
        "sheet_count": len(sheets),
        "sheets": sheets,
    }


def _source_csv_census(path: Path) -> dict[str, object]:
    rows = 0
    cols = 0
    with open(path, encoding="utf-8-sig", newline="") as handle:
        for row in csv.reader(handle):
            rows += 1
            cols = max(cols, len(row))
    return {
        "method": "csv_source_scan",
        "status": "available",
        "file_type": "csv",
        "sheet_count": 1,
        "sheets": [{
            "sheet_index": 0,
            "sheet_name": path.stem,
            "n_rows": rows,
            "n_cols": cols,
            "range_status": "available",
        }],
    }


def _source_xls_census(path: Path) -> dict[str, object]:
    """盘点旧式 XLS 的工作表目录；结构能力仍由主解析器降级处理。"""
    import xlrd

    book = xlrd.open_workbook(str(path), on_demand=True)
    try:
        sheets = [
            {
                "sheet_index": index,
                "sheet_name": str(book.sheet_by_index(index).name),
                "n_rows": int(book.sheet_by_index(index).nrows),
                "n_cols": int(book.sheet_by_index(index).ncols),
                "range_status": "available",
            }
            for index in range(book.nsheets)
        ]
    finally:
        book.release_resources()
    return {
        "method": "xls_source_directory",
        "status": "available",
        "file_type": "xls",
        "sheet_count": len(sheets),
        "sheets": sheets,
    }


def _source_workbook_census(path: Path, file_type: str) -> dict[str, object]:
    """从源文件建立独立 Sheet 清单；失败时返回可序列化的不可用状态。"""
    try:
        if file_type == "xlsx":
            return _source_xlsx_census(path)
        if file_type == "xls":
            return _source_xls_census(path)
        if file_type == "csv":
            return _source_csv_census(path)
    except Exception as exc:  # noqa: BLE001 - 入口必须保留失败证据并继续主解析
        return {
            "method": "source_scan",
            "status": "unavailable",
            "file_type": file_type,
            "reason": f"{type(exc).__name__}: {exc}",
            "sheet_count": None,
            "sheets": [],
        }
    return {
        "method": "source_scan",
        "status": "unavailable",
        "file_type": file_type,
        "reason": "当前文件类型没有可用的源工作簿 Sheet 盘点器",
        "sheet_count": None,
        "sheets": [],
    }


def _compare_source_census(census: dict[str, object], sheets: list[SheetRecord]) -> tuple[str, list[str]]:
    """比较源 Sheet 目录与解析落库前的结果，返回状态和差异说明。"""
    if census.get("status") != "available":
        return "unavailable", [str(census.get("reason") or "源文件 Sheet 范围无法盘点")]
    source_sheets = census.get("sheets")
    if not isinstance(source_sheets, list):
        return "unavailable", ["源文件 Sheet 清单不是数组"]
    differences: list[str] = []
    if len(source_sheets) != len(sheets):
        differences.append(
            f"源文件 Sheet 数量={len(source_sheets)}，解析结果数量={len(sheets)}"
        )
    parsed_indices: list[object] = [sheet.sheet_index for sheet in sheets]
    expected_indices = list(range(len(sheets)))
    if (
        any(isinstance(index, bool) or not isinstance(index, int) for index in parsed_indices)
        or parsed_indices != expected_indices
        or len(parsed_indices) != len(set(parsed_indices))
    ):
        differences.append(
            f"解析结果 Sheet 索引必须从 0 开始连续且唯一："
            f"实际={parsed_indices}，期望={expected_indices}"
        )
    for index, (source, parsed) in enumerate(zip(source_sheets, sheets, strict=False)):
        if not isinstance(source, dict):
            differences.append(f"源文件第{index + 1}个 Sheet 目录项无效")
            continue
        source_name = str(source.get("sheet_name") or "")
        if source.get("sheet_index") != parsed.sheet_index:
            differences.append(
                f"第{index + 1}个 Sheet 索引不一致："
                f"源文件={source.get('sheet_index')!r}/解析={parsed.sheet_index!r}"
            )
        if source_name != parsed.sheet_name:
            differences.append(
                f"第{index + 1}个 Sheet 名称不一致：源文件「{source_name}」/解析「{parsed.sheet_name}」"
            )
        if source.get("range_status") != "available":
            differences.append(f"Sheet「{source_name}」源文件使用范围无法独立盘点")
        source_rows = source.get("n_rows")
        source_cols = source.get("n_cols")
        if source_rows is not None and int(source_rows) != int(parsed.n_rows):
            differences.append(
                f"Sheet「{source_name}」行数不一致：源文件={source_rows}/解析={parsed.n_rows}"
            )
        if source_cols is not None and int(source_cols) != int(parsed.n_cols):
            differences.append(
                f"Sheet「{source_name}」列数不一致：源文件={source_cols}/解析={parsed.n_cols}"
            )
        source_cell_count = source.get("populated_cell_count")
        if source_cell_count is not None:
            parsed_cell_count = sum(
                1
                for cell in parsed.cells
                if cell.is_formula or cell.raw_value not in (None, "")
            )
            if int(source_cell_count) != parsed_cell_count:
                differences.append(
                    f"Sheet「{source_name}」有值单元格数不一致："
                    f"源文件={source_cell_count}/解析={parsed_cell_count}"
                )
        source_row_count = source.get("populated_row_count")
        if source_row_count is not None:
            parsed_rows = {
                int(cell.row)
                for cell in parsed.cells
                if cell.is_formula or cell.raw_value not in (None, "")
            }
            if int(source_row_count) != len(parsed_rows):
                differences.append(
                    f"Sheet「{source_name}」有值物理行数不一致："
                    f"源文件={source_row_count}/解析={len(parsed_rows)}"
                )
        source_digest = source.get("populated_coordinate_sha256")
        if source_digest:
            parsed_coordinates = sorted({
                f"{int(cell.row)}:{int(cell.col)}"
                for cell in parsed.cells
                if cell.is_formula or cell.raw_value not in (None, "")
            })
            parsed_digest = hashlib.sha256(
                "\n".join(parsed_coordinates).encode("utf-8")
            ).hexdigest()
            if str(source_digest) != parsed_digest:
                differences.append(
                    f"Sheet「{source_name}」有值单元格坐标指纹不一致"
                )
    return ("complete" if not differences else "mismatch"), differences


def _attach_source_census(result: ParseResult, census: dict[str, object]) -> ParseResult:
    stats = dict(result.stats)
    status, differences = _compare_source_census(census, result.sheets)
    stats["source_census"] = census
    stats["source_census_status"] = status
    if differences:
        stats["source_census_differences"] = differences
    result.stats = stats
    return result


def parse_file(path: Path, file_type: str) -> ParseResult:
    """按文件类型分发。未知类型/未实现类型返回明确状态。"""
    path = Path(path)
    if file_type in UNSUPPORTED:
        return ParseResult(parser="none", status="unsupported",
                           error=f"parser for '{file_type}' not implemented yet (planned Phase 6+)")
    if not path.exists():
        return ParseResult(parser="none", status="failed", error=f"file not found: {path}")
    census = _source_workbook_census(path, file_type)
    try:
        if file_type == "xlsx":
            return _attach_source_census(parse_xlsx(path), census)
        if file_type == "xls":
            return _attach_source_census(parse_xls(path), census)
        if file_type == "csv":
            return _attach_source_census(parse_csv(path), census)
        if file_type in UNSUPPORTED:
            return ParseResult(parser="none", status="unsupported",
                               error=f"parser for '{file_type}' not implemented yet (planned Phase 6+)")
        return _attach_source_census(
            ParseResult(parser="none", status="failed", error=f"unknown file_type: {file_type}"),
            census,
        )
    except Exception as exc:  # 解析器级失败必须显式报告，不静默
        return _attach_source_census(
            ParseResult(parser="unknown", status="failed", error=f"{type(exc).__name__}: {exc}"),
            census,
        )


def persist_parse_result(
    conn: sqlite3.Connection,
    file_id: int,
    result: ParseResult,
    *,
    commit: bool = True,
) -> int:
    """把解析结果落库。返回 batch_id。

    ``failed``/``unsupported`` 解析批次也必须落库；调用方可以传
    ``commit=False``，把批次和对应的 Evidence 放在同一个外层事务内。
    """
    _validate_parse_result_for_persistence(result)
    now = datetime.now().isoformat(timespec="seconds")
    stats = dict(result.stats)
    if result.error:
        # 解析错误是输入证据的一部分；不要只放在内存 ImportReport 里，
        # 否则项目重新打开后无法解释为何该文件没有 Sheet。
        stats["error"] = result.error

    def _insert() -> int:
        cur = conn.execute(
            "INSERT INTO parse_batches(file_id, parser, parsed_at, status, stats_json) VALUES (?,?,?,?,?)",
            (file_id, result.parser, now, result.status, _dumps(stats)),
        )
        batch_id = cur.lastrowid
        for sheet in result.sheets:
            cur = conn.execute(
                """INSERT INTO raw_sheets(batch_id, sheet_index, sheet_name, n_rows, n_cols,
                   visible_state,
                   merged_ranges_json, hidden_rows_json, hidden_cols_json,
                   auto_filter_ref, filter_conditions_json, table_ranges_json,
                   filter_state, formula_metadata_json)
                   VALUES (?,?,?,?,?,?,?,?,?,?,?,?,?,?)""",
                (batch_id, sheet.sheet_index, sheet.sheet_name, sheet.n_rows, sheet.n_cols,
                 sheet.visible_state,
                 _dumps(sheet.merged_ranges), _dumps(sheet.hidden_rows), _dumps(sheet.hidden_cols),
                 sheet.auto_filter_ref, _dumps(sheet.filter_conditions), _dumps(sheet.table_ranges),
                 sheet.filter_state, _dumps(sheet.formula_metadata)),
            )
            sheet_id = cur.lastrowid
            conn.executemany(
                """INSERT INTO raw_cells(sheet_id, row, col, raw_value, cached_value,
                   is_formula, is_number_stored_as_text, num_fmt, formula_cache_status)
                   VALUES (?,?,?,?,?,?,?,?,?)""",
                [
                    (sheet_id, c.row, c.col, c.raw_value, c.cached_value,
                     int(c.is_formula), int(c.is_number_stored_as_text), c.num_fmt,
                     c.formula_cache_status)
                    for c in sheet.cells
                ],
            )
        return int(batch_id)

    if commit:
        with conn:
            return _insert()
    return _insert()


def _dumps(obj) -> str:
    import json

    return json.dumps(obj, ensure_ascii=False)

"""Excel 解析保真层（ADR-008 第一段）。

原则：
- 只"读"，不"解释"：逐格落库 raw_cells，保留公式/缓存值/文本数字/合并拓扑/隐藏行列；
- 永远只读只读副本（originals/），绝不写回；
- .xlsx 用 openpyxl 双读（公式视图 + 缓存值视图）；
- .xls 用 xlrd（无公式缓存与合并单元格信息，作为已知限制记录）；
- .csv 单文件即单表；
- pdf/docx/image 在本 Phase 只登记"解析器未实现"，不假装解析。
"""
from __future__ import annotations

import csv
import sqlite3
from dataclasses import dataclass, field
from datetime import datetime
from pathlib import Path

from costguard.core.engine.money import NotANumberError, to_decimal

# 前多少行参与表头识别
HEADER_SCAN_ROWS = 15
SHEET_MAX_EMPTY_TAIL = 200


@dataclass
class CellRecord:
    row: int  # 1-based
    col: int  # 1-based
    raw_value: str | None
    cached_value: str | None
    is_formula: bool
    is_number_stored_as_text: bool
    num_fmt: str


@dataclass
class SheetRecord:
    sheet_index: int
    sheet_name: str
    n_rows: int
    n_cols: int
    merged_ranges: list[str] = field(default_factory=list)
    hidden_rows: list[int] = field(default_factory=list)
    hidden_cols: list[int] = field(default_factory=list)
    cells: list[CellRecord] = field(default_factory=list)


@dataclass
class ParseResult:
    parser: str
    status: str  # 'ok' | 'unsupported' | 'failed'
    sheets: list[SheetRecord] = field(default_factory=list)
    stats: dict = field(default_factory=dict)
    error: str = ""


def _num_stored_as_text(value) -> bool:
    """值是字符串、且看起来是数字（含千分位/¥/括号负数）。"""
    if not isinstance(value, str):
        return False
    try:
        to_decimal(value)
        return True
    except NotANumberError:
        return False


def parse_xlsx(path: Path) -> ParseResult:
    import openpyxl

    wb_formula = openpyxl.load_workbook(path, data_only=False, read_only=False)
    wb_cached = openpyxl.load_workbook(path, data_only=True, read_only=False)
    result = ParseResult(parser="openpyxl", status="ok")
    stats = {"n_sheets": 0, "n_cells": 0, "n_formulas": 0, "n_text_numbers": 0, "error_cells": []}

    try:
        for idx, ws in enumerate(wb_formula.worksheets):
            ws_c = wb_cached.worksheets[idx]
            sheet = SheetRecord(
                sheet_index=idx,
                sheet_name=ws.title,
                n_rows=ws.max_row or 0,
                n_cols=ws.max_column or 0,
                merged_ranges=[str(r) for r in ws.merged_cells.ranges],
                hidden_rows=sorted(r for r, dim in ws.row_dimensions.items() if dim.hidden),
                hidden_cols=sorted(
                    (c.column if isinstance(c, int) else openpyxl.utils.column_index_from_string(c))
                    for c, dim in ws.column_dimensions.items()
                    if dim.hidden
                ),
            )
            for row in ws.iter_rows():
                for cell in row:
                    v = cell.value
                    if v is None and cell.number_format in ("General", ""):
                        continue
                    cached = ws_c.cell(row=cell.row, column=cell.column).value
                    is_formula = cell.data_type == "f"
                    rec = CellRecord(
                        row=cell.row,
                        col=cell.column,
                        raw_value=str(v) if v is not None else None,
                        cached_value=_cached_str(cached),
                        is_formula=is_formula,
                        is_number_stored_as_text=_num_stored_as_text(v),
                        num_fmt=cell.number_format or "",
                    )
                    sheet.cells.append(rec)
                    if is_formula:
                        stats["n_formulas"] += 1
                    if rec.is_number_stored_as_text:
                        stats["n_text_numbers"] += 1
                    if isinstance(v, str) and v.startswith("#"):
                        stats["error_cells"].append(f"{ws.title}!{cell.coordinate}:{v}")
            result.sheets.append(sheet)
            stats["n_sheets"] += 1
            stats["n_cells"] += len(sheet.cells)
    finally:
        wb_formula.close()
        wb_cached.close()
    result.stats = stats
    return result


def _cached_str(v) -> str | None:
    if v is None:
        return None
    return str(v)


def parse_xls(path: Path) -> ParseResult:
    import xlrd

    book = xlrd.open_workbook(str(path))
    result = ParseResult(parser="xlrd", status="ok")
    stats = {"n_sheets": 0, "n_cells": 0, "n_formulas": 0, "n_text_numbers": 0,
             "limitations": "xls: no formula cache/merged-cell info via xlrd 2.x"}
    for idx in range(book.nsheets):
        sh = book.sheet_by_index(idx)
        sheet = SheetRecord(
            sheet_index=idx,
            sheet_name=sh.name,
            n_rows=sh.nrows,
            n_cols=sh.ncols,
            hidden_rows=sorted(r for r, info in getattr(sh, "rowinfo_map", {}).items() if info.hidden),
            hidden_cols=sorted(c for c, info in getattr(sh, "colinfo_map", {}).items() if info.hidden),
        )
        for r in range(sh.nrows):
            for c in range(sh.ncols):
                cell = sh.cell(r, c)
                if cell.ctype == 0:  # EMPTY
                    continue
                text = _xls_cell_text(cell)
                sheet.cells.append(
                    CellRecord(
                        row=r + 1, col=c + 1,
                        raw_value=text, cached_value=text,
                        is_formula=False,
                        is_number_stored_as_text=_num_stored_as_text(text),
                        num_fmt="",
                    )
                )
        result.sheets.append(sheet)
        stats["n_sheets"] += 1
        stats["n_cells"] += len(sheet.cells)
    book.release_resources()
    result.stats = stats
    return result


def _xls_cell_text(cell) -> str:
    if cell.ctype == 2:  # NUMBER
        if float(cell.value).is_integer():
            return str(int(cell.value))
        return repr(cell.value)
    if cell.ctype == 3:  # DATE
        try:
            import xlrd

            dt = xlrd.xldate_as_datetime(cell.value, 0)
            return dt.strftime("%Y-%m-%d")
        except Exception:
            return str(cell.value)
    return str(cell.value)


def parse_csv(path: Path) -> ParseResult:
    result = ParseResult(parser="csv", status="ok")
    stats = {"n_sheets": 1, "n_cells": 0, "n_formulas": 0, "n_text_numbers": 0}
    sheet = SheetRecord(sheet_index=0, sheet_name=path.stem, n_rows=0, n_cols=0)
    n_rows = 0
    with open(path, encoding="utf-8-sig", newline="") as f:
        for row in csv.reader(f):
            n_rows += 1
            for c, v in enumerate(row, start=1):
                if v == "":
                    continue
                sheet.cells.append(
                    CellRecord(row=n_rows, col=c, raw_value=v, cached_value=v,
                               is_formula=False, is_number_stored_as_text=_num_stored_as_text(v), num_fmt="")
                )
    sheet.n_rows = n_rows
    sheet.n_cols = max((rec.col for rec in sheet.cells), default=0)
    result.sheets = [sheet]
    stats["n_cells"] = len(sheet.cells)
    result.stats = stats
    return result


UNSUPPORTED = {"pdf", "doc", "docx", "image"}


def parse_file(path: Path, file_type: str) -> ParseResult:
    """按文件类型分发。未知类型/未实现类型返回明确状态。"""
    path = Path(path)
    if file_type in UNSUPPORTED:
        return ParseResult(parser="none", status="unsupported",
                           error=f"parser for '{file_type}' not implemented yet (planned Phase 6+)")
    if not path.exists():
        return ParseResult(parser="none", status="failed", error=f"file not found: {path}")
    try:
        if file_type == "xlsx":
            return parse_xlsx(path)
        if file_type == "xls":
            return parse_xls(path)
        if file_type == "csv":
            return parse_csv(path)
        if file_type in UNSUPPORTED:
            return ParseResult(parser="none", status="unsupported",
                               error=f"parser for '{file_type}' not implemented yet (planned Phase 6+)")
        return ParseResult(parser="none", status="failed", error=f"unknown file_type: {file_type}")
    except Exception as exc:  # 解析器级失败必须显式报告，不静默
        return ParseResult(parser="unknown", status="failed", error=f"{type(exc).__name__}: {exc}")


def persist_parse_result(conn: sqlite3.Connection, file_id: int, result: ParseResult) -> int:
    """把解析结果落库。返回 batch_id。"""
    now = datetime.now().isoformat(timespec="seconds")
    with conn:
        cur = conn.execute(
            "INSERT INTO parse_batches(file_id, parser, parsed_at, status, stats_json) VALUES (?,?,?,?,?)",
            (file_id, result.parser, now, result.status, _dumps(result.stats)),
        )
        batch_id = cur.lastrowid
        for sheet in result.sheets:
            cur = conn.execute(
                """INSERT INTO raw_sheets(batch_id, sheet_index, sheet_name, n_rows, n_cols,
                   merged_ranges_json, hidden_rows_json, hidden_cols_json) VALUES (?,?,?,?,?,?,?,?)""",
                (batch_id, sheet.sheet_index, sheet.sheet_name, sheet.n_rows, sheet.n_cols,
                 _dumps(sheet.merged_ranges), _dumps(sheet.hidden_rows), _dumps(sheet.hidden_cols)),
            )
            sheet_id = cur.lastrowid
            conn.executemany(
                """INSERT INTO raw_cells(sheet_id, row, col, raw_value, cached_value,
                   is_formula, is_number_stored_as_text, num_fmt) VALUES (?,?,?,?,?,?,?,?)""",
                [
                    (sheet_id, c.row, c.col, c.raw_value, c.cached_value,
                     int(c.is_formula), int(c.is_number_stored_as_text), c.num_fmt)
                    for c in sheet.cells
                ],
            )
    return batch_id


def _dumps(obj) -> str:
    import json

    return json.dumps(obj, ensure_ascii=False)
